"""AI connectivity and repair checks for the universal self-test."""

import json
import os

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication

from universal_core import (
    AIClient,
    FieldRule,
    SiteTemplate,
    UniversalExtractor,
    ai_parse_task,
    ai_repair_fields,
    ai_suggest_fields,
    diagnose_ai_settings,
    export_table_data,
    load_ai_call_logs,
    load_ai_settings,
    page_snapshot_from_html,
    summarize_ai_call_logs,
    test_ai_provider_connectivity,
)
from universal_ui import AIWorker


def verify_ai_workflow(window, settings, html, data_dir, self_test_stage):
    connectivity_providers = {
        "openai": {
            "enabled": True,
            "name": "OpenAI",
            "api_key": "test-key",
            "base_url": "https://example.com/v1",
            "model": "fake-model",
            "active_api_key_name": "好 Key",
            "api_keys": [{"name": "好 Key", "key": "test-key"}],
        },
        "deepseek": {
            "enabled": True,
            "name": "DeepSeek",
            "api_key": "bad-key",
            "base_url": "https://example.com/v1",
            "model": "fake-model",
            "active_api_key_name": "坏 Key",
            "api_keys": [{"name": "坏 Key", "key": "bad-key"}],
        },
        "thunderbit": {
            "enabled": True,
            "name": "Thunderbit",
            "api_key": "skip-key",
            "base_url": "https://example.com/v1",
            "model": "tb-model",
            "active_api_key_name": "跳过 Key",
            "api_keys": [{"name": "跳过 Key", "key": "skip-key"}],
        },
    }
    connectivity_seed = dict(settings)
    connectivity_seed["providers"] = connectivity_providers
    connectivity_result = test_ai_provider_connectivity(
        connectivity_seed,
        ["openai", "deepseek", "thunderbit"],
        now_text="2026-06-09 09:30:00",
    )
    connectivity_rows = connectivity_result.get("results", [])
    if not any(item.get("provider") == "openai" and item.get("status") == "成功" for item in connectivity_rows):
        raise AssertionError("批量模型测试未记录成功厂商")
    if not any(item.get("provider") == "deepseek" and item.get("status") == "失败" and "invalid key" in item.get("message", "").lower() for item in connectivity_rows):
        raise AssertionError("批量模型测试未记录失败原因")
    if not any(item.get("provider") == "thunderbit" and item.get("status") == "跳过" for item in connectivity_rows):
        raise AssertionError("批量模型测试未跳过第三方抽取接口")
    window.on_ai_result("test_provider_connectivity", connectivity_result)
    connectivity_settings = load_ai_settings()
    tested_openai = connectivity_settings.get("providers", {}).get("openai", {})
    tested_deepseek = connectivity_settings.get("providers", {}).get("deepseek", {})
    if tested_openai.get("connection_status") != "成功" or tested_openai.get("connection_tested_at") != "2026-06-09 09:30:00":
        raise AssertionError("批量模型测试成功状态未保存")
    if tested_deepseek.get("connection_status") != "失败" or "invalid key" not in tested_deepseek.get("connection_error", "").lower():
        raise AssertionError("批量模型测试失败状态未保存")
    failed_key = next((item for item in tested_deepseek.get("api_keys", []) if item.get("name") == "坏 Key"), {})
    if failed_key.get("status") != "失败" or "invalid key" not in failed_key.get("last_error", "").lower():
        raise AssertionError("批量模型测试未同步 Key 失败状态")
    window.refresh_ai_provider_overview()
    connectivity_text = "\n".join(
        window.ai_provider_overview_table.item(row, 8).text()
        for row in range(window.ai_provider_overview_table.rowCount())
        if window.ai_provider_overview_table.item(row, 8)
    )
    if "2026-06-09 09:30:00" not in connectivity_text or "invalid key" not in connectivity_text.lower():
        raise AssertionError("厂商总览未展示连接测试时间或失败原因")

    test_result = AIClient(settings).test_connection()
    if not test_result.get("ok"):
        raise AssertionError("AI API 测试失败")
    self_test_stage("ai api connection OK")
    good_diagnosis = diagnose_ai_settings(settings)
    if not good_diagnosis.get("ok"):
        raise AssertionError("AI 配置诊断误判正常配置")
    window.on_ai_result("diagnose_api", good_diagnosis)
    if window.ai_diagnosis_table.rowCount() < 6:
        raise AssertionError("AI 配置诊断结果未进入表格")

    suggested = ai_suggest_fields("https://example.com/item", html, "建议列", settings)
    if not suggested.get("fields"):
        raise AssertionError("AI 建议列失败")
    retry_settings = dict(settings)
    retry_settings.update(
        {
            "api_key": "bad-key",
            "active_api_key_name": "坏 Key",
            "api_keys": [
                {"name": "坏 Key", "key": "bad-key", "status": "失败", "last_error": "invalid key"},
                {"name": "好 Key", "key": "test-key", "status": "可用", "last_tested_at": "2026-06-09 00:00:00"},
            ],
        }
    )
    retry_worker = AIWorker(
        "suggest_fields",
        retry_settings,
        {"url": "https://example.com/item", "html": html, "goal": "建议列"},
    )
    retry_result = retry_worker.run_with_key_retry()
    if retry_result.get("_auto_switched_key") != "好 Key" or not retry_result.get("fields"):
        raise AssertionError("AI 任务未自动切换可用 Key 重试")
    retry_worker.write_call_log("成功", result=retry_result, duration_ms=123)
    retry_worker_without_key = AIWorker(
        "suggest_fields",
        {**retry_settings, "api_keys": [{"name": "坏 Key", "key": "bad-key", "status": "失败"}]},
        {"url": "https://example.com/item", "html": html, "goal": "建议列"},
    )
    try:
        retry_worker_without_key.run_with_key_retry()
        raise AssertionError("无可用 Key 时不应重试成功")
    except RuntimeError as exc:
        retry_worker_without_key.write_call_log("失败", error_text=str(exc), duration_ms=45)
        if "401" not in str(exc) and "invalid key" not in str(exc).lower():
            raise
    ai_logs = load_ai_call_logs(20)
    if not any(item.get("status") == "成功" and item.get("auto_switched_key") == "好 Key" for item in ai_logs):
        raise AssertionError("AI 调用日志未记录自动换 Key 成功")
    if not any(item.get("status") == "失败" and "invalid key" in item.get("error", "").lower() for item in ai_logs):
        raise AssertionError("AI 调用日志未记录失败原因")
    window.fill_ai_call_log_table(ai_logs)
    if window.ai_call_log_table.rowCount() < 2:
        raise AssertionError("AI 调用日志未进入 UI 表格")
    summary_rows = summarize_ai_call_logs(ai_logs)
    summary_row = next(
        (
            item for item in summary_rows
            if item.get("model") == "fake-model" and item.get("key_name") == "好 Key"
        ),
        {},
    )
    if summary_row.get("total_calls") != 1 or summary_row.get("success_count") != 1:
        raise AssertionError("AI 用量汇总未统计成功调用")
    failed_summary_row = next(
        (
            item for item in summary_rows
            if item.get("key_name") == "坏 Key" and item.get("failure_count") == 1
        ),
        {},
    )
    if "invalid key" not in failed_summary_row.get("latest_error", "").lower():
        raise AssertionError("AI 用量汇总未保留最近失败原因")
    window.fill_ai_call_summary_table(summary_rows)
    if window.ai_call_summary_table.rowCount() < 2:
        raise AssertionError("AI 用量汇总未进入 UI 表格")

    ai_log_export_xlsx = os.path.join(data_dir, "ai_call_logs_export.xlsx")
    ai_log_export_csv = os.path.join(data_dir, "ai_call_logs_export.csv")
    ai_log_export_json = os.path.join(data_dir, "ai_call_logs_export.json")
    ai_summary_export_xlsx = os.path.join(data_dir, "ai_call_summary_export.xlsx")
    ai_summary_export_csv = os.path.join(data_dir, "ai_call_summary_export.csv")
    ai_summary_export_json = os.path.join(data_dir, "ai_call_summary_export.json")
    window.export_ai_call_logs_to_file(ai_log_export_xlsx)
    window.export_ai_call_logs_to_file(ai_log_export_csv)
    window.export_ai_call_logs_to_file(ai_log_export_json)
    window.export_ai_call_summary_to_file(ai_summary_export_xlsx)
    window.export_ai_call_summary_to_file(ai_summary_export_csv)
    window.export_ai_call_summary_to_file(ai_summary_export_json)
    for exported_path in (
        ai_log_export_xlsx,
        ai_log_export_csv,
        ai_log_export_json,
        ai_summary_export_xlsx,
        ai_summary_export_csv,
        ai_summary_export_json,
    ):
        if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
            raise AssertionError(f"AI 调用日志导出失败：{exported_path}")
    with open(ai_log_export_json, "r", encoding="utf-8") as file_obj:
        exported_log_payload = json.load(file_obj)
    if not exported_log_payload.get("rows"):
        raise AssertionError("AI 调用日志 JSON 导出内容为空")
    with open(ai_summary_export_json, "r", encoding="utf-8") as file_obj:
        exported_summary_payload = json.load(file_obj)
    if not any(row.get("Key", "").startswith("好 Key") for row in exported_summary_payload.get("rows", [])):
        raise AssertionError("AI 用量汇总 JSON 导出内容错误")
    window.clear_ai_call_logs_and_refresh()
    if load_ai_call_logs(20) or window.ai_call_log_table.rowCount() != 0 or window.ai_call_summary_table.rowCount() != 0:
        raise AssertionError("AI 调用日志清空失败")

    task = ai_parse_task("按自然语言生成采集任务 actions", page_snapshot_from_html("https://example.com/item", html), settings)
    if not task.get("actions"):
        raise AssertionError("AI Agent 动作规划失败")
    window.latest_ai_result = task
    window.apply_ai_task(task)
    if window.ai_task_plan_table.rowCount() < 3 or "AI 自检模板" not in window.ai_task_plan_label.text():
        raise AssertionError("自然语言任务计划未进入预览表")
    plan_text = "\n".join(
        window.ai_task_plan_table.item(row, column).text()
        for row in range(window.ai_task_plan_table.rowCount())
        for column in range(window.ai_task_plan_table.columnCount())
        if window.ai_task_plan_table.item(row, column)
    )
    if "extract" not in plan_text or "标题" not in plan_text:
        raise AssertionError("自然语言任务计划预览缺少动作或字段")
    window.copy_current_ai_task_plan()
    copied_plan_text = QApplication.clipboard().text() or getattr(window, "last_clipboard_text", "")
    if "AI 自检模板" not in copied_plan_text or "actions" not in copied_plan_text:
        raise AssertionError("自然语言任务计划复制失败")
    if not window.apply_current_ai_task_plan():
        raise AssertionError("自然语言任务计划未应用成功")
    if window.template_combo.currentText() != "AI 自检模板" or window.page_limit_input.value() != 2:
        raise AssertionError("自然语言任务计划未写入模板或采集参数")
    if window.tabs.tabText(window.tabs.currentIndex()) != "一键采集":
        raise AssertionError("自然语言任务计划应用后不应离开统一普通界面")

    window.apply_ai_fields(suggested)
    if window.ai_suggest_table.rowCount() < 2:
        raise AssertionError("AI 建议列未进入确认表格")
    second_enable = window.ai_suggest_table.item(1, 0)
    second_enable.setCheckState(Qt.CheckState.Unchecked)
    window.apply_checked_ai_fields_to_template()
    if window.field_table.rowCount() != 1:
        raise AssertionError("AI 建议列确认应用失败")
    if window.field_table.item(0, 0).text() != "标题":
        raise AssertionError("AI 建议列编辑结果未写入模板 UI")
    window.show_ai_suggested_fields(suggested.get("fields", []))
    preview_record = UniversalExtractor(
        SiteTemplate(
            "AI 预采自检模板",
            field_rules=window.suggested_field_rules_from_table(),
        )
    ).extract(html, "https://example.com/item")
    window.show_preview_record(preview_record, window.suggested_field_rules_from_table())
    if window.ai_table.columnCount() < 3 or window.ai_table.rowCount() != 1:
        raise AssertionError("AI 预采表格未生成")
    if "测试商品 A" not in (window.ai_table.item(0, 1).text() if window.ai_table.item(0, 1) else ""):
        raise AssertionError("AI 预采字段值错误")
    if window.ai_quality_table.rowCount() < 2:
        raise AssertionError("AI 预采质量检测未生成")
    if window.ai_quality_table.columnCount() != 6 or "字段质量评分" not in window.ai_quality_score_label.text():
        raise AssertionError("AI 字段质量评分未显示")
    table_columns, table_rows = window.ai_table_data()
    if "标题" not in table_columns or not table_rows:
        raise AssertionError("AI 表格数据读取失败")
    ai_export_xlsx = os.path.join(data_dir, "ai_table_export.xlsx")
    ai_export_csv = os.path.join(data_dir, "ai_table_export.csv")
    ai_export_json = os.path.join(data_dir, "ai_table_export.json")
    export_table_data(ai_export_xlsx, table_columns, table_rows, sheet_name="AI表格结果")
    export_table_data(ai_export_csv, table_columns, table_rows, sheet_name="AI表格结果")
    export_table_data(ai_export_json, table_columns, table_rows, sheet_name="AI表格结果")
    for exported_path in (ai_export_xlsx, ai_export_csv, ai_export_json):
        if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
            raise AssertionError(f"AI 表格导出失败：{exported_path}")
    from openpyxl import load_workbook

    ai_workbook = load_workbook(ai_export_xlsx)
    if ai_workbook.sheetnames[:3] != ["AI表格结果", "字段说明", "导出摘要"]:
        raise AssertionError("AI 表格 Excel 未生成工作簿结构")
    if ai_workbook["AI表格结果"].freeze_panes != "A2" or not ai_workbook["AI表格结果"].auto_filter.ref:
        raise AssertionError("AI 表格 Excel 未冻结表头或启用筛选")
    if ai_workbook["导出摘要"].cell(5, 2).value != len(table_rows):
        raise AssertionError("AI 表格 Excel 摘要行数错误")
    with open(ai_export_json, "r", encoding="utf-8") as file_obj:
        exported_json = json.load(file_obj)
    if not exported_json.get("rows") or "测试商品 A" not in json.dumps(exported_json, ensure_ascii=False):
        raise AssertionError("AI 表格 JSON 导出内容错误")
    window.fill_ai_table(["字段A", "字段B"], [["含\t制表符", "含\n换行"]])
    window.copy_ai_table_to_clipboard()
    clipboard_text = QApplication.clipboard().text() or getattr(window, "last_clipboard_text", "")
    if "含 制表符" not in clipboard_text or "含 换行" not in clipboard_text or "\t制表符" in clipboard_text:
        raise AssertionError("AI 表格复制到剪贴板失败")
    window.fill_ai_table(table_columns, table_rows)
    quality_issues = window.analyze_preview_quality(
        [
            FieldRule("空字段", ".not-exists"),
            FieldRule("重复一", "h1"),
            FieldRule("重复二", "h1"),
        ],
        {"空字段": "", "重复一": "同一个值", "重复二": "同一个值"},
    )
    problems = [issue.get("problem", "") for issue in quality_issues]
    if not any("空值" in problem for problem in problems):
        raise AssertionError("AI 质量检测未识别空值列")
    if not any("重复" in problem for problem in problems):
        raise AssertionError("AI 质量检测未识别重复列")
    summary = window.quality_summary(quality_issues)
    if summary.get("score", 100) >= 80 or summary.get("need_fix") != 1 or summary.get("need_confirm") < 1:
        raise AssertionError("AI 字段质量评分未识别低质量字段")
    window.fill_quality_table(quality_issues)
    if "需要修复" not in window.ai_quality_score_label.text():
        raise AssertionError("AI 字段质量评分未提示修复")
    if int(window.ai_quality_table.item(0, 1).text()) >= 50:
        raise AssertionError("AI 字段质量评分未给空字段低分")
    repaired = ai_repair_fields(
        "https://example.com/item",
        html,
        [FieldRule("标题", "h1"), FieldRule("价格", ".missing")],
        quality_issues,
        "修复字段",
        settings,
    )
    if not repaired.get("fields") or repaired["fields"][0].get("selector") != "h1#main-title":
        raise AssertionError("AI 问题列修复失败")
    window.apply_repaired_fields(repaired)
    if window.ai_suggest_table.item(0, 2).text() != "h1#main-title":
        raise AssertionError("AI 修复字段未回填建议表")
    if not window.apply_repaired_fields_to_template() or window.field_table.item(0, 1).text() != "h1#main-title":
        raise AssertionError("AI 修复字段未一键应用到模板")
    self_test_stage("ai repair fields OK")
