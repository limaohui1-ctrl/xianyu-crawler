"""
export_excel.py — export harvest results to Excel (.xlsx).

Uses openpyxl. Falls back gracefully if not installed.
"""

import json
import os
import time
from typing import List, Optional

EXCEL_FIELDS = [
    ("序号", "index"),
    ("标题", "title"),
    ("URL", "url"),
    ("来源网站", "source_domain"),
    ("资料类型", "doc_type"),
    ("正文摘要", "summary"),
    ("关键词命中", "keyword_hits"),
    ("质量评分", "quality_score"),
    ("质量状态", "quality_status"),
    ("采集状态", "status"),
    ("失败原因", "error"),
    ("是否重复", "is_duplicate"),
    ("重复原因", "duplicate_reason"),
    ("采集时间", "harvest_time"),
    ("发布者", "author"),
    ("发布时间", "publish_time"),
    ("正文长度", "text_length"),
]


def export_excel(articles: List[dict],
                 output_path: str = "",
                 include_duplicates: bool = True) -> str:
    """
    Export articles to Excel file.

    Args:
        articles: List of ContentRecord dicts.
        output_path: Output .xlsx path. Auto-generated if empty.
        include_duplicates: Whether to include duplicates.

    Returns:
        Absolute path to the generated .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Run: pip install openpyxl"
        )

    if not output_path:
        os.makedirs("acs_data/harvest", exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join("acs_data/harvest", f"harvest_{timestamp}.xlsx")

    # Filter duplicates if requested
    if not include_duplicates:
        articles = [a for a in articles if not a.get("is_duplicate")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采集结果"

    # ── Styles ──
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Quality color fills
    quality_fills = {
        "高质量": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "可用": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "需复核": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "失败": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    }

    # ── Write headers ──
    headers = [h[0] for h in EXCEL_FIELDS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Write data ──
    for row_idx, article in enumerate(articles, 2):
        for col_idx, (header, key) in enumerate(EXCEL_FIELDS, 1):
            value = _format_value(article, key, row_idx - 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border

        # Color by quality status
        quality = article.get("quality_status", "")
        if quality in quality_fills:
            for col_idx in range(1, len(EXCEL_FIELDS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = quality_fills[quality]

    # ── Column widths ──
    col_widths = {
        1: 6,    # 序号
        2: 35,   # 标题
        3: 45,   # URL
        4: 18,   # 来源网站
        5: 10,   # 资料类型
        6: 45,   # 正文摘要
        7: 10,   # 关键词命中
        8: 10,   # 质量评分
        9: 10,   # 质量状态
        10: 10,  # 采集状态
        11: 25,  # 失败原因
        12: 10,  # 是否重复
        13: 20,  # 重复原因
        14: 20,  # 采集时间
        15: 15,  # 发布者
        16: 14,  # 发布时间
        17: 10,  # 正文长度
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_FIELDS))}{len(articles) + 1}"

    wb.save(output_path)
    return os.path.abspath(output_path)


def _format_value(article: dict, key: str, idx: int) -> str:
    """Format a single field value for Excel."""
    if key == "index":
        return idx + 1
    if key == "is_duplicate":
        return "是" if article.get("is_duplicate") else ""
    if key == "harvest_time":
        return article.get("harvest_time") or time.strftime("%Y-%m-%d %H:%M:%S")
    val = article.get(key, "")
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)[:500]
    return str(val)[:1000] if val else ""


def is_excel_available() -> bool:
    """Check if openpyxl is installed."""
    try:
        import openpyxl
        return True
    except ImportError:
        return False
