"""Queue workflow checks for the universal self-test."""

import os

from PyQt6.QtWidgets import QApplication
from universal_core import export_records, records_to_tsv


def verify_queue_workflow(window, db, data_dir, run_id, run_config, run_risks, linked_records, record, self_test_stage):
    run_results_xlsx = os.path.join(data_dir, "run_results_export.xlsx")
    run_results_csv = os.path.join(data_dir, "run_results_export.csv")
    export_records(run_results_xlsx, db.records_for_run(run_id))
    export_records(run_results_csv, db.records_for_run(run_id))
    for exported_path in (run_results_xlsx, run_results_csv):
        if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
            raise AssertionError(f"当前任务结果导出失败：{exported_path}")
    from openpyxl import load_workbook

    workbook = load_workbook(run_results_xlsx)
    if workbook.sheetnames[:3] != ["采集结果", "字段说明", "导出摘要"]:
        raise AssertionError("Excel 工作簿未按采集结果/字段说明/导出摘要组织")
    result_sheet = workbook["采集结果"]
    if result_sheet.freeze_panes != "A2" or not result_sheet.auto_filter.ref:
        raise AssertionError("Excel 工作簿未冻结表头或启用筛选")
    if workbook["字段说明"].cell(1, 1).value != "字段" or workbook["导出摘要"].cell(3, 2).value != 1:
        raise AssertionError("Excel 工作簿字段说明或摘要内容错误")
    with open(run_results_csv, "r", encoding="utf-8-sig") as file_obj:
        if "任务关联结果 A" not in file_obj.read():
            raise AssertionError("当前任务结果 CSV 导出内容错误")
    sheets_text = records_to_tsv(db.records_for_run(run_id))
    if not sheets_text.startswith("采集时间\t网址\t域名") or "任务关联结果 A" not in sheets_text:
        raise AssertionError("Sheets 复制文本格式错误")
    window.records = linked_records
    if not window.copy_records_to_sheets(window.records, "自检本次结果"):
        raise AssertionError("复制本次结果到 Sheets 失败")
    if "采集时间\t网址\t域名" not in window.last_clipboard_text or "任务关联结果 A" not in window.last_clipboard_text:
        raise AssertionError("复制到 Sheets 未写入剪贴板文本")
    if not window.copy_current_results_to_sheets():
        raise AssertionError("复制当前结果到 Sheets 失败")
    if "已复制" not in window.result_export_hint_label.text():
        raise AssertionError("复制当前结果后导出引导未反馈成功")

    window.update_collect_progress(
        {
            "processed": 1,
            "success": 1,
            "failed": 0,
            "total": 2,
            "current_url": "https://example.com/item",
            "status": "running",
        }
    )
    if window.collect_progress_bar.maximum() != 2 or window.collect_progress_bar.value() != 1:
        raise AssertionError("采集进度条未按任务进度更新")
    if "已处理 1/2" not in window.collect_progress_label.text() or "https://example.com/item" not in window.collect_progress_label.text():
        raise AssertionError("采集进度文字未显示处理数量和当前网址")
    window.update_collect_progress(
        {
            "processed": 1,
            "success": 1,
            "failed": 0,
            "total": 2,
            "current_url": "https://example.com/item",
            "stage": "采集页面",
            "status": "running",
        }
    )
    if "阶段：采集页面" not in window.collect_progress_label.text():
        raise AssertionError("采集进度文字未显示内核阶段")
    window.url_input.setPlainText("https://example.com/item\nhttps://example.com/list")
    window.page_limit_input.setValue(3)
    window.subpage_checkbox.setChecked(True)
    window.selected_subpage_urls = ["https://example.com/detail/1"]
    estimated_queue = window.estimated_task_queue(window.urls_from_input())
    if len(estimated_queue) != 8:
        raise AssertionError("任务预估队列数量错误")
    window.fill_task_queue_table(estimated_queue)
    if window.task_queue_table.rowCount() != 8:
        raise AssertionError("任务预估队列表未显示")
    window.update_task_queue_progress({"stage": "采集页面", "current_url": "https://example.com/item"})
    if window.task_queue_table.item(0, 0).text() != "运行中" or window.task_queue_table.item(0, 2).text() != "采集页面":
        raise AssertionError("运行队列未显示当前采集阶段")
    window.update_task_queue_progress({"stage": "页面完成", "current_url": "https://example.com/item"})
    if window.task_queue_table.item(0, 0).text() != "已完成":
        raise AssertionError("运行队列未显示完成状态")
    self_test_stage("queue workflow OK")

    queue_result_record = dict(record)
    queue_result_record["url"] = "https://example.com/item"
    queue_result_record["title"] = "队列关联结果"
    queue_result_record["error"] = ""
    window.add_record(queue_result_record)
    if not any(source.get("url") == "https://example.com/item" and int(source.get("result_count") or 0) >= 1 for source in window.task_queue_rows):
        raise AssertionError("当前队列未关联结果数量")
    window.task_queue_status_filter.setCurrentText("全部状态")
    window.task_queue_table.selectRow(0)
    window.view_selected_queue_result()
    selected_current = window.selected_record_from_table(window.result_table)
    if not selected_current or selected_current.get("url") != "https://example.com/item":
        raise AssertionError("当前队列查看结果未定位结果表")

    queue_error_record = dict(record)
    queue_error_record["url"] = "https://example.com/list"
    queue_error_record["title"] = ""
    queue_error_record["error"] = "Timeout 30000ms exceeded"
    window.add_record(queue_error_record)
    failed_source = next((source for source in window.task_queue_rows if source.get("url") == "https://example.com/list"), {})
    if failed_source.get("error_category") != "网络超时" or "访问间隔" not in failed_source.get("error_advice", ""):
        raise AssertionError("当前队列未显示错误分类和修复建议")
    window.update_task_queue_progress({"stage": "页面完成", "current_url": "https://example.com/list", "failed_item": True})
    window.task_queue_status_filter.setCurrentText("失败")
    if window.task_queue_table.rowCount() != 1 or window.task_queue_table.item(0, 3).text() != "https://example.com/list":
        raise AssertionError("运行队列失败筛选未生效")
    window.task_queue_table.selectRow(0)
    window.update_queue_detail_panel()
    if "网络超时" not in window.queue_detail_output.toPlainText() or "访问间隔" not in window.queue_detail_output.toPlainText():
        raise AssertionError("运行队列详情未显示失败原因和建议")
    if "失败" not in window.queue_summary_label.text():
        raise AssertionError("运行队列摘要未显示失败数量")
    recovery_text = window.failure_recovery_label.text()
    for expected in ("失败自恢复", "网络超时", "重试失败项", "启用真实浏览器", "调低速度"):
        if expected not in recovery_text:
            raise AssertionError(f"失败自恢复面板缺失：{expected}")
    window.use_browser_checkbox.setChecked(False)
    if not window.enable_browser_recovery() or not window.use_browser_checkbox.isChecked():
        raise AssertionError("失败自恢复未启用真实浏览器")
    window.delay_input.setValue(0)
    if not window.slow_down_recovery() or window.delay_input.value() < 3:
        raise AssertionError("失败自恢复未调低访问速度")
    window.copy_selected_queue_error()
    clipboard_text = getattr(window, "last_clipboard_text", "") or QApplication.clipboard().text()
    if "https://example.com/list" not in clipboard_text or "Timeout 30000ms exceeded" not in clipboard_text:
        raise AssertionError("运行队列错误复制内容不完整")
    selected_retry_calls = []
    original_start_collecting_for_selected_queue = window.start_collecting
    window.start_collecting = lambda: selected_retry_calls.append(window.urls_from_input())
    window.retry_selected_queue_item()
    window.start_collecting = original_start_collecting_for_selected_queue
    if selected_retry_calls != [["https://example.com/list"]]:
        raise AssertionError("选中队列项重试未回填并触发采集")
    window.task_queue_status_filter.setCurrentText("未完成")
    if window.task_queue_table.rowCount() < 1:
        raise AssertionError("运行队列未完成筛选未生效")
    retry_calls = []
    original_start_collecting_for_queue = window.start_collecting
    window.start_collecting = lambda: retry_calls.append(window.urls_from_input())
    window.retry_incomplete_queue_items()
    window.start_collecting = original_start_collecting_for_queue
    if retry_calls != [["https://example.com/list"]]:
        raise AssertionError("失败/未完成队列重试未回填并触发采集")
    window.task_queue_status_filter.setCurrentText("全部状态")

    queue_run_id = db.start_run(run_config, run_risks)
    window.current_run_id = queue_run_id
    window.persist_current_run_queue_snapshot()
    queue_archive_record = dict(record)
    queue_archive_record["url"] = "https://example.com/item"
    queue_archive_record["title"] = "队列档案关联结果"
    queue_archive_record["fingerprint"] = ""
    queue_archive_record["run_id"] = queue_run_id
    db.save_record(queue_archive_record, skip_unchanged=False)
    queue_run = next((item for item in db.recent_runs(20) if item.get("id") == queue_run_id), None)
    queue_snapshot = (queue_run.get("config") or {}).get("task_queue_snapshot") if queue_run else []
    if not queue_snapshot or not any(item.get("status") == "失败" and item.get("url") == "https://example.com/list" for item in queue_snapshot):
        raise AssertionError("任务队列快照未写入任务档案")
    return {"queue_run_id": queue_run_id}
