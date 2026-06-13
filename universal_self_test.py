"""Noninteractive regression self-test for the universal collector UI."""

import gc
import json
import os
import sys
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication, QFileDialog, QLineEdit, QMessageBox, QProgressBar

from core_firecrawl import (
    FirecrawlClient,
    FirecrawlConfig,
    firecrawl_document_to_record,
    merge_firecrawl_extract_record,
    merge_firecrawl_interact_record,
)
from universal_core import (
    AI_CALL_LOG_FILE,
    AI_PROVIDER_PRESETS,
    AI_SETTINGS_FILE,
    AIClient,
    CHANGE_ALERT_STATE_FILE,
    CollectorDatabase,
    DB_FILE,
    DEFAULT_PAGE_LIMIT,
    FIELD_HEADERS,
    FieldRule,
    RISK_CONFIRMATION_FILE,
    SCHEDULE_FILE,
    SiteTemplate,
    TEMPLATE_FILE,
    UniversalCollector,
    UniversalExtractor,
    ai_parse_task,
    ai_provider_preset_health,
    ai_provider_runtime_overview,
    ai_repair_fields,
    ai_suggest_fields,
    analyze_collect_task,
    append_ai_repair_history,
    classify_error,
    cleanup_user_data,
    diagnose_ai_settings,
    download_images_from_records,
    ensure_runtime_dirs,
    export_records,
    export_table_data,
    extract_emails_and_phones,
    load_ai_call_logs,
    load_ai_repair_history,
    load_ai_settings,
    load_change_alert_states,
    load_risk_confirmations,
    load_schedules,
    normalize_url,
    page_snapshot_from_html,
    records_to_tsv,
    refresh_ai_provider_models,
    runtime_self_test_error_log_file,
    runtime_startup_log_file,
    runtime_diagnostic_log_file,
    save_ai_settings,
    scene_template_presets,
    summarize_ai_call_logs,
    test_ai_provider_connectivity,
)
from universal_ui import AIWorker, UniversalMainWindow
from universal_self_test_runtime import prepare_self_test_runtime
from universal_self_test_launchers import verify_home_tabs, verify_launcher_layout, verify_simple_panel
from universal_self_test_ai_history import verify_repair_history_flow
from universal_self_test_ui_smoke import verify_collect_wizard_smoke
from universal_self_test_results import verify_result_workflow
from universal_self_test_queue import verify_queue_workflow
from universal_self_test_ai_flow import verify_ai_workflow

def run_universal_self_test():
    runtime = prepare_self_test_runtime(
        (
            ("db_file", "UNIVERSAL_COLLECTOR_DB_FILE", DB_FILE),
            ("template_file", "UNIVERSAL_COLLECTOR_TEMPLATE_FILE", TEMPLATE_FILE),
            ("ai_settings_file", "UNIVERSAL_COLLECTOR_AI_SETTINGS_FILE", AI_SETTINGS_FILE),
            ("ai_call_log_file", "UNIVERSAL_COLLECTOR_AI_CALL_LOG_FILE", AI_CALL_LOG_FILE),
            ("schedule_file", "UNIVERSAL_COLLECTOR_SCHEDULE_FILE", SCHEDULE_FILE),
            ("change_alert_state_file", "UNIVERSAL_COLLECTOR_CHANGE_ALERT_STATE_FILE", CHANGE_ALERT_STATE_FILE),
            ("risk_confirmation_file", "UNIVERSAL_COLLECTOR_RISK_CONFIRMATION_FILE", RISK_CONFIRMATION_FILE),
        )
    )
    safe_root = runtime["safe_root"]
    data_dir = runtime["data_dir"]
    self_test_stage = runtime["self_test_stage"]
    db_file = runtime["db_file"]
    template_file = runtime["template_file"]
    ai_settings_file = runtime["ai_settings_file"]
    ai_call_log_file = runtime["ai_call_log_file"]
    schedule_file = runtime["schedule_file"]
    change_alert_state_file = runtime["change_alert_state_file"]
    risk_confirmation_file = runtime["risk_confirmation_file"]
    normalize_cases = [
        ("example.com/path#frag", "", "https://example.com/path"),
        ("https://EXAMPLE.com:443/path/?utm_source=x&id=1#frag", "", "https://example.com/path?id=1"),
        ("//example.com/a?fbclid=abc&b=2", "", "https://example.com/a?b=2"),
        ("/detail?id=1&utm_medium=y#tab", "https://example.com/base/", "https://example.com/detail?id=1"),
        ("javascript:void(0)", "", ""),
    ]
    for raw_url, base_url, expected_url in normalize_cases:
        normalized_url = normalize_url(raw_url, base_url)
        if normalized_url != expected_url:
            raise AssertionError(f"URL 归一化不符合预期：{raw_url} + {base_url} => {normalized_url} != {expected_url}")
    firecrawl_calls = []

    def fake_firecrawl_transport(endpoint, payload, headers, timeout):
        firecrawl_calls.append({"endpoint": endpoint, "payload": dict(payload), "headers": dict(headers), "timeout": timeout})
        if endpoint == "/v2/scrape":
            return {
                "success": True,
                "id": "scrape-job-1",
                "data": {
                    "markdown": "# Firecrawl 标题\n\n这是 Firecrawl 抽取出的正文资料，包含商品说明和页面重点。",
                    "html": "<html><body><a href='/detail'>详情</a><img src='/a.png' alt='图'></body></html>",
                    "links": [{"url": "/next", "text": "下一页"}],
                    "metadata": {
                        "title": "Firecrawl 元标题",
                        "sourceURL": "https://example.com/item?utm_source=x&id=1",
                        "author": "Firecrawl 作者",
                        "ogImage": "/cover.jpg",
                    },
                },
            }
        if endpoint == "/v2/parse":
            return {
                "success": True,
                "data": {
                    "markdown": "# Parse 标题\n\nFirecrawl Parse 文件正文",
                    "metadata": {"title": "Parse 标题", "sourceURL": "https://example.com/parse-source"},
                    "links": [{"url": "https://example.com/parse-link", "text": "文件链接"}],
                },
            }
        if endpoint == "/v2/scrape/scrape-job-1/interact":
            return {
                "success": True,
                "output": "已点击展开更多，得到隐藏内容",
                "exitCode": 0,
                "interactiveLiveViewUrl": "https://live.example.test/session",
            }
        if endpoint == "/v2/map":
            return {"success": True, "links": ["https://example.com/a?utm_source=x", {"url": "/b", "title": "B"}]}
        if endpoint == "/v2/search":
            return {
                "success": True,
                "data": {
                    "web": [
                        {"url": "https://example.com/search-a?utm_campaign=x", "title": "搜索 A", "description": "A 摘要"},
                        {"link": "https://example.com/search-b", "title": "搜索 B"},
                    ],
                    "images": [{"imageUrl": "https://example.com/image-result.jpg", "alt": "图片结果"}],
                },
            }
        if endpoint == "/v2/extract":
            return {"success": True, "id": "extract-job-1", "status": "processing"}
        if endpoint == "/v2/extract/extract-job-1":
            return {
                "success": True,
                "status": "completed",
                "data": {
                    "title": "抽取标题",
                    "price": "199",
                    "author": "抽取作者",
                    "published_time": "2026-06-10",
                    "specs": {"颜色": "黑色", "库存": 3},
                },
                "sources": [{"url": "https://example.com/item"}],
            }
        if endpoint == "/v2/batch/scrape":
            return {"success": True, "id": "batch-job-1", "status": "scraping"}
        if endpoint == "/v2/batch/scrape/batch-job-1":
            return {
                "success": True,
                "status": "completed",
                "completed": 2,
                "total": 2,
                "data": [
                    {
                        "url": "https://example.com/batch-a",
                        "markdown": "# Batch A\n\n批量抓取正文 A",
                        "metadata": {"title": "Batch A"},
                    },
                    {
                        "url": "https://example.com/batch-b",
                        "markdown": "# Batch B\n\n批量抓取正文 B",
                        "metadata": {"title": "Batch B"},
                    },
                ],
            }
        if endpoint == "/v2/crawl":
            return {"success": True, "id": "crawl-job-1", "status": "scraping"}
        if endpoint == "/v2/crawl/crawl-job-1":
            return {
                "success": True,
                "status": "completed",
                "completed": 2,
                "total": 2,
                "data": [
                    {
                        "url": "https://example.com/crawl-root",
                        "markdown": "# Crawl Root\n\n深度采集首页",
                        "metadata": {"title": "Crawl Root"},
                    },
                    {
                        "url": "https://example.com/crawl-detail",
                        "markdown": "# Crawl Detail\n\n深度采集详情",
                        "metadata": {"title": "Crawl Detail"},
                    },
                ],
            }
        raise AssertionError(f"未知 Firecrawl endpoint：{endpoint}")

    firecrawl_config = FirecrawlConfig.from_dict(
        {
            "enabled": True,
            "api_key": "fc-self-test",
            "base_url": "https://api.firecrawl.dev",
            "formats": "markdown,html,links",
            "map_limit": 2,
            "timeout_seconds": 12,
        }
    )
    firecrawl_client = FirecrawlClient(firecrawl_config, transport=fake_firecrawl_transport)
    if FirecrawlClient(FirecrawlConfig.from_dict({"enabled": True, "base_url": "http://127.0.0.1:3002/v2"})).endpoint_url("/v2/scrape") != "http://127.0.0.1:3002/v2/scrape":
        raise AssertionError("Firecrawl 自托管 /v2 地址不应重复拼接版本路径")
    firecrawl_document = firecrawl_client.scrape("https://example.com/item")
    firecrawl_record = firecrawl_document_to_record(firecrawl_document, "https://example.com/item", template_name="自检模板")
    if firecrawl_calls[0]["endpoint"] != "/v2/scrape" or firecrawl_calls[0]["headers"].get("Authorization") != "Bearer fc-self-test":
        raise AssertionError(f"Firecrawl scrape 请求未携带正确 endpoint/header：{firecrawl_calls[0]}")
    if firecrawl_calls[0]["payload"].get("origin") != "universal-web-collector":
        raise AssertionError("Firecrawl 请求未标记来源")
    if firecrawl_record.get("title") != "Firecrawl 元标题" or "Firecrawl 抽取" not in firecrawl_record.get("body", ""):
        raise AssertionError(f"Firecrawl 文档未映射为采集记录：{firecrawl_record}")
    if firecrawl_record.get("url") != "https://example.com/item?id=1" or firecrawl_record.get("template_name") != "自检模板 + Firecrawl":
        raise AssertionError(f"Firecrawl URL/模板映射错误：{firecrawl_record}")
    if not any(item.get("url") == "https://example.com/next" for item in firecrawl_record.get("links", [])):
        raise AssertionError(f"Firecrawl 链接未归一化：{firecrawl_record.get('links')}")
    if not any(item.get("url") == "https://example.com/cover.jpg" for item in firecrawl_record.get("images", [])):
        raise AssertionError(f"Firecrawl 图片未归一化：{firecrawl_record.get('images')}")
    firecrawl_map_links = firecrawl_client.map("https://example.com/")
    if firecrawl_map_links[0] != "https://example.com/a?utm_source=x" or firecrawl_calls[-1]["payload"].get("ignoreQueryParameters") is not True:
        raise AssertionError(f"Firecrawl Map 请求/响应不符合预期：{firecrawl_calls[-1]} / {firecrawl_map_links}")
    if "fc-self-test" in json.dumps(firecrawl_config.safe_dict(), ensure_ascii=False):
        raise AssertionError("Firecrawl 安全配置摘要不应包含明文 API Key")
    firecrawl_search_results = firecrawl_client.search("测试搜索", limit=3, sources=["web", "images"])
    search_urls = [item.get("url") for item in firecrawl_search_results]
    if "https://example.com/search-a" not in search_urls or "https://example.com/image-result.jpg" not in search_urls:
        raise AssertionError(f"Firecrawl Search 结果未归一化：{firecrawl_search_results}")
    if firecrawl_calls[-1]["payload"].get("query") != "测试搜索" or firecrawl_calls[-1]["payload"].get("sources") != ["web", "images"]:
        raise AssertionError(f"Firecrawl Search 请求参数错误：{firecrawl_calls[-1]}")
    extract_payload = firecrawl_client.extract(["https://example.com/item"])
    merged_extract_record = merge_firecrawl_extract_record(
        {
            "title": "",
            "price": "",
            "published_time": "",
            "author": "",
            "body": "原始正文",
            "tables": [],
        },
        extract_payload,
    )
    if merged_extract_record.get("title") != "抽取标题" or merged_extract_record.get("price") != "199":
        raise AssertionError(f"Firecrawl Extract 结果未补齐基础字段：{merged_extract_record}")
    if "Firecrawl 结构化抽取" not in merged_extract_record.get("body", "") or not merged_extract_record.get("tables"):
        raise AssertionError(f"Firecrawl Extract 结果未进入正文/表格：{merged_extract_record}")
    parse_file_path = os.path.join(data_dir, "firecrawl_parse_self_test.txt")
    os.makedirs(os.path.dirname(parse_file_path), exist_ok=True)
    with open(parse_file_path, "w", encoding="utf-8") as f:
        f.write("Firecrawl Parse 自检文件")
    parse_table = firecrawl_client.parse_file_to_table(parse_file_path, instruction="整理文件")
    if parse_table.get("source") != "firecrawl_parse" or not any("Parse 标题" in str(row) for row in parse_table.get("rows", [])):
        raise AssertionError(f"Firecrawl Parse 文件解析未输出表格：{parse_table}")
    interact_payload = firecrawl_client.interact("scrape-job-1", prompt="点击展开更多")
    merged_interact_record = merge_firecrawl_interact_record({"body": "原始正文", "tables": []}, interact_payload)
    if "隐藏内容" not in merged_interact_record.get("body", "") or not merged_interact_record.get("tables"):
        raise AssertionError(f"Firecrawl Interact 结果未合并到记录：{merged_interact_record}")
    batch_payload = firecrawl_client.batch_scrape(["https://example.com/batch-a", "https://example.com/batch-b"])
    if batch_payload.get("status") != "completed" or len(batch_payload.get("data", [])) != 2:
        raise AssertionError(f"Firecrawl Batch 轮询结果错误：{batch_payload}")
    batch_start_call = next((item for item in firecrawl_calls if item.get("endpoint") == "/v2/batch/scrape"), {})
    if batch_start_call.get("payload", {}).get("maxConcurrency") != firecrawl_config.batch_max_concurrency:
        raise AssertionError(f"Firecrawl Batch 请求未携带并发配置：{batch_start_call}")
    crawl_payload = firecrawl_client.crawl("https://example.com/crawl-root")
    if crawl_payload.get("status") != "completed" or len(crawl_payload.get("data", [])) != 2:
        raise AssertionError(f"Firecrawl Crawl 轮询结果错误：{crawl_payload}")
    crawl_start_call = next((item for item in firecrawl_calls if item.get("endpoint") == "/v2/crawl"), {})
    if crawl_start_call.get("payload", {}).get("limit") != firecrawl_config.crawl_limit or crawl_start_call.get("payload", {}).get("maxDiscoveryDepth") != firecrawl_config.crawl_max_depth:
        raise AssertionError(f"Firecrawl Crawl 请求未携带页数/深度配置：{crawl_start_call}")
    ensure_runtime_dirs()
    with open(schedule_file, "w", encoding="utf-8") as f:
        f.write("{bad schedule json")
    if load_schedules() != []:
        raise AssertionError("损坏任务计划应降级为空列表")
    diagnostic_file = runtime_diagnostic_log_file()
    if not os.path.exists(diagnostic_file):
        raise AssertionError("可恢复异常未写入诊断日志")
    with open(diagnostic_file, "r", encoding="utf-8") as f:
        diagnostic_text = f.read()
    if "读取任务计划失败" not in diagnostic_text or "JSONDecodeError" not in diagnostic_text:
        raise AssertionError(f"损坏任务计划诊断日志缺少关键信息：{diagnostic_text}")
    if os.path.exists(schedule_file):
        os.remove(schedule_file)
    app = QApplication.instance() or QApplication(sys.argv)
    window = UniversalMainWindow()
    self_test_stage("window ready")
    verify_launcher_layout(window)
    verify_home_tabs(window)
    verify_simple_panel(window)
    for attr_name in (
        "firecrawl_enabled_checkbox",
        "firecrawl_api_key_input",
        "firecrawl_base_url_input",
        "firecrawl_map_checkbox",
        "firecrawl_search_checkbox",
        "firecrawl_search_query_input",
        "firecrawl_search_limit_input",
        "firecrawl_extract_checkbox",
        "firecrawl_extract_prompt_input",
        "firecrawl_batch_checkbox",
        "firecrawl_batch_concurrency_input",
        "firecrawl_crawl_checkbox",
        "firecrawl_crawl_limit_input",
        "firecrawl_crawl_depth_input",
        "firecrawl_parse_checkbox",
        "firecrawl_interact_checkbox",
        "firecrawl_interact_wait_input",
        "firecrawl_interact_prompt_input",
    ):
        if not hasattr(window, attr_name):
            raise AssertionError(f"Firecrawl 配置控件缺失：{attr_name}")
    window.firecrawl_enabled_checkbox.setChecked(True)
    window.firecrawl_api_key_input.setText("fc-ui-self-test")
    window.firecrawl_base_url_input.setText("https://api.firecrawl.dev")
    window.firecrawl_map_checkbox.setChecked(True)
    window.firecrawl_search_checkbox.setChecked(True)
    window.firecrawl_search_query_input.setText("智能采集 自检")
    window.firecrawl_search_limit_input.setValue(7)
    window.firecrawl_extract_checkbox.setChecked(True)
    window.firecrawl_extract_prompt_input.setText("提取标题、价格、作者、时间和规格")
    window.firecrawl_batch_checkbox.setChecked(True)
    window.firecrawl_batch_concurrency_input.setValue(6)
    window.firecrawl_crawl_checkbox.setChecked(True)
    window.firecrawl_crawl_limit_input.setValue(12)
    window.firecrawl_crawl_depth_input.setValue(3)
    window.firecrawl_parse_checkbox.setChecked(True)
    window.firecrawl_interact_checkbox.setChecked(True)
    window.firecrawl_interact_wait_input.setValue(1500)
    window.firecrawl_interact_prompt_input.setText("点击展开更多")
    ui_firecrawl_config = window.current_firecrawl_config(include_secret=True)
    if not ui_firecrawl_config.get("enabled") or ui_firecrawl_config.get("api_key") != "fc-ui-self-test":
        raise AssertionError(f"Firecrawl UI 配置未读取运行时密钥：{ui_firecrawl_config}")
    stored_firecrawl_config = window.current_run_config(["https://example.com/"]).get("firecrawl", {})
    if stored_firecrawl_config.get("api_key") or not stored_firecrawl_config.get("api_key_present"):
        raise AssertionError(f"Firecrawl 任务档案配置不应保存明文 Key：{stored_firecrawl_config}")
    if stored_firecrawl_config.get("base_url") != "https://api.firecrawl.dev" or not stored_firecrawl_config.get("use_map"):
        raise AssertionError(f"Firecrawl 任务档案配置缺少 API/Map 摘要：{stored_firecrawl_config}")
    if (
        not stored_firecrawl_config.get("use_search")
        or stored_firecrawl_config.get("search_query") != "智能采集 自检"
        or stored_firecrawl_config.get("search_limit") != 7
        or not stored_firecrawl_config.get("use_extract")
        or "标题" not in stored_firecrawl_config.get("extract_prompt", "")
        or not stored_firecrawl_config.get("use_batch")
        or stored_firecrawl_config.get("batch_max_concurrency") != 6
        or not stored_firecrawl_config.get("use_crawl")
        or stored_firecrawl_config.get("crawl_limit") != 12
        or stored_firecrawl_config.get("crawl_max_depth") != 3
        or not stored_firecrawl_config.get("use_parse")
        or not stored_firecrawl_config.get("use_interact")
        or stored_firecrawl_config.get("interact_wait_ms") != 1500
        or "展开更多" not in stored_firecrawl_config.get("interact_prompt", "")
    ):
        raise AssertionError(f"Firecrawl 高级配置未进入任务档案摘要：{stored_firecrawl_config}")
    if window.simple_export_retry_report():
        raise AssertionError("普通人首页不应在没有重抓效果时导出空报告")
    if "重抓效果报告" not in getattr(window, "last_simple_message", ("", ""))[1]:
        raise AssertionError("普通人首页空重抓报告导出未给出明确提示")
    if window.tabs.currentWidget().objectName() != "simpleWorkbench":
        raise AssertionError("普通人首页未使用采集工作台页面结构")
    if window.simple_input_box.title() != "采集任务" or window.simple_main_splitter.orientation() != Qt.Orientation.Horizontal:
        raise AssertionError("普通人首页未采用任务区加左右分栏工作台布局")
    if window.simple_ai_box.isChecked() or window.simple_recent_box.isChecked():
        raise AssertionError("普通人首页高级设置和最近结果默认应折叠降噪")
    if len(window.simple_step_labels) != 3 or "1 输入" not in window.simple_step_labels[0].text():
        raise AssertionError("普通人面板未展示三步流程")
    if window.simple_start_button.text() != "确认并采集":
        raise AssertionError("普通人主按钮未采用两步确认采集语义")
    if window.simple_stop_button.isEnabled():
        raise AssertionError("普通人停止按钮不应在空闲时可点")
    window.refresh_simple_field_table()
    if "自动识别" not in window.simple_column_card_label.text() or "标题" not in window.simple_column_card_label.text():
        raise AssertionError(f"普通人首页未展示准备抓取的列：{window.simple_column_card_label.text()}")
    if window.simple_column_table.rowCount() < 4:
        raise AssertionError("普通人首页准备抓取的列未显示可操作清单")
    first_column_item = window.simple_column_table.item(0, 0)
    first_name_item = window.simple_column_table.item(0, 1)
    first_column_name = first_name_item.text() if first_name_item else ""
    if not first_column_item or first_column_item.checkState() != Qt.CheckState.Checked or not first_column_name:
        raise AssertionError("普通人首页准备抓取的列未默认启用")
    first_column_item.setCheckState(Qt.CheckState.Unchecked)
    unchecked_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    if first_column_name in unchecked_headers:
        raise AssertionError("普通人首页取消列勾选后整理表未同步隐藏")
    first_column_row = -1
    for row in range(window.simple_column_table.rowCount()):
        name_item = window.simple_column_table.item(row, 1)
        if name_item and name_item.text() == first_column_name:
            first_column_row = row
            break
    if first_column_row < 0:
        raise AssertionError("普通人首页取消列勾选后列清单未保留可恢复入口")
    restored_item = window.simple_column_table.item(first_column_row, 0)
    restored_item.setCheckState(Qt.CheckState.Checked)
    window.simple_column_table.selectRow(first_column_row)
    if not window.delete_selected_simple_columns():
        raise AssertionError("普通人首页删除列按钮不可用")
    deleted_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    if first_column_name in deleted_headers:
        raise AssertionError("普通人首页删除列后整理表仍显示该列")
    window.simple_column_hidden.clear()
    window.simple_column_enabled.clear()
    window.refresh_simple_field_table()
    if not hasattr(window, "simple_expert_toggle_button"):
        raise AssertionError("普通界面应提供专家模式切换按钮")
    if window.simple_expert_toggle_button.text() != "专家模式":
        raise AssertionError("专家模式按钮默认文案应为'专家模式'")
    # 验证切换专家模式后tab可见
    window.toggle_expert_mode()
    if not window.expert_mode_enabled:
        raise AssertionError("点击专家模式按钮应切换到专家模式")
    window.toggle_expert_mode()
    if window.expert_mode_enabled:
        raise AssertionError("再次点击专家模式按钮应返回简化模式")
    if window.simple_real_check_button.text() != "真实自检":
        raise AssertionError("普通人面板缺少真实抓取自检入口")
    if window.simple_depth_combo.currentData() != "deep":
        raise AssertionError("普通人一键采集应默认使用深度采集")
    window.simple_url_input.setPlainText("https://example.com/")
    window.url_input.setPlainText("https://example.com/")  # build_sample_verification_report uses url_input
    if "等待样本" not in window.build_sample_verification_report().get("summary", ""):
        raise AssertionError("普通人抽样验证空状态未提示等待样本")
    if not window.simple_ai_provider_combo.count() or not window.simple_ai_model_combo.count():
        raise AssertionError("普通人面板未提供 API 厂商和模型选择")
    window.on_real_scrape_check_result(
        {
            "ok": True,
            "row_count": 1,
            "title": "Example Domain",
            "body_preview": "This domain is for use in documentation examples",
            "link_count": 1,
            "error": "",
        }
    )
    if "真实自检通过" not in window.simple_status_label.text() or "Example Domain" not in window.simple_status_label.text():
        raise AssertionError(f"普通人真实自检结果未用简单语言展示：{window.simple_status_label.text()}")
    if "正文预览" not in window.simple_progress_label.text():
        raise AssertionError("普通人真实自检未展示抓到的正文预览")
    simple_headers = [
        window.simple_result_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_result_table.columnCount())
    ]
    if simple_headers != ["状态", "标题", "内容", "网址", "图片", "完整度", "错误"]:
        raise AssertionError(f"普通人结果表仍暴露专家字段：{simple_headers}")
    window.simple_url_input.setPlainText("https://example.com/")
    window.simple_goal_input.setPlainText("抓标题、价格、正文、图片和链接")
    window.sync_simple_inputs_to_background()
    if "https://example.com/" not in window.url_input.toPlainText() or "抓标题、价格、正文、图片和链接" not in window.ai_prompt_input.toPlainText():
        raise AssertionError("普通人面板未同步到后台采集/AI 配置")
    if "2 后台采集：进行中" not in window.simple_step_labels[1].text():
        raise AssertionError("普通人面板未进入后台采集步骤")
    direct_starts = []
    original_simple_start_collecting = window.start_collecting
    window._self_test_start_hook = lambda urls, risks: direct_starts.append({"urls": list(urls), "risks": list(risks)})
    window.start_collecting = lambda skip_confirmation=False, runtime_overrides=None: direct_starts.append(
        {
            "urls": list(window.urls_from_input()),
            "risks": [],
            "skip_confirmation": skip_confirmation,
            "runtime_overrides": dict(runtime_overrides or {}),
        }
    )
    try:
        window.simple_url_input.setPlainText("www.example.com")
        window.simple_goal_input.setPlainText("抓网页资料")
        if not window.simple_prepare_and_start_collect():
            raise AssertionError("普通人一键采集未直接启动")
        if not direct_starts or direct_starts[-1].get("urls") != ["https://www.example.com"]:
            raise AssertionError(f"普通人一键采集未自动补全并启动网址：{direct_starts}")
        if window.template_combo.currentText() != "通用自动识别":
            raise AssertionError("普通人一键采集未使用通用自动识别模板")
        if window.subpage_checkbox.isChecked() or window.page_limit_input.value() != 3:
            raise AssertionError("普通人一键采集应默认自动翻页但不暴露复杂子页面开关")
        runtime_overrides = direct_starts[-1].get("runtime_overrides", {})
        if not runtime_overrides.get("scrape_subpages") or runtime_overrides.get("subpage_limit") != 12 or not runtime_overrides.get("simple_auto_subpages"):
            raise AssertionError(f"普通人一键采集未启用深度详情页补全：{runtime_overrides}")
        if runtime_overrides.get("simple_collect_depth") != "深度":
            raise AssertionError(f"普通人一键采集未记录采集深度：{runtime_overrides}")
        normal_index = window.simple_depth_combo.findData("normal")
        window.simple_depth_combo.setCurrentIndex(normal_index)
        if not window.simple_prepare_and_start_collect():
            raise AssertionError("普通深度切换后未启动")
        normal_overrides = direct_starts[-1].get("runtime_overrides", {})
        if normal_overrides.get("subpage_limit") != 3 or normal_overrides.get("simple_collect_depth") != "普通":
            raise AssertionError(f"普通模式采集深度参数错误：{normal_overrides}")
        complete_index = window.simple_depth_combo.findData("complete")
        window.simple_depth_combo.setCurrentIndex(complete_index)
        if not window.simple_prepare_and_start_collect():
            raise AssertionError("完整深度切换后未启动")
        complete_overrides = direct_starts[-1].get("runtime_overrides", {})
        if complete_overrides.get("subpage_limit") != 30 or complete_overrides.get("simple_collect_depth") != "完整":
            raise AssertionError(f"完整模式采集深度参数错误：{complete_overrides}")
        if window.page_limit_input.value() != 5 or window.scroll_times_input.value() < 5:
            raise AssertionError("完整模式未提高分页或滚动采集参数")
        window.clear_current_results()
        window.simple_url_input.setPlainText("https://example.com/dual")
        window.sync_simple_inputs_to_background()
        if not window.simple_run_strategy_comparison():
            raise AssertionError("缺少实测样本时未自动启动普通/完整双跑")
        dual_normal_overrides = direct_starts[-1].get("runtime_overrides", {})
        if (
            dual_normal_overrides.get("simple_collect_depth") != "普通"
            or dual_normal_overrides.get("subpage_limit") != 3
            or dual_normal_overrides.get("skip_unchanged") is not False
        ):
            raise AssertionError(f"实测对比首轮未使用普通模式：{dual_normal_overrides}")
        if not window.strategy_dual_run_active or window.strategy_dual_run_step != "普通":
            raise AssertionError("实测对比未记录普通首轮状态")
        if not window.maybe_continue_strategy_dual_run("finished"):
            raise AssertionError("实测对比普通模式结束后未自动接力完整模式")
        dual_complete_overrides = direct_starts[-1].get("runtime_overrides", {})
        if dual_complete_overrides.get("simple_collect_depth") != "完整" or dual_complete_overrides.get("subpage_limit") != 30:
            raise AssertionError(f"实测对比第二轮未使用完整模式：{dual_complete_overrides}")
        window.records = [
            {
                "simple_collect_depth": "普通",
                "completeness_score": 20,
                "completeness_label": "20% 偏少",
                "completeness_missing": ["图片", "链接"],
                "images": [],
                "links": [],
                "tables": [],
                "url": "https://example.com/dual",
            },
            {
                "simple_collect_depth": "完整",
                "completeness_score": 90,
                "completeness_label": "90% 完整",
                "completeness_missing": [],
                "images": [{"url": "x"}],
                "links": [{"url": "y"}],
                "tables": [[["a"]]],
                "url": "https://example.com/dual",
            },
        ]
        if window.maybe_continue_strategy_dual_run("finished"):
            raise AssertionError("实测对比完整模式结束后不应继续启动第三轮")
        if not window.finalize_strategy_dual_run_report() or "推荐 完整" not in window.simple_strategy_compare_label.text():
            raise AssertionError(f"实测对比双跑结束后未生成完整推荐：{window.simple_strategy_compare_label.text()}")
        window.clear_current_results()
    finally:
        window.start_collecting = original_simple_start_collecting
        if hasattr(window, "_self_test_start_hook"):
            delattr(window, "_self_test_start_hook")
    window.simple_url_input.setPlainText("https://example.com/")
    window.simple_goal_input.setPlainText("抓标题、价格、正文、图片和链接")
    window.sync_simple_inputs_to_background()
    window.add_record(
        {
            "collected_at": "2026-06-09 10:40:00",
            "url": "https://example.com/",
            "domain": "example.com",
            "template_name": "通用自动识别",
            "title": "Example Domain",
            "price": "9.9",
            "body": "示例正文",
            "images": [{"url": "https://example.com/a.png"}],
            "links": [{"text": "More", "url": "https://example.com/more"}],
            "tables": [[["字段", "值"], ["名称", "示例"]]],
            "fingerprint": "simple",
        }
    )
    if window.simple_result_table.rowCount() < 1 or window.simple_result_table.item(0, 1).text() != "Example Domain":
        raise AssertionError("后台采集结果未同步到普通人面板")
    if "示例正文" not in window.simple_result_table.item(0, 2).text():
        raise AssertionError("普通人面板未展示采集内容摘要")
    if "%" not in window.simple_result_table.item(0, 5).text():
        raise AssertionError("普通人结果表未展示资料完整度")
    if not isinstance(window.simple_result_table.cellWidget(0, 5), QProgressBar):
        raise AssertionError("普通人结果表完整度未使用进度条展示")
    if "缺少" not in window.simple_result_table.item(0, 0).toolTip():
        raise AssertionError("普通人结果表未在状态提示中展示缺项")
    if "共 1 条" not in window.simple_result_summary_label.text() or "平均完整度" not in window.simple_result_summary_label.text():
        raise AssertionError("普通人面板未显示结果摘要")
    if "诊断建议" not in window.simple_diagnosis_label.text():
        raise AssertionError("普通人面板未显示采集诊断建议")
    first_sample_report = window.build_sample_verification_report()
    if not first_sample_report.get("recommendation") or "样本" not in first_sample_report.get("summary", ""):
        raise AssertionError(f"普通首页抽样验证未给出样本推荐：{first_sample_report}")
    strategy_report = window.build_strategy_comparison_report(
        [
            {
                "url": "https://example.com/item",
                "title": "普通样本",
                "body": "普通模式只抓到很短的列表摘要。",
                "images": [],
                "links": [],
                "tables": [],
                "simple_collect_depth": "普通",
            },
            {
                "url": "https://example.com/item",
                "title": "完整样本",
                "body": "完整模式抓到了详情页正文，包含价格、库存、规格、图片、分页和子链接资料，内容明显更完整。",
                "price": "88",
                "published_time": "2026-06-09",
                "author": "Example",
                "images": [{"url": "https://example.com/a.png"}],
                "links": [{"text": "详情", "url": "https://example.com/detail"}],
                "tables": [[["规格", "值"]]],
                "simple_collect_depth": "完整",
            },
        ]
    )
    if strategy_report.get("best") != "完整" or strategy_report.get("delta", 0) <= 0:
        raise AssertionError(f"普通/完整实测对比未推荐完整模式：{strategy_report}")
    original_strategy_records = list(window.records)
    window.records = [
        {
            "simple_collect_depth": "普通",
            "completeness_score": 20,
            "completeness_label": "20% 偏少",
            "completeness_missing": ["图片", "链接", "表格/规格"],
            "images": [],
            "links": [],
            "tables": [],
            "url": "https://example.com/item",
        },
        {
            "simple_collect_depth": "完整",
            "completeness_score": 95,
            "completeness_label": "95% 完整",
            "completeness_missing": [],
            "images": [{"url": "x"}],
            "links": [{"url": "y"}],
            "tables": [[["a"]]],
            "url": "https://example.com/item",
        },
    ]
    try:
        if not window.simple_run_strategy_comparison() or "推荐 完整" not in window.simple_strategy_compare_label.text():
            raise AssertionError(f"普通首页实测对比未展示完整推荐：{window.simple_strategy_compare_label.text()}")
        if window.simple_depth_combo.currentData() != "complete":
            raise AssertionError("普通首页实测对比推荐完整后未切换完整模式")
    finally:
        window.records = original_strategy_records
    if "Example Domain" not in window.simple_preview_title_label.text():
        raise AssertionError("普通人结果预览未展示标题")
    if "示例正文" not in window.simple_preview_body_output.toPlainText():
        raise AssertionError("普通人结果预览未展示正文")
    if (
        "图片 1" not in window.simple_preview_counts_label.text()
        or "链接 1" not in window.simple_preview_counts_label.text()
        or "表格 1" not in window.simple_preview_counts_label.text()
        or "完整度" not in window.simple_preview_counts_label.text()
    ):
        raise AssertionError("普通人结果预览未展示资料数量")
    simple_field_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    for expected_header in ("网址", "标题", "价格", "正文", "图片", "链接", "完整度"):
        if expected_header not in simple_field_headers:
            raise AssertionError(f"普通人智能字段表缺少列：{expected_header}")
    if "本地规则" not in window.simple_field_status_label.text() or "关键列都有内容" not in window.simple_field_status_label.text():
        raise AssertionError(f"普通人字段状态未说明本地整理结果：{window.simple_field_status_label.text()}")
    simple_field_text = "\t".join(
        window.simple_field_table.item(0, column).text() if window.simple_field_table.item(0, column) else ""
        for column in range(window.simple_field_table.columnCount())
    )
    for expected_value in ("Example Domain", "9.9", "示例正文", "https://example.com/a.png", "https://example.com/more", "%"):
        if expected_value not in simple_field_text:
            raise AssertionError(f"普通人智能字段表缺少值：{expected_value}")
    before_merge_simple_rows = window.simple_result_table.rowCount()
    window.simple_merge_subpage_results = True
    window.add_record(
        {
            "collected_at": "2026-06-09 10:41:00",
            "url": "https://example.com/more",
            "domain": "example.com",
            "template_name": "通用自动识别",
            "title": "详情页",
            "body": "详情页补充参数：蓝色、大号、有库存",
            "images": [{"url": "https://example.com/detail.png"}],
            "links": [],
            "tables": [],
            "fingerprint": "detail",
        }
    )
    if len(window.records) < 2 or window.simple_result_table.rowCount() != before_merge_simple_rows:
        raise AssertionError("普通首页详情页补全不应新增普通结果行")
    if "详情页补充参数" not in window.records[0].get("body", ""):
        raise AssertionError("普通首页未把详情页正文合并到主记录")
    if not window.records[0].get("simple_detail_enriched") or "https://example.com/more" not in window.records[0].get("simple_detail_urls", []):
        raise AssertionError("普通首页未标记详情页补全来源")
    if "detail.png" not in "\t".join(item.get("url", "") for item in window.records[0].get("images", []) or []):
        raise AssertionError("普通首页未把详情页图片合并到主记录")
    merged_simple_text = "\t".join(
        window.simple_field_table.item(0, column).text() if window.simple_field_table.item(0, column) else ""
        for column in range(window.simple_field_table.columnCount())
    )
    if "详情页补充参数" not in merged_simple_text:
        raise AssertionError("普通首页按要求整理表未展示合并后的详情资料")
    if not window.records[0].get("completeness_score") or "完整度" not in window.simple_preview_counts_label.text():
        raise AssertionError("普通首页详情合并后未刷新资料完整度")
    window.simple_merge_subpage_results = False
    window.add_record(
        {
            "collected_at": "2026-06-09 10:42:00",
            "url": "https://example.com/weak",
            "domain": "example.com",
            "template_name": "通用自动识别",
            "title": "低完整度页",
            "body": "",
            "images": [],
            "links": [],
            "tables": [],
            "fingerprint": "weak",
        }
    )
    low_row = window.simple_result_table.rowCount() - 1
    if not isinstance(window.simple_result_table.cellWidget(low_row, 5), QProgressBar):
        raise AssertionError("低完整度结果未展示完整度条")
    if "缺少" not in window.simple_result_table.item(low_row, 1).toolTip():
        raise AssertionError("低完整度结果标题未提示缺少资料")
    low_quality_urls = window.low_quality_urls()
    if "https://example.com/weak" not in low_quality_urls:
        raise AssertionError(f"低完整度筛选未发现弱结果：{low_quality_urls}")
    weak_diagnosis = next(
        (row for row in window.simple_crawl_diagnosis_rows() if row.get("url") == "https://example.com/weak"),
        {},
    )
    if weak_diagnosis.get("reason") != "疑似动态加载" or "完整模式" not in weak_diagnosis.get("advice", ""):
        raise AssertionError(f"低完整度动态加载诊断不正确：{weak_diagnosis}")
    if "诊断建议" not in window.simple_diagnosis_label.text() or "动态加载" not in window.simple_diagnosis_label.text():
        raise AssertionError(f"普通首页诊断摘要未指出主要原因：{window.simple_diagnosis_label.text()}")
    dynamic_sample_report = window.build_sample_verification_report()
    if dynamic_sample_report.get("recommendation") != "完整" or "动态加载" not in dynamic_sample_report.get("summary", ""):
        raise AssertionError(f"动态样本抽样验证未推荐完整模式：{dynamic_sample_report}")
    if not window.simple_run_sample_verification() or "推荐 完整" not in window.simple_sample_verify_label.text():
        raise AssertionError(f"普通首页抽样验证未展示完整模式推荐：{window.simple_sample_verify_label.text()}")
    if window.simple_depth_combo.currentData() != "complete" or window.scroll_times_input.value() < 3:
        raise AssertionError("抽样验证推荐完整模式后未调整采集深度")
    blocked_diagnosis = window.crawl_diagnosis_for_record(
        {
            "url": "https://example.com/blocked",
            "title": "",
            "body": "",
            "images": [],
            "links": [],
            "tables": [],
            "error": "403 验证码拦截",
        }
    )
    if blocked_diagnosis.get("reason") != "反爬或权限限制" or "真实浏览器" not in blocked_diagnosis.get("advice", ""):
        raise AssertionError(f"反爬/权限诊断不正确：{blocked_diagnosis}")
    pagination_diagnosis = window.crawl_diagnosis_for_record(
        {
            "url": "https://example.com/list",
            "title": "列表页",
            "body": "列表摘要",
            "images": [],
            "links": [{"text": "下一页", "url": "https://example.com/list?page=2"}],
            "tables": [],
        }
    )
    if pagination_diagnosis.get("reason") != "分页可能未继续" or "下一页" not in pagination_diagnosis.get("advice", ""):
        raise AssertionError(f"分页未继续诊断不正确：{pagination_diagnosis}")
    detail_diagnosis = window.crawl_diagnosis_for_record(
        {
            "url": "https://example.com/list-item",
            "title": "列表项",
            "body": "这是列表页摘要内容，正文已经足够长，但图片、价格和规格可能藏在同站详情页里，需要继续展开详情。",
            "images": [],
            "links": [{"text": "详情", "url": "https://example.com/detail"}],
            "tables": [],
        }
    )
    if detail_diagnosis.get("reason") != "子链接未展开" or "子链接" not in detail_diagnosis.get("advice", ""):
        raise AssertionError(f"子链接未展开诊断不正确：{detail_diagnosis}")
    no_subpage_diagnosis = window.crawl_diagnosis_for_record(
        {
            "url": "https://example.com/no-links",
            "title": "无子链接页",
            "body": "页面正文已经抓到足够长的产品介绍，但没有发现可以继续进入的详情页链接，需要扫描或手动选择子页面。",
            "price": "88",
            "published_time": "2026-06-09",
            "author": "Example",
            "images": [],
            "links": [],
            "tables": [[["规格", "值"]]],
        }
    )
    if no_subpage_diagnosis.get("reason") != "子链接候选不足" or "扫描" not in no_subpage_diagnosis.get("advice", ""):
        raise AssertionError(f"子链接候选不足诊断不正确：{no_subpage_diagnosis}")
    field_rule_diagnosis = window.crawl_diagnosis_for_record(
        {
            "url": "https://example.com/custom",
            "title": "字段规则页",
            "body": "页面正文已经抓到很多文字内容，说明网页可以访问，但用户需要的价格、作者、时间、图片等结构化字段没有被当前规则提取出来。",
            "images": [],
            "links": [],
            "tables": [],
        }
    )
    if field_rule_diagnosis.get("reason") != "字段规则可能不匹配" or "AI 建议列" not in field_rule_diagnosis.get("advice", ""):
        raise AssertionError(f"字段规则诊断不正确：{field_rule_diagnosis}")
    repair_records_backup = list(window.records)
    window.records = [
        {
            "url": "https://example.com/list",
            "title": "分页页",
            "body": "列表摘要",
            "images": [],
            "links": [{"text": "下一页", "url": "https://example.com/list?page=2"}],
            "tables": [],
        },
        {
            "url": "https://example.com/list-item",
            "title": "子链接页",
            "body": "这是列表页摘要内容，正文已经足够长，但图片、价格和规格可能藏在同站详情页里，需要继续展开详情。",
            "images": [],
            "links": [{"text": "详情", "url": "https://example.com/detail"}],
            "tables": [],
        },
        {
            "url": "https://example.com/blocked",
            "title": "",
            "body": "",
            "images": [],
            "links": [],
            "tables": [],
            "error": "403 验证码拦截",
        },
        {
            "url": "https://example.com/custom",
            "title": "字段规则页",
            "body": "页面正文已经抓到很多文字内容，说明网页可以访问，但用户需要的价格、作者、时间、图片等结构化字段没有被当前规则提取出来。",
            "images": [],
            "links": [],
            "tables": [],
        },
    ]
    repair_groups = window.simple_repair_plan_groups()
    if (
        "https://example.com/list" not in repair_groups.get("pagination", {}).get("urls", [])
        or "https://example.com/list-item" not in repair_groups.get("subpages", {}).get("urls", [])
        or "https://example.com/blocked" not in repair_groups.get("login", {}).get("urls", [])
        or "https://example.com/custom" not in repair_groups.get("fields", {}).get("urls", [])
    ):
        raise AssertionError(f"修复方案未按原因分组：{repair_groups}")
    window.refresh_simple_result_summary()
    if "分页 1 条" not in window.simple_repair_plan_label.text() or "子链接 1 条" not in window.simple_repair_plan_label.text():
        raise AssertionError(f"修复方案摘要未展示分页/子链接分组：{window.simple_repair_plan_label.text()}")
    repair_starts = []
    original_start_for_repair_plan = window.start_collecting
    window.start_collecting = lambda skip_confirmation=False, runtime_overrides=None: repair_starts.append(
        {
            "urls": list(window.urls_from_input()),
            "skip_confirmation": skip_confirmation,
            "runtime_overrides": dict(runtime_overrides or {}),
        }
    )
    try:
        if not window.simple_apply_repair_plan_action("pagination"):
            raise AssertionError("分页修复方案未启动")
    finally:
        window.start_collecting = original_start_for_repair_plan
    if repair_starts[-1].get("urls") != ["https://example.com/list"]:
        raise AssertionError(f"分页修复方案未只重抓分页网址：{repair_starts}")
    repair_overrides = repair_starts[-1].get("runtime_overrides", {})
    if repair_overrides.get("simple_collect_depth") != "完整" or repair_overrides.get("skip_unchanged") is not False:
        raise AssertionError(f"分页修复方案未使用完整重抓参数：{repair_overrides}")
    repair_baseline = getattr(window, "low_quality_retry_baseline", {}).get("https://example.com/list", {})
    if not repair_baseline or repair_baseline.get("reason", "").find("分页") < 0:
        raise AssertionError(f"分页修复方案未保存重抓前对照基线：{repair_baseline}")
    window.add_record(
        {
            "collected_at": "2026-06-09 10:42:30",
            "url": "https://example.com/list",
            "domain": "example.com",
            "template_name": "通用自动识别",
            "title": "分页页",
            "price": "66",
            "published_time": "2026-06-09",
            "author": "Example",
            "body": "这是完整模式翻页后补到的正文内容，包含第二页、第三页和更多列表资料，明显比原来的列表摘要更完整。",
            "images": [{"url": "https://example.com/list.jpg"}],
            "links": [
                {"text": "详情1", "url": "https://example.com/detail/1"},
                {"text": "详情2", "url": "https://example.com/detail/2"},
            ],
            "tables": [[["页码", "内容"]]],
        }
    )
    repair_report_row = next(
        (row for row in getattr(window, "low_quality_retry_report_rows", []) if row.get("url") == "https://example.com/list"),
        {},
    )
    if not repair_report_row or repair_report_row.get("link_delta", 0) <= 0 or repair_report_row.get("body_delta", 0) <= 0:
        raise AssertionError(f"分页修复方案未生成真实采集对照报告：{repair_report_row}")
    repair_report_text = window.simple_retry_report_label.text()
    if "多抓正文" not in repair_report_text or "链接" not in repair_report_text:
        raise AssertionError(f"真实采集对照摘要未展示新增资料数量：{repair_report_text}")
    window.records = repair_records_backup
    window.low_quality_retry_baseline = {}
    window.low_quality_retry_active = False
    window.low_quality_retry_report_rows = []
    window.refresh_simple_result_summary()
    diagnosis_retry_starts = []
    original_start_for_diagnosis_action = window.start_collecting
    window.start_collecting = lambda skip_confirmation=False, runtime_overrides=None: diagnosis_retry_starts.append(
        {
            "urls": list(window.urls_from_input()),
            "skip_confirmation": skip_confirmation,
            "runtime_overrides": dict(runtime_overrides or {}),
        }
    )
    try:
        if not window.simple_apply_diagnosis_action():
            raise AssertionError("普通首页应用动态加载诊断建议失败")
    finally:
        window.start_collecting = original_start_for_diagnosis_action
    if not diagnosis_retry_starts or "https://example.com/weak" not in diagnosis_retry_starts[-1].get("urls", []):
        raise AssertionError(f"应用诊断建议未发起低完整度重抓：{diagnosis_retry_starts}")
    diagnosis_overrides = diagnosis_retry_starts[-1].get("runtime_overrides", {})
    if diagnosis_overrides.get("simple_collect_depth") != "完整" or diagnosis_overrides.get("skip_unchanged") is not False:
        raise AssertionError(f"应用诊断建议未使用完整重抓参数：{diagnosis_overrides}")
    if window.simple_depth_combo.currentData() != "complete" or not window.use_browser_checkbox.isChecked() or window.scroll_times_input.value() < 3:
        raise AssertionError("应用动态加载诊断建议未切换完整采集配置")
    diagnosis_records_backup = list(window.records)
    window.records = [
        {
            "url": "https://example.com/blocked",
            "title": "",
            "body": "",
            "images": [],
            "links": [],
            "tables": [],
            "error": "403 验证码拦截",
        }
    ]
    window.keep_login_checkbox.setChecked(False)
    window.use_browser_checkbox.setChecked(False)
    window.delay_input.setValue(0)
    window.scroll_times_input.setValue(0)
    window.refresh_simple_result_summary()
    if not window.simple_apply_diagnosis_action():
        raise AssertionError("普通首页应用反爬诊断建议失败")
    if not window.use_browser_checkbox.isChecked() or not window.keep_login_checkbox.isChecked() or window.delay_input.value() < 3:
        raise AssertionError("应用反爬诊断建议未启用真实浏览器、保留登录和慢速访问")
    if "真实浏览器" not in window.simple_status_label.text() or "保留登录" not in window.simple_status_label.text():
        raise AssertionError(f"应用反爬诊断建议未更新状态：{window.simple_status_label.text()}")
    blocked_sample_report = window.build_sample_verification_report()
    if blocked_sample_report.get("recommendation") != "登录浏览器":
        raise AssertionError(f"反爬样本抽样验证未推荐登录浏览器：{blocked_sample_report}")
    window.records = [
        {
            "url": "https://example.com/custom",
            "title": "字段规则页",
            "body": "页面正文已经抓到很多文字内容，说明网页可以访问，但用户需要的价格、作者、时间、图片等结构化字段没有被当前规则提取出来。",
            "images": [],
            "links": [],
            "tables": [],
        }
    ]
    window.simple_url_input.setPlainText("https://example.com/custom")
    captured_diagnosis_suggest = []
    original_maybe_suggest_for_diagnosis = window.maybe_start_simple_ai_suggest_fields
    window.maybe_start_simple_ai_suggest_fields = lambda urls: captured_diagnosis_suggest.append(list(urls)) or True
    try:
        if not window.simple_apply_diagnosis_action():
            raise AssertionError("普通首页应用字段规则诊断建议失败")
    finally:
        window.maybe_start_simple_ai_suggest_fields = original_maybe_suggest_for_diagnosis
    if captured_diagnosis_suggest != [["https://example.com/custom"]]:
        raise AssertionError(f"应用字段规则诊断建议未触发 AI 建议列：{captured_diagnosis_suggest}")
    window.records = diagnosis_records_backup
    window.refresh_simple_result_summary()
    low_quality_queue = window.low_quality_retry_queue()
    weak_queue_row = next((row for row in low_quality_queue if row.get("url") == "https://example.com/weak"), {})
    if not weak_queue_row or "低完整度页" not in weak_queue_row.get("title", "") or not weak_queue_row.get("missing"):
        raise AssertionError(f"低完整度重抓预览队列缺少标题、完整度或缺项：{low_quality_queue}")
    selected_preview_urls = window.confirm_low_quality_retry_queue(low_quality_queue)
    if "https://example.com/weak" not in selected_preview_urls:
        raise AssertionError(f"低完整度重抓预览未默认选择弱结果：{selected_preview_urls}")
    low_quality_starts = []
    original_low_quality_start = window.start_collecting
    window.start_collecting = lambda skip_confirmation=False, runtime_overrides=None: low_quality_starts.append(
        {
            "urls": list(window.urls_from_input()),
            "skip_confirmation": skip_confirmation,
            "runtime_overrides": dict(runtime_overrides or {}),
        }
    )
    try:
        if not window.simple_retry_low_quality_items():
            raise AssertionError("低完整度一键重抓未启动")
    finally:
        window.start_collecting = original_low_quality_start
    if not low_quality_starts or "https://example.com/weak" not in low_quality_starts[-1].get("urls", []):
        raise AssertionError(f"低完整度一键重抓未带入弱结果网址：{low_quality_starts}")
    low_quality_overrides = low_quality_starts[-1].get("runtime_overrides", {})
    if low_quality_overrides.get("subpage_limit") != 30 or low_quality_overrides.get("simple_collect_depth") != "完整" or low_quality_overrides.get("skip_unchanged") is not False:
        raise AssertionError(f"低完整度一键重抓未使用完整重采参数：{low_quality_overrides}")
    if "https://example.com/weak" not in getattr(window, "low_quality_retry_baseline", {}):
        raise AssertionError("低完整度重抓未保存重抓前完整度基线")
    window.add_record(
        {
            "collected_at": "2026-06-09 10:43:00",
            "url": "https://example.com/weak",
            "domain": "example.com",
            "template_name": "通用自动识别",
            "title": "低完整度页",
            "price": "88",
            "published_time": "2026-06-09",
            "author": "Example",
            "body": "这是重抓后补充出的详情正文，包含足够多的产品描述、参数、规格和页面内容，用于证明完整模式采集到了更多资料。",
            "images": [{"url": "https://example.com/weak.jpg"}],
            "links": [{"text": "详情", "url": "https://example.com/detail"}],
            "tables": [[["规格", "值"], ["颜色", "黑色"]]],
            "fingerprint": "weak-retry-rich",
        }
    )
    retry_report_rows = getattr(window, "low_quality_retry_report_rows", [])
    report_row = next((row for row in retry_report_rows if row.get("url") == "https://example.com/weak"), {})
    if not report_row or report_row.get("after", 0) <= report_row.get("before", 0):
        raise AssertionError(f"低完整度重抓未生成前后提升报告：{retry_report_rows}")
    if "图片" not in report_row.get("captured", "") or "价格" not in report_row.get("captured", ""):
        raise AssertionError(f"低完整度重抓报告未展示补到的字段：{report_row}")
    retry_report_text = window.simple_retry_report_label.text()
    if "重抓效果" not in retry_report_text or "+" not in retry_report_text or "补到" not in retry_report_text:
        raise AssertionError(f"低完整度重抓效果摘要不完整：{retry_report_text}")
    retry_columns, retry_rows = window.retry_report_table_data()
    for expected_header in ("网址", "重抓原因", "重抓前完整度", "重抓后完整度", "提升分数", "多抓正文字数", "新增图片", "新增链接", "新增表格", "补到资料", "仍缺资料"):
        if expected_header not in retry_columns:
            raise AssertionError(f"低完整度重抓报告导出列缺失：{retry_columns}")
    if not retry_rows or "https://example.com/weak" not in "\t".join(str(value) for value in retry_rows[0]):
        raise AssertionError(f"低完整度重抓报告导出行缺失：{retry_rows}")
    if not window.simple_export_retry_report():
        raise AssertionError("低完整度重抓报告导出失败")
    retry_export_path = getattr(window, "last_simple_retry_report_export_path", "")
    if not os.path.exists(retry_export_path) or not retry_export_path.endswith(".xlsx"):
        raise AssertionError("低完整度重抓报告未保存为 Excel")
    from openpyxl import load_workbook as load_retry_report_workbook

    retry_report_workbook = load_retry_report_workbook(retry_export_path)
    if retry_report_workbook.sheetnames[0] != "重抓效果报告":
        raise AssertionError("低完整度重抓报告 Excel 工作表名称错误")
    retry_report_sheet = retry_report_workbook["重抓效果报告"]
    retry_saved_headers = [
        retry_report_sheet.cell(1, column).value
        for column in range(1, retry_report_sheet.max_column + 1)
    ]
    for expected_header in ("网址", "重抓前完整度", "重抓后完整度", "提升分数", "补到资料"):
        if expected_header not in retry_saved_headers:
            raise AssertionError(f"低完整度重抓报告 Excel 缺少列：{retry_saved_headers}")
    retry_saved_row = "\t".join(
        str(retry_report_sheet.cell(2, column).value or "")
        for column in range(1, retry_report_sheet.max_column + 1)
    )
    for expected_value in ("https://example.com/weak", "图片", "价格"):
        if expected_value not in retry_saved_row:
            raise AssertionError(f"低完整度重抓报告 Excel 缺少值：{expected_value} / {retry_saved_row}")
    window.simple_merge_subpage_results = False
    window.set_simple_flow_step("导出")
    window.simple_ai_field_rules = []
    fallback_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    window.apply_simple_ai_fields(
        {
            "fields": [
                {"name": "AI标题", "selector": "h1", "attr": "text", "multiple": False},
                {"name": "AI链接", "selector": "a", "attr": "href", "multiple": True},
            ]
        }
    )
    ai_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    if "AI标题" not in ai_headers or "AI链接" not in ai_headers or "价格" in ai_headers:
        raise AssertionError(f"普通首页 AI 建议列未接管整理表：{ai_headers}")
    if "AI 建议" not in window.simple_column_card_label.text() or "AI标题" not in window.simple_column_card_label.text():
        raise AssertionError("普通首页 AI 建议列未同步到列卡片")
    if "AI 智能建议" not in window.simple_field_status_label.text() or "暂未抓到：AI标题" not in window.simple_field_status_label.text():
        raise AssertionError(f"普通首页 AI 字段状态未提示缺失列：{window.simple_field_status_label.text()}")
    window.on_ai_result("simple_suggest_fields", {"error": "模拟失败"})
    failed_headers = [
        window.simple_field_table.horizontalHeaderItem(column).text()
        for column in range(window.simple_field_table.columnCount())
    ]
    if failed_headers != fallback_headers:
        raise AssertionError("普通首页 AI 建议失败后未回退本地字段规则")
    if "本地规则" not in window.simple_field_status_label.text():
        raise AssertionError("普通首页 AI 失败后字段状态未回退本地规则")
    window.ai_base_url_input.setText("https://api.openai.com/v1")
    window.ai_key_input.setText("")
    window.ai_key_name_input.setText("默认 Key")
    if window.simple_has_ai_settings():
        raise AssertionError("普通首页无 API Key 时不应调用远程建议列")
    window.ai_key_input.setText("sk-self-test")
    if not window.simple_has_ai_settings():
        raise AssertionError("普通首页有 API Key 时应允许远程建议列")
    captured_simple_suggest = []
    original_simple_fetch_static = UniversalCollector.fetch_static
    original_simple_run_ai_worker = window.run_ai_worker
    UniversalCollector.fetch_static = lambda self, url: "<html><body><h1>Example Domain</h1><a href='/more'>More</a></body></html>"
    window.run_ai_worker = lambda action, payload=None: captured_simple_suggest.append((action, payload or {}))
    try:
        if not window.maybe_start_simple_ai_suggest_fields(["https://example.com/"]):
            raise AssertionError("普通首页未触发 API 智能建议列")
    finally:
        UniversalCollector.fetch_static = original_simple_fetch_static
        window.run_ai_worker = original_simple_run_ai_worker
    if not captured_simple_suggest or captured_simple_suggest[0][0] != "simple_suggest_fields":
        raise AssertionError("普通首页 AI 智能建议列未走专用后台动作")
    if "Example Domain" not in captured_simple_suggest[0][1].get("html", ""):
        raise AssertionError("普通首页 AI 智能建议列未携带网页快照")
    window.refresh_simple_recent_area()
    if window.simple_recent_records_table.rowCount() < 1:
        raise AssertionError("普通人首页未显示最近采集记录")
    if "3 导出：进行中" not in window.simple_step_labels[2].text():
        raise AssertionError("普通人面板采到结果后未进入导出步骤")
    if window.simple_export_button.text() != "自动保存":
        raise AssertionError("普通人面板导出按钮未改为自动保存")
    if not window.simple_auto_save_results():
        raise AssertionError("普通人面板网页结果自动保存失败")
    web_export_path = getattr(window, "last_simple_export_path", "")
    if not os.path.exists(web_export_path) or not web_export_path.endswith(".xlsx"):
        raise AssertionError("普通人面板网页结果未保存为 Excel")
    from openpyxl import load_workbook as load_simple_workbook

    simple_workbook = load_simple_workbook(web_export_path)
    if simple_workbook.sheetnames[0] != "按要求整理结果":
        raise AssertionError("普通人面板网页结果未优先导出按要求整理表")
    simple_sheet = simple_workbook["按要求整理结果"]
    saved_headers = [simple_sheet.cell(1, column).value for column in range(1, simple_sheet.max_column + 1)]
    for expected_header in ("网址", "标题", "价格", "正文", "图片", "链接", "完整度"):
        if expected_header not in saved_headers:
            raise AssertionError(f"普通人自动保存 Excel 缺少列：{expected_header}")
    saved_row_text = "\t".join(str(simple_sheet.cell(2, column).value or "") for column in range(1, simple_sheet.max_column + 1))
    for expected_value in ("Example Domain", "9.9", "示例正文", "https://example.com/a.png", "https://example.com/more"):
        if expected_value not in saved_row_text:
            raise AssertionError(f"普通人自动保存 Excel 缺少值：{expected_value}")
    if window.simple_recent_files_table.rowCount() < 1:
        raise AssertionError("普通人首页未显示最近保存的 Excel")
    window.simple_recent_files_table.selectRow(0)
    if window.selected_simple_recent_file_path() != web_export_path:
        raise AssertionError("普通人首页未选中最近保存的 Excel")
    if not window.open_selected_simple_recent_file() or getattr(window, "last_simple_open_path", "") != os.path.abspath(web_export_path):
        raise AssertionError("普通人首页打开最近 Excel 文件失败")
    if not window.open_simple_recent_export_folder() or getattr(window, "last_simple_open_path", "") != os.path.abspath(window.simple_export_dir()):
        raise AssertionError("普通人首页打开 Excel 文件夹失败")
    contact_result = window.simple_extract_contacts()
    if not isinstance(contact_result, dict) or "emails" not in contact_result or "phones" not in contact_result:
        raise AssertionError("普通人首页未能使用联系方式提取能力")
    image_results = window.simple_download_images()
    if not isinstance(image_results, list):
        raise AssertionError("普通人首页未能使用图片下载能力")
    window.simple_url_input.setPlainText("https://example.com/")
    schedule_count_before = len(window.schedules or [])
    scheduled_item = window.simple_add_schedule()
    if not scheduled_item or len(window.schedules or []) != schedule_count_before + 1:
        raise AssertionError("普通人首页未能创建定时监控")
    window.simple_ai_provider_combo.setCurrentIndex(max(0, window.simple_ai_provider_combo.findData("openai")))
    window.simple_ai_key_input.setText("sk-self-test")
    if not window.save_simple_ai_settings() or window.ai_settings.get("provider") != "openai":
        raise AssertionError("普通人首页未能保存 API 设置")
    if not window.simple_suggest_columns_now():
        raise AssertionError("普通人首页未能触发 AI 建议列入口")
    window.update_collect_progress({"processed": 1, "success": 1, "failed": 0, "total": 2, "stage": "采集页面", "status": "running"})
    if "后台采集中" not in window.simple_progress_label.text() or "复杂步骤" not in window.simple_status_label.text():
        raise AssertionError("普通人面板未用简单语言展示后台进度")
    window.task_queue_rows = [
        {
            "status": "失败",
            "type": "实际",
            "stage": "页面完成",
            "url": "https://example.com/slow",
            "error": "Timeout 30000ms exceeded",
            "error_category": "网络超时",
            "error_advice": "建议调低速度或稍后重试",
        }
    ]
    window.apply_task_queue_filters()
    window.update_collect_progress({"processed": 1, "success": 0, "failed": 1, "total": 1, "stage": "页面完成", "status": "partial"})
    if "重试失败" not in window.simple_progress_label.text() or "网站可能较慢" not in window.simple_progress_label.text():
        raise AssertionError(f"普通人面板未提示慢网页重试：{window.simple_progress_label.text()}")
    retry_calls_from_simple = []
    original_start_collecting_for_simple_retry = window.start_collecting
    window.start_collecting = lambda skip_confirmation=False, runtime_overrides=None: retry_calls_from_simple.append(
        {"urls": list(window.urls_from_input()), "skip_confirmation": skip_confirmation}
    )
    try:
        if not window.simple_retry_failed_items():
            raise AssertionError("普通人首页重试失败入口不可用")
    finally:
        window.start_collecting = original_start_collecting_for_simple_retry
    if retry_calls_from_simple != [{"urls": ["https://example.com/slow"], "skip_confirmation": True}]:
        raise AssertionError(f"普通人首页重试失败未回填并启动：{retry_calls_from_simple}")
    class _SelfTestStopWorker:
        def __init__(self):
            self.stopped = False

        def stop(self):
            self.stopped = True

        def should_stop(self):
            return self.stopped

    window.worker = _SelfTestStopWorker()
    window.set_collecting_buttons_state(True)
    if window.simple_start_button.isEnabled() or not window.simple_stop_button.isEnabled():
        raise AssertionError("普通人面板采集中按钮状态不正确")
    window.stop_collecting()
    if "正在停止" not in window.simple_stop_button.text() or "正在停止采集" not in window.simple_status_label.text():
        raise AssertionError("普通人面板停止采集未给出明确反馈")
    window.worker = None
    window.set_collecting_buttons_state(False)
    if not window.simple_start_button.isEnabled() or window.simple_stop_button.isEnabled():
        raise AssertionError("普通人面板采集结束后按钮状态未恢复")
    window.clear_current_results()
    if window.simple_result_table.rowCount() != 0 or window.simple_result_summary_label.text() != "结果：暂无":
        raise AssertionError("普通人面板清空结果后摘要未复位")
    if "未选择结果" not in window.simple_preview_title_label.text() or window.simple_preview_body_output.toPlainText():
        raise AssertionError("普通人面板清空后预览未复位")
    if window.simple_field_table.rowCount() != 0:
        raise AssertionError("普通人面板清空后智能字段表未复位")
    if "暂无结果" not in window.simple_field_status_label.text() and "采到结果后" not in window.simple_field_status_label.text():
        raise AssertionError("普通人面板清空后字段状态未复位")
    if "1 输入：进行中" not in window.simple_step_labels[0].text():
        raise AssertionError("普通人面板清空后未回到输入步骤")
    simple_csv_path = os.path.join(data_dir, "simple_input.csv")
    with open(simple_csv_path, "w", encoding="utf-8") as f:
        f.write("name,price\n苹果,3\n")
    captured_simple_ai = []
    original_simple_run_ai_worker = window.run_ai_worker
    window.run_ai_worker = lambda action, payload=None: captured_simple_ai.append((action, payload or {}))
    try:
        window.simple_url_input.setPlainText(simple_csv_path)
        window.simple_goal_input.setPlainText("整理成表格")
        if not window.simple_prepare_and_start_collect():
            raise AssertionError("普通人面板未接受本地 CSV 文件")
    finally:
        window.run_ai_worker = original_simple_run_ai_worker
    if not captured_simple_ai or captured_simple_ai[0][0] != "extract_file":
        raise AssertionError("普通人一键采集未自动识别本地文件")
    if captured_simple_ai[0][1].get("file_path") != simple_csv_path:
        raise AssertionError("普通人一键采集未把本地文件交给后台文件提取")
    if window.simple_target_kind("https://example.com/report.pdf") != "media_url":
        raise AssertionError("普通人一键采集未识别 PDF/图片网址")
    if window.simple_target_kind("https://example.com/") != "web_url":
        raise AssertionError("普通人一键采集误判普通网页")
    window.on_ai_result("extract_file", {"columns": ["name", "price"], "rows": [["苹果", "3"]]})
    if "文件已转成表格" not in window.simple_status_label.text() or "3 导出：进行中" not in window.simple_step_labels[2].text():
        raise AssertionError("普通人面板未展示文件提取完成状态")
    window.records = []
    if not window.simple_auto_save_results():
        raise AssertionError("普通人面板文件表格自动保存失败")
    file_export_path = getattr(window, "last_simple_export_path", "")
    if not os.path.exists(file_export_path) or not file_export_path.endswith(".xlsx"):
        raise AssertionError("普通人面板文件表格未保存为 Excel")
    if window.simple_recent_files_table.rowCount() < 1:
        raise AssertionError("普通人首页未刷新文件转表格保存记录")
    window.simple_recent_files_table.selectRow(0)
    if not window.open_selected_simple_recent_file() or getattr(window, "last_simple_open_path", "") != os.path.abspath(file_export_path):
        raise AssertionError("普通人首页未能打开文件转表格 Excel")
    window.clear_current_results()
    window.set_expert_mode(True)
    visible_expert_tabs = [
        window.tabs.tabText(index)
        for index in range(window.tabs.count())
        if window.tabs.tabText(index) in getattr(window, "expert_tab_names", []) and window.tabs.isTabVisible(index)
    ]
    if not visible_expert_tabs:
        raise AssertionError("专家模式未显示专家页")
    if not window.show_main_tab("历史与监控"):
        raise AssertionError("专家模式下无法跳转到历史页")
    window.set_workspace_mode(True)
    visible_workspace_tabs = [
        window.tabs.tabText(index)
        for index in range(window.tabs.count())
        if window.tabs.tabText(index) in getattr(window, "expert_tab_names", []) and window.tabs.isTabVisible(index)
    ]
    if not visible_workspace_tabs:
        raise AssertionError("高级工作区模式未显示专家页")
    window.set_workspace_mode(False)
    window.set_expert_mode(False)
    if window.tabs.tabText(window.tabs.currentIndex()) != "一键采集":
        raise AssertionError("统一普通界面未保持在一键采集面板")
    self_test_stage("simple panel OK")
    if not os.path.normcase(runtime_startup_log_file()).startswith(os.path.normcase(data_dir) + os.sep):
        raise AssertionError("通用入口启动日志未使用通用数据目录")
    if not os.path.normcase(runtime_self_test_error_log_file()).startswith(os.path.normcase(data_dir) + os.sep):
        raise AssertionError("通用入口自检日志未使用通用数据目录")
    if window.template_combo.count() < 6:
        raise AssertionError("站点模板库未初始化")
    presets = scene_template_presets()
    if len(presets) < 9 or "房产房源页" not in presets or "本地服务页" not in presets:
        raise AssertionError("场景模板库不完整")
    preset_index = window.scene_preset_combo.findText("房产房源页")
    if preset_index < 0:
        raise AssertionError("场景模板下拉框未加载房产模板")
    window.scene_preset_combo.setCurrentIndex(preset_index)
    window.apply_scene_preset()
    if window.template_type_combo.currentData() != "real_estate":
        raise AssertionError("场景模板类型未套用")
    if window.field_table.rowCount() < 4 or not window.next_page_selector_input.text():
        raise AssertionError("场景模板字段或分页设置未套用")
    for attr_name in (
        "template_market_search_input",
        "template_market_category_combo",
        "template_market_recommend_label",
        "template_market_table",
        "template_market_install_button",
        "template_market_apply_button",
    ):
        if not hasattr(window, attr_name):
            raise AssertionError(f"模板市场组件缺失：{attr_name}")
    if window.template_market_table.rowCount() < 6:
        raise AssertionError("模板市场内置模板过少")
    window.template_market_search_input.setText("企业")
    if window.template_market_table.rowCount() < 1:
        raise AssertionError("模板市场搜索失败")
    market_text = "\n".join(
        window.template_market_table.item(row, 1).text()
        for row in range(window.template_market_table.rowCount())
        if window.template_market_table.item(row, 1)
    )
    if "企业黄页页" not in market_text:
        raise AssertionError("模板市场未搜索到企业黄页模板")
    window.template_market_table.selectRow(0)
    if not window.install_market_template(True):
        raise AssertionError("模板市场安装失败")
    if window.template_combo.currentText() != "企业黄页页":
        raise AssertionError("模板市场安装后未用于当前采集任务")
    if window.template_type_combo.currentData() != "company" or window.field_table.rowCount() < 3:
        raise AssertionError("模板市场安装后未加载模板编辑内容")
    window.template_market_search_input.clear()
    wizard_index = window.wizard_scene_combo.findText("电商商品页")
    if wizard_index < 0:
        raise AssertionError("向导场景下拉框未加载电商模板")
    window.wizard_scene_combo.setCurrentIndex(wizard_index)
    preset_index = window.scene_preset_combo.findText("电商商品页")
    if preset_index >= 0:
        window.scene_preset_combo.setCurrentIndex(preset_index)
    window.ai_url_input.setText("https://example.com/product/1")
    window.ai_prompt_input.clear()
    window.ai_next_page_selector_input.clear()
    window.ai_page_limit_input.setValue(3)
    window.ai_scroll_times_input.setValue(2)
    window.latest_preview_url = "https://example.com/product/1"
    window.latest_preview_html = """
    <html><head><title>Product list</title></head><body>
      <h1>商品列表</h1>
      <a class="card" href="/product/10001">商品 A 价格 19 元</a>
      <a class="card" href="/product/10002">商品 B 价格 29 元</a>
      <a class="next" rel="next" href="/page/2">下一页</a>
      <img src="/a.jpg" alt="商品图">
      <img src="/b.jpg" alt="商品图">
      <p>库存 30，SKU 可选，加入购物车</p>
    </body></html>
    """
    window.configure_collect_wizard()
    if window.url_input.toPlainText().strip() != "https://example.com/product/1":
        raise AssertionError("向导未同步采集网址")
    if window.template_combo.currentText() != "电商商品页" or window.template_type_combo.currentData() != "ecommerce":
        raise AssertionError("向导未套用并选中电商模板")
    if not window.latest_market_recommendations:
        raise AssertionError("向导未生成模板市场推荐")
    if "电商商品页" not in window.template_market_recommend_label.text():
        raise AssertionError("模板市场未展示向导推荐模板")
    selected_market_item = window.template_market_table.item(window.template_market_table.currentRow(), 1)
    if not selected_market_item or selected_market_item.text() != "电商商品页":
        raise AssertionError("模板市场未自动选中向导推荐模板")
    if window.template_market_search_input.text().strip() != "电商商品页":
        raise AssertionError("模板市场未按推荐模板自动筛选")
    if "商品标题" not in window.ai_prompt_input.toPlainText():
        raise AssertionError("向导未填入场景默认需求")
    if window.page_limit_input.value() < 2 or window.scroll_times_input.value() < 2:
        raise AssertionError("向导未自动推荐分页/滚动参数")
    if not window.next_page_selector_input.text():
        raise AssertionError("向导未同步下一页选择器")
    if window.subpage_limit_input.value() < 2 or not window.subpage_checkbox.isChecked():
        raise AssertionError("向导未根据列表页开启子页面抓取")
    if "列表页" not in window.ai_task_plan_label.text() or window.ai_task_plan_table.rowCount() < 5:
        raise AssertionError("向导未展示页面结构诊断计划")
    wizard_text = json.dumps(window.latest_wizard_analysis_rows, ensure_ascii=False)
    if "页面类型" not in wizard_text or "疑似详情链接" not in wizard_text or "模型用途" not in wizard_text:
        raise AssertionError("向导诊断结果未保存")
    if window.ai_use_case_combo.currentData() != "cheap_batch":
        raise AssertionError("列表页向导未自动选择便宜批量模型用途")
    if window.ai_provider_combo.currentData() != "qwen" or window.current_ai_model_text() != "qwen-flash":
        raise AssertionError("列表页向导未切换到批量推荐模型")
    verify_collect_wizard_smoke(window)
    for attr_name in (
        "detail_title_label",
        "detail_body_output",
        "image_scroll",
        "detail_link_table",
        "detail_table_view",
        "history_detail_body_output",
        "risk_check_button",
        "estimate_task_button",
        "risk_table",
        "risk_summary_label",
        "collect_progress_bar",
        "collect_progress_label",
        "result_status_label",
        "result_export_hint_label",
        "new_user_flow_label",
        "task_queue_table",
        "task_queue_status_filter",
        "task_queue_type_filter",
        "retry_incomplete_button",
        "retry_selected_queue_button",
        "view_queue_result_button",
        "copy_queue_error_button",
        "queue_summary_label",
        "queue_detail_title_label",
        "queue_detail_output",
        "failure_recovery_label",
        "enable_browser_recovery_button",
        "slow_down_recovery_button",
        "change_alert_table",
        "change_alert_status_label",
        "change_alert_filter_combo",
        "refresh_change_alerts_button",
        "export_change_alerts_button",
        "change_alert_mark_handled_button",
        "change_alert_mark_ignored_button",
        "change_alert_mark_unread_button",
        "change_report_table",
        "generate_change_report_button",
        "export_change_report_button",
        "history_sections_tabs",
        "schedule_table",
        "add_schedule_button",
        "toggle_schedule_button",
        "run_schedule_now_button",
        "delete_schedule_button",
        "run_table",
        "export_runs_button",
        "reuse_run_button",
        "rerun_task_button",
        "resume_run_queue_button",
        "run_detail_title_label",
        "run_detail_summary_output",
        "run_detail_url_table",
        "run_detail_risk_table",
        "run_detail_queue_table",
        "run_detail_result_table",
        "export_run_results_button",
        "view_run_queue_result_button",
        "run_detail_json_output",
        "scene_preset_combo",
        "apply_scene_preset_button",
        "wizard_scene_combo",
        "wizard_configure_button",
        "ai_two_click_prepare_button",
        "ai_two_click_start_button",
        "ai_auto_fix_preflight_button",
        "auto_fix_preflight_button",
        "tray_icon",
        "last_unread_alert_notice_key",
        "notification_events",
        "last_clipboard_text",
        "overview_unread_label",
        "overview_schedule_label",
        "overview_failed_label",
        "overview_records_label",
        "overview_status_label",
        "overview_run_table",
        "overview_record_table",
        "overview_new_collect_button",
        "overview_ai_button",
        "overview_alerts_button",
        "overview_schedule_button",
        "overview_failed_button",
        "overview_refresh_button",
        "copy_sheets_button",
        "copy_history_sheets_button",
    ):
        if not hasattr(window, attr_name):
            raise AssertionError(f"详情预览组件缺失：{attr_name}")
    if "变更提醒" not in [window.history_sections_tabs.tabText(index) for index in range(window.history_sections_tabs.count())]:
        raise AssertionError("历史与监控页未展示变更提醒")
    if window.history_sections_tabs.tabText(4) != "计划采集":
        raise AssertionError("历史与监控页未展示计划采集")
    if not window.show_history_section("计划采集"):
        raise AssertionError("历史页应能定位到计划采集子分区")
    window.refresh_overview()
    if "未读变更" not in window.overview_unread_label.text() or "计划采集" not in window.overview_schedule_label.text():
        raise AssertionError("监控概览指标未刷新")
    if not hasattr(window, "use_browser_checkbox") or not window.use_browser_checkbox.isChecked():
        raise AssertionError("动态网页采集默认未开启")
    if not hasattr(window, "keep_login_checkbox"):
        raise AssertionError("登录态动态采集开关未创建")
    if not hasattr(window, "skip_unchanged_checkbox") or not window.skip_unchanged_checkbox.isChecked():
        raise AssertionError("反重复默认开关未创建")
    window.url_input.setPlainText("https://example.com/a\nhttps://example.com/b")
    window.keep_login_checkbox.setChecked(True)
    window.delay_input.setValue(1)
    window.page_limit_input.setValue(30)
    window.subpage_checkbox.setChecked(True)
    window.subpage_limit_input.setValue(30)
    window.field_table.setRowCount(0)
    window.add_field_row(FieldRule("邮箱", "a[href^=mailto]", "href", True))
    risks = window.run_preflight_check()
    risk_text = json.dumps(risks, ensure_ascii=False)
    for expected in ("robots.txt", "登录态", "敏感字段", "采集规模"):
        if expected not in risk_text:
            raise AssertionError(f"抓取前风险检查未识别：{expected}")
    if window.risk_table.rowCount() < 4:
        raise AssertionError("抓取前风险表未显示检查结果")
    risk_summary = window.risk_summary_label.text()
    for expected in ("高风险", "robots.txt", "登录态", "敏感字段"):
        if expected not in risk_summary:
            raise AssertionError(f"抓取前风险摘要未展示：{expected}")
    if not window.auto_fix_before_start():
        raise AssertionError("开始前自动修复未处理高风险配置")
    if window.keep_login_checkbox.isChecked():
        raise AssertionError("开始前自动修复未关闭登录态")
    if window.delay_input.value() < 1:
        raise AssertionError("开始前自动修复未调高访问间隔")
    if window.page_limit_input.value() > 10 or window.subpage_limit_input.value() > 10:
        raise AssertionError("开始前自动修复未限制大批量范围")
    if window.ai_page_limit_input.value() != window.page_limit_input.value():
        raise AssertionError("开始前自动修复未同步 AI 分页页数")
    if window.task_queue_table.rowCount() < 1:
        raise AssertionError("开始前自动修复未刷新任务队列")
    fixed_risk_text = json.dumps(window.collect_preflight_risks(), ensure_ascii=False)
    for fixed_item in ("登录态", "访问频率", "采集规模", "动态浏览器"):
        if fixed_item in fixed_risk_text:
            raise AssertionError(f"开始前自动修复后仍残留可自动处理项：{fixed_item}")
    if "敏感字段" not in fixed_risk_text:
        raise AssertionError("开始前自动修复不应擅自删除敏感字段提醒")
    remaining_confirm = window.remaining_confirmation_risks(window.collect_preflight_risks())
    remaining_text = json.dumps(remaining_confirm, ensure_ascii=False)
    if "robots.txt" not in remaining_text or "敏感字段" not in remaining_text:
        raise AssertionError("开始前风险确认未保留必须人工确认的风险")
    if window.unconfirmed_preflight_risks(window.collect_preflight_risks()) != remaining_confirm:
        raise AssertionError("首次风险确认不应被记忆跳过")
    original_self_test_flag = os.environ.get("UNIVERSAL_COLLECTOR_SELF_TEST")
    original_question = QMessageBox.question
    start_attempts = []

    try:
        os.environ["UNIVERSAL_COLLECTOR_SELF_TEST"] = "0"
        QMessageBox.question = staticmethod(lambda *args, **kwargs: QMessageBox.StandardButton.No)
        window._self_test_start_hook = lambda urls, risks: start_attempts.append(list(urls))
        window.worker = None
        window.start_collecting()
        if start_attempts:
            raise AssertionError("用户取消风险确认后仍启动了采集")
        if "已取消" not in window.collect_progress_label.text():
            raise AssertionError("用户取消风险确认后未显示取消状态")
        QMessageBox.question = staticmethod(lambda *args, **kwargs: QMessageBox.StandardButton.Yes)
        window.worker = None
        window.start_collecting()
        if start_attempts != [window.urls_from_input()]:
            raise AssertionError("用户确认风险后未继续启动采集")
        remembered_states = load_risk_confirmations()
        if not remembered_states:
            raise AssertionError("用户确认风险后未保存确认记忆")
        remembered_risks = window.unconfirmed_preflight_risks(window.collect_preflight_risks())
        if remembered_risks:
            raise AssertionError("同站点同风险确认后仍重复要求确认")
        question_calls = []
        QMessageBox.question = staticmethod(lambda *args, **kwargs: question_calls.append(args) or QMessageBox.StandardButton.No)
        window.worker = None
        window.start_collecting()
        if question_calls:
            raise AssertionError("同站点同风险记忆有效时仍弹出确认")
        if start_attempts[-1] != window.urls_from_input():
            raise AssertionError("风险确认记忆有效时未继续启动采集")
        window.add_field_row(FieldRule("手机号", ".phone", "text", True))
        changed_unconfirmed = window.unconfirmed_preflight_risks(window.collect_preflight_risks())
        if not changed_unconfirmed or "手机号" not in json.dumps(changed_unconfirmed, ensure_ascii=False):
            raise AssertionError("敏感字段变化后未重新要求确认")
    finally:
        if hasattr(window, "_self_test_start_hook"):
            delattr(window, "_self_test_start_hook")
        QMessageBox.question = original_question
        if original_self_test_flag is None:
            os.environ.pop("UNIVERSAL_COLLECTOR_SELF_TEST", None)
        else:
            os.environ["UNIVERSAL_COLLECTOR_SELF_TEST"] = original_self_test_flag
    window.keep_login_checkbox.setChecked(False)
    window.use_browser_checkbox.setChecked(True)
    window.delay_input.setValue(1)
    window.page_limit_input.setValue(DEFAULT_PAGE_LIMIT)
    window.subpage_checkbox.setChecked(False)
    window.subpage_limit_input.setValue(0)
    for attr_name in (
        "ai_provider_combo",
        "ai_config_box",
        "ai_workspace_box",
        "ai_quick_scrape_box",
        "ai_depth_scrape_box",
        "ai_agent_workspace_box",
        "ai_file_tools_box",
        "advanced_ai_boxes",
        "ai_model_combo",
        "ai_model_count_label",
        "ai_model_search_input",
        "ai_suggest_table",
        "ai_diagnose_button",
        "ai_apply_preset_button",
        "ai_diagnosis_table",
        "ai_call_log_table",
        "ai_call_summary_table",
        "ai_repair_history_table",
        "ai_repair_diff_table",
        "ai_refresh_call_logs_button",
        "ai_export_call_logs_button",
        "ai_refresh_call_summary_button",
        "ai_export_call_summary_button",
        "ai_refresh_repair_history_button",
        "ai_export_repair_history_button",
        "ai_apply_best_repair_history_button",
        "ai_apply_selected_repair_history_button",
        "ai_compare_selected_repair_history_button",
        "ai_apply_selected_repair_fields_button",
        "ai_clear_call_logs_button",
        "ai_task_plan_label",
        "ai_task_plan_table",
        "ai_apply_task_plan_button",
        "ai_run_task_plan_button",
        "ai_copy_task_plan_button",
        "pagination_table",
        "subpage_link_table",
        "ai_quality_table",
        "ai_quality_score_label",
        "result_quality_table",
        "result_quality_score_label",
        "result_quality_repair_button",
        "repair_quality_report_label",
        "repair_quality_report_table",
        "ai_apply_repaired_fields_button",
        "ai_provider_boundary_label",
        "api_health_label",
        "ai_apply_fields_button",
        "ai_base_url_input",
        "ai_key_input",
        "ai_key_name_input",
        "ai_key_combo",
        "ai_key_show_button",
        "ai_key_add_button",
        "ai_key_delete_button",
        "ai_key_use_available_button",
        "ai_provider_overview_table",
        "ai_provider_overview_refresh_button",
        "ai_provider_overview_switch_button",
        "ai_provider_models_refresh_button",
        "ai_provider_connectivity_test_button",
        "ai_suggest_button",
        "ai_preview_extract_button",
        "ai_repair_fields_button",
        "ai_auto_fix_preflight_button",
        "ai_export_table_button",
        "ai_copy_table_button",
        "preview_pagination_button",
        "apply_pagination_button",
        "ai_next_page_selector_input",
        "ai_page_limit_input",
        "ai_scroll_times_input",
        "scan_subpages_button",
        "apply_subpages_button",
        "ai_agent_button",
        "ai_file_button",
        "email_phone_button",
        "image_download_button",
        "schedule_button",
        "subpage_checkbox",
        "subpage_limit_input",
    ):
        if not hasattr(window, attr_name):
            raise AssertionError(f"AI/Thunderbit 风格组件缺失：{attr_name}")
    if window.ai_provider_combo.findText("Thunderbit 抽取接口（第三方接口）") < 0:
        raise AssertionError("Thunderbit 未以第三方抽取接口身份展示")
    if window.ai_quick_scrape_box.title() != "2 次点击网页转表格":
        raise AssertionError("AI 新手抓取入口未突出 2 次点击网页转表格")
    if window.ai_depth_scrape_box.title() != "高级设置：分页与子页面":
        raise AssertionError("分页与子页面未收纳到高级设置")
    if window.ai_agent_workspace_box.title() != "高级设置：网页自动化 Agent":
        raise AssertionError("网页 Agent 未收纳到高级设置")
    if window.ai_file_tools_box.title() != "高级设置：文件与线索工具":
        raise AssertionError("PDF/图片/邮箱电话工具未收纳到高级设置")
    if len(window.advanced_ai_boxes) != 3:
        raise AssertionError("高级设置分区数量错误")
    if not window.ai_depth_scrape_box.isCheckable() or window.ai_depth_scrape_box.isChecked():
        raise AssertionError("高级分页设置应默认折叠")
    if not window.ai_next_page_selector_input.isHidden():
        raise AssertionError("高级分页控件默认不应占用新手第一屏")
    window.ai_depth_scrape_box.setChecked(True)
    window.apply_advanced_ai_visibility()
    if window.ai_next_page_selector_input.isHidden():
        raise AssertionError("高级分页设置展开后不可见")
    window.ai_depth_scrape_box.setChecked(False)
    window.apply_advanced_ai_visibility()
    if window.ai_format_combo.findText("第三方抽取接口") < 0:
        raise AssertionError("第三方抽取接口格式未展示")
    if "模型 API 配置" in window.ai_config_box.title():
        raise AssertionError("AI 配置仍被旧文案描述为模型厂商配置")
    if "仅远程 API" not in window.ai_config_box.title() or "配置中心" not in window.ai_config_box.title():
        raise AssertionError("AI 配置区未明确为远程 API 配置中心")
    if not hasattr(window, "overview_product_boundary_label") or "所有网站" not in window.overview_product_boundary_label.text():
        raise AssertionError("总览页未说明主产品面向所有网站")
    thunderbit_index = window.ai_provider_combo.findData("thunderbit")
    if thunderbit_index < 0:
        raise AssertionError("Thunderbit 第三方抽取接口预设缺失")
    window.ai_provider_combo.setCurrentIndex(thunderbit_index)
    if "第三方抽取接口" not in window.ai_provider_boundary_label.text() or "不是通用大模型" not in window.ai_provider_boundary_label.text():
        raise AssertionError("Thunderbit 抽取接口未与大模型 API 明确区分")
    if window.ai_fetch_models_button.isEnabled() or window.ai_fetch_models_button.text() != "无需拉取模型":
        raise AssertionError("第三方抽取接口不应展示为可拉取模型")
    openai_index = window.ai_provider_combo.findData("openai")
    window.ai_provider_combo.setCurrentIndex(openai_index)
    if "模型 API" not in window.ai_provider_boundary_label.text() or not window.ai_fetch_models_button.isEnabled():
        raise AssertionError("模型 API 厂商切换后未恢复模型配置状态")
    lead_result = extract_emails_and_phones(
        [
            {
                "url": "https://example.com/contact",
                "title": "联系页",
                "body": "请联系 sales@example.com 或 138 0013 8000",
                "links": ["mailto:support@example.com"],
            }
        ]
    )
    if "sales@example.com" not in lead_result.get("emails", []) or not lead_result.get("rows"):
        raise AssertionError("邮箱/电话结构化提取失败")
    window.records = [
        {
            "url": "https://example.com/contact",
            "title": "联系页",
            "body": "请联系 sales@example.com 或 138 0013 8000",
            "links": ["mailto:support@example.com"],
        }
    ]
    window.extract_email_phone_current()
    lead_columns, lead_rows = window.ai_table_data()
    if lead_columns != ["内容", "类型", "来源标题", "来源网址"] or not any("sales@example.com" in row for row in lead_rows):
        raise AssertionError("邮箱/电话未进入结构化结果表")
    required_providers = (
        "openai",
        "deepseek",
        "anthropic",
        "gemini",
        "qwen",
        "hunyuan",
        "doubao",
        "kimi",
        "zhipu",
        "xai",
        "mistral",
        "groq",
        "together",
        "perplexity",
        "openrouter",
        "siliconflow",
        "thunderbit",
        "custom",
    )
    missing_providers = [provider for provider in required_providers if window.ai_provider_combo.findData(provider) < 0]
    if missing_providers:
        raise AssertionError(f"AI 服务商预设缺失：{missing_providers}")
    if len(AI_PROVIDER_PRESETS) < 16:
        raise AssertionError("AI 服务商数量过少")
    provider_overview_rows = ai_provider_runtime_overview(window.ai_settings)
    if len(provider_overview_rows) != len(AI_PROVIDER_PRESETS):
        raise AssertionError("AI 厂商运行总览未覆盖全部厂商")
    overview_openai = next((item for item in provider_overview_rows if item.get("provider") == "openai"), {})
    if overview_openai.get("model_count", 0) < 12 or overview_openai.get("config_status") not in {"正常", "需确认", "错误"}:
        raise AssertionError("AI 厂商运行总览未正确统计 OpenAI")
    window.refresh_ai_provider_overview()
    if window.ai_provider_overview_table.rowCount() != len(AI_PROVIDER_PRESETS):
        raise AssertionError("AI 厂商总览表行数不正确")
    silicon_row = -1
    for row in range(window.ai_provider_overview_table.rowCount()):
        marker = window.ai_provider_overview_table.item(row, 0)
        if marker and marker.data(Qt.ItemDataRole.UserRole) == "siliconflow":
            silicon_row = row
            break
    if silicon_row < 0:
        raise AssertionError("AI 厂商总览表缺少硅基流动")
    window.ai_provider_overview_table.selectRow(silicon_row)
    window.switch_ai_provider_from_overview()
    if window.ai_provider_combo.currentData() != "siliconflow":
        raise AssertionError("AI 厂商总览切换厂商失败")
    openai_index_for_overview = window.ai_provider_combo.findData("openai")
    window.ai_provider_combo.setCurrentIndex(openai_index_for_overview)
    unhealthy_presets = [
        item for item in ai_provider_preset_health()
        if item.get("status") == "错误" or (item.get("status") == "需补充" and item.get("provider") != "custom")
    ]
    if unhealthy_presets:
        raise AssertionError(f"AI 模型预设不健康：{unhealthy_presets}")
    openai_models = (window.ai_settings.get("providers") or {}).get("openai", {}).get("models", [])
    if len(openai_models) < 12 or AI_PROVIDER_PRESETS["openai"]["default_model"] not in openai_models:
        raise AssertionError("OpenAI 预设模型数量过少")
    qwen_models = (window.ai_settings.get("providers") or {}).get("qwen", {}).get("model_cache", [])
    if len(qwen_models) < 15 or "qwen-plus-latest" not in qwen_models:
        raise AssertionError("通义千问预设模型未扩展")
    for provider_key in ("deepseek", "doubao", "kimi", "openrouter", "siliconflow"):
        provider_models = (window.ai_settings.get("providers") or {}).get(provider_key, {}).get("model_cache", [])
        if len(provider_models) < 3:
            raise AssertionError(f"{provider_key} 预设模型数量过少")
    default_model_index = window.ai_model_combo.findData(AI_PROVIDER_PRESETS["openai"]["default_model"])
    if default_model_index < 0 or "[推荐]" not in window.ai_model_combo.itemText(default_model_index):
        raise AssertionError("默认模型未展示推荐标签")
    if "推荐" not in window.ai_model_hint_label.text():
        raise AssertionError("当前模型提示未展示标签")
    window.ai_model_search_input.setText("低价")
    if window.ai_model_combo.count() < 1:
        raise AssertionError("模型标签搜索失败")
    low_cost_model = window.current_ai_model_text()
    if not low_cost_model or "[" in low_cost_model:
        raise AssertionError("模型标签显示污染了真实模型名")
    window.save_ai_settings_from_ui()
    saved_tag_settings = load_ai_settings()
    saved_model = saved_tag_settings.get("providers", {}).get("openai", {}).get("model", "")
    if "[" in saved_model or "]" in saved_model:
        raise AssertionError("保存模型时不应写入标签文本")
    window.ai_model_search_input.clear()
    window.ai_key_input.setText("use-case-keep-key")
    window.ai_key_name_input.setText("用途测试 Key")
    expected_use_cases = {
        "web_scrape": ("openai", "gpt-5.2"),
        "vision_file": ("gemini", "gemini-2.5-flash"),
        "cheap_batch": ("qwen", "qwen-flash"),
        "strong_reasoning": ("deepseek", "deepseek-reasoner"),
    }
    for use_case_key, (expected_provider, expected_model) in expected_use_cases.items():
        use_case_index = window.ai_use_case_combo.findData(use_case_key)
        if use_case_index < 0:
            raise AssertionError(f"用途预设缺失：{use_case_key}")
        window.ai_use_case_combo.setCurrentIndex(use_case_index)
        window.apply_ai_use_case_preset()
        if window.ai_provider_combo.currentData() != expected_provider:
            raise AssertionError(f"用途预设未切换厂商：{use_case_key}")
        if window.current_ai_model_text() != expected_model:
            raise AssertionError(f"用途预设未切换模型：{use_case_key}")
        if window.ai_key_input.text().strip() != "use-case-keep-key":
            raise AssertionError("用途预设不应清空 API Key")
        use_case_settings = load_ai_settings()
        saved_provider = use_case_settings.get("provider")
        saved_provider_settings = (use_case_settings.get("providers") or {}).get(expected_provider, {})
        if saved_provider != expected_provider or saved_provider_settings.get("model") != expected_model:
            raise AssertionError(f"用途预设未保存：{use_case_key}")
        if "[" in saved_provider_settings.get("model", ""):
            raise AssertionError("用途预设保存了显示标签")
    openai_index_for_manual_lock = window.ai_provider_combo.findData("openai")
    if openai_index_for_manual_lock < 0:
        raise AssertionError("OpenAI 预设缺失")
    window.ai_provider_combo.setCurrentIndex(openai_index_for_manual_lock)
    window.refresh_ai_model_combo("manual-locked-model")
    window.ai_auto_apply_use_case_checkbox.setChecked(False)
    locked_plan = analyze_collect_task(
        "https://example.com/products",
        html=window.latest_preview_html,
        user_goal="抓取商品列表",
        preferred_scene="电商商品页",
    )
    if not window.apply_wizard_use_case(locked_plan):
        raise AssertionError("关闭自动切换后仍应记录向导用途建议")
    if window.ai_provider_combo.currentData() != "openai" or window.current_ai_model_text() != "manual-locked-model":
        raise AssertionError("向导用途建议不应覆盖手动锁定模型")
    window.save_ai_settings_from_ui()
    locked_settings = load_ai_settings()
    if locked_settings.get("providers", {}).get("openai", {}).get("auto_apply_use_case") is not False:
        raise AssertionError("向导自动切换模型开关未保存")
    window.ai_auto_apply_use_case_checkbox.setChecked(True)
    openai_index_after_use_case = window.ai_provider_combo.findData("openai")
    window.ai_provider_combo.setCurrentIndex(openai_index_after_use_case)
    window.ai_model_search_input.setText("gpt")
    if window.ai_model_combo.count() < 1:
        raise AssertionError("模型搜索过滤失败")
    if "/" not in window.ai_model_count_label.text():
        raise AssertionError("模型数量提示未展示搜索过滤状态")
    window.ai_model_search_input.clear()
    if "可选模型：" not in window.ai_model_count_label.text():
        raise AssertionError("模型数量提示缺失")
    window.ai_model_combo.setCurrentText("manual-self-test-model")
    window.save_ai_settings_from_ui()
    reloaded_settings = load_ai_settings()
    if "manual-self-test-model" not in reloaded_settings.get("providers", {}).get("openai", {}).get("model_cache", []):
        raise AssertionError("手动填写模型未写入缓存")
    bad_diagnosis = diagnose_ai_settings(
        {
            "provider": "openai",
            "provider_name": "OpenAI",
            "api_format": "openai_compatible",
            "base_url": "api.openai.com/v1",
            "models_url": "not-a-url",
            "model": "maybe-typo-model",
            "models": ["gpt-5-mini"],
            "model_cache": ["gpt-5-mini"],
            "api_key": "",
        }
    )
    if bad_diagnosis.get("ok") or not any(item.get("level") == "错误" for item in bad_diagnosis.get("checks", [])):
        raise AssertionError("AI 配置诊断未识别错误配置")
    window.fill_ai_diagnosis_table(bad_diagnosis.get("checks", []))
    window.refresh_api_health_summary(bad_diagnosis)
    if "API 体检：错误" not in window.api_health_label.text():
        raise AssertionError("API 体检总览未展示错误状态")
    if window.ai_diagnosis_table.rowCount() < 4:
        raise AssertionError("AI 配置诊断表未展示检查项")
    openai_index = window.ai_provider_combo.findData("openai")
    window.ai_provider_combo.setCurrentIndex(openai_index)
    window.ai_format_combo.setCurrentIndex(window.ai_format_combo.findData("gemini"))
    window.ai_base_url_input.setText("api.openai.com")
    window.ai_models_url_input.setText("bad-models-url")
    window.ai_model_combo.setCurrentText("wrong-model")
    window.ai_key_input.setText("keep-this-key")
    window.ai_key_name_input.setText("主 Key")
    window.add_or_update_ai_key()
    window.ai_key_name_input.setText("备用 Key")
    window.ai_key_input.setText("backup-key-123456")
    window.add_or_update_ai_key()
    if window.ai_key_combo.count() < 2:
        raise AssertionError("多 API Key 未进入下拉列表")
    main_index = window.ai_key_combo.findData("主 Key")
    if main_index < 0:
        raise AssertionError("命名 API Key 未保存")
    window.ai_key_combo.setCurrentIndex(main_index)
    if window.ai_key_input.text().strip() != "keep-this-key":
        raise AssertionError("API Key 切换未回填")
    window.ai_key_show_button.click()
    if window.ai_key_input.echoMode() != QLineEdit.EchoMode.Normal:
        raise AssertionError("API Key 显示按钮失效")
    window.ai_key_show_button.click()
    if window.ai_key_input.echoMode() != QLineEdit.EchoMode.Password:
        raise AssertionError("API Key 隐藏按钮失效")
    window.on_ai_result("test_api", {"ok": True, "message": "连接成功"})
    available_settings = load_ai_settings()
    available_keys = available_settings.get("providers", {}).get("openai", {}).get("api_keys", [])
    main_key_state = next((item for item in available_keys if item.get("name") == "主 Key"), {})
    if main_key_state.get("status") != "可用" or not main_key_state.get("last_tested_at"):
        raise AssertionError("API Key 成功状态未保存")
    if "可用" not in window.ai_key_combo.currentText():
        raise AssertionError("API Key 下拉未显示可用状态")
    backup_index = window.ai_key_combo.findData("备用 Key")
    if backup_index < 0:
        raise AssertionError("备用 API Key 未保留")
    window.ai_key_combo.setCurrentIndex(backup_index)
    window.update_current_ai_key_status("失败", "HTTP 401 invalid key")
    failed_settings = load_ai_settings()
    failed_keys = failed_settings.get("providers", {}).get("openai", {}).get("api_keys", [])
    backup_state = next((item for item in failed_keys if item.get("name") == "备用 Key"), {})
    if backup_state.get("status") != "失败" or "401" not in backup_state.get("last_error", ""):
        raise AssertionError("API Key 失败状态未保存")
    if "备用 Key" not in window.api_health_label.text() or "失败" not in window.api_health_label.text():
        raise AssertionError("API 体检总览未展示失败 Key")
    window.switch_to_available_ai_key()
    if window.ai_key_name_input.text().strip() != "主 Key" or window.ai_key_input.text().strip() != "keep-this-key":
        raise AssertionError("未自动切换到可用 API Key")
    if "主 Key" not in window.api_health_label.text() or "可用" not in window.api_health_label.text():
        raise AssertionError("API 体检总览未展示已切换的可用 Key")
    window.apply_recommended_ai_settings()
    if window.ai_format_combo.currentData() != "openai_compatible":
        raise AssertionError("一键修复未恢复接口格式")
    if window.ai_base_url_input.text().strip() != AI_PROVIDER_PRESETS["openai"]["base_url"]:
        raise AssertionError("一键修复未恢复 Base URL")
    if window.ai_models_url_input.text().strip() != AI_PROVIDER_PRESETS["openai"]["models_url"]:
        raise AssertionError("一键修复未恢复模型列表 URL")
    if window.current_ai_model_text() != AI_PROVIDER_PRESETS["openai"]["default_model"]:
        raise AssertionError("一键修复未恢复默认模型")
    if window.ai_key_input.text().strip() != "keep-this-key":
        raise AssertionError("一键修复不应清空 API Key")
    fixed_settings = load_ai_settings()
    fixed_openai = fixed_settings.get("providers", {}).get("openai", {})
    if fixed_openai.get("api_key") != "keep-this-key" or fixed_openai.get("base_url") != AI_PROVIDER_PRESETS["openai"]["base_url"]:
        raise AssertionError("一键修复未保存到用户设置")
    if "API 体检：正常" not in window.api_health_label.text():
        raise AssertionError("API 体检总览未展示修复后的正常状态")
    if len(fixed_openai.get("api_keys", [])) < 2:
        raise AssertionError("一键修复不应丢失多 API Key")
    backup_index = window.ai_key_combo.findData("备用 Key")
    if backup_index < 0:
        raise AssertionError("备用 API Key 未保留")
    window.ai_key_combo.setCurrentIndex(backup_index)
    window.delete_current_ai_key()
    after_delete_settings = load_ai_settings()
    after_delete_keys = after_delete_settings.get("providers", {}).get("openai", {}).get("api_keys", [])
    if any(item.get("name") == "备用 Key" for item in after_delete_keys):
        raise AssertionError("删除 API Key 未落盘")
    if window.ai_diagnosis_table.rowCount() < 6:
        raise AssertionError("一键修复后未刷新诊断表")
    old_ai_settings_file = AI_SETTINGS_FILE
    migration_settings_path = os.path.join(os.getcwd(), "self_test_runtime", "legacy_ai_settings_models.json")
    with open(migration_settings_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "provider": "openai",
                "providers": {
                    "openai": {
                        "provider": "openai",
                        "provider_name": "OpenAI",
                        "api_format": "openai_compatible",
                        "base_url": "https://api.openai.com/v1",
                        "models_url": "https://api.openai.com/v1/models",
                        "model": "old-only-model",
                        "models": ["old-only-model"],
                        "model_cache": ["old-only-model"],
                        "api_key": "legacy-key-abcdef",
                    }
                },
            },
            f,
            ensure_ascii=False,
        )
    migrated_settings = load_ai_settings(migration_settings_path)
    migrated_openai_models = migrated_settings.get("providers", {}).get("openai", {}).get("model_cache", [])
    if "old-only-model" not in migrated_openai_models or "gpt-5-mini" not in migrated_openai_models:
        raise AssertionError("旧 AI 配置未自动合并新版内置模型")
    migrated_openai_keys = migrated_settings.get("providers", {}).get("openai", {}).get("api_keys", [])
    if not migrated_openai_keys or migrated_openai_keys[0].get("key") != "legacy-key-abcdef":
        raise AssertionError("旧单 API Key 未迁移为命名 Key")
    window.click_tag_input.setText("h1")
    window.click_id_input.setText("main-title")
    window.generate_selector_from_helper()
    if "h1#main-title" not in window.field_table.item(window.field_table.rowCount() - 1, 1).text():
        raise AssertionError("可视化字段选择辅助失败")
    template = SiteTemplate(
        name="自检商品模板",
        field_rules=[
            FieldRule("标题", "h1"),
            FieldRule("价格", ".price"),
            FieldRule("时间", "time"),
            FieldRule("作者", ".author"),
            FieldRule("图片", "img", "src", True),
            FieldRule("链接", "a", "href", True),
        ],
    )
    html = """
    <html>
      <head><title>备用标题</title></head>
      <body>
        <main>
          <h1 id="main-title">测试商品 A</h1>
          <div class="price">￥128.50</div>
          <time datetime="2026-06-08 22:30">2026-06-08</time>
          <span class="author">测试作者</span>
          <p>这是一段用于验证正文自动识别的内容，长度足够进入正文候选。</p>
          <img src="/a.jpg" alt="图一">
          <a href="/detail">详情链接</a>
          <table><tr><th>规格</th><td>蓝色</td></tr></table>
        </main>
      </body>
    </html>
    """
    result_ctx = verify_result_workflow(window, template, html, self_test_stage)
    record = result_ctx["record"]
    db = result_ctx["db"]
    first = result_ctx["first"]
    duplicate = result_ctx["duplicate"]
    second = result_ctx["second"]
    repair_payload = result_ctx["repair_payload"]
    verify_repair_history_flow(window, self_test_stage)
    sample_html = """
    <html><body>
      <h1 class="fixed-title">样本标题</h1>
      <span class="fixed-price">￥88.00</span>
      <article class="fixed-body">修复后短正文</article>
    </body></html>
    """
    sample_records = [
        {"url": "https://example.com/sample/1"},
        {"url": "https://example.com/sample/2"},
        {"url": "https://example.com/sample/3"},
    ]
    original_fetch_snapshot_html = window.fetch_snapshot_html
    original_records_for_sample = list(window.records)
    try:
        window.records = sample_records
        window.fetch_snapshot_html = lambda url: sample_html
        sample_rules = [
            FieldRule("标题", ".fixed-title"),
            FieldRule("价格", ".fixed-price"),
            FieldRule("正文", ".fixed-body"),
        ]
        verified_samples = window.verify_repaired_fields_on_samples(sample_rules, limit=3)
    finally:
        window.records = original_records_for_sample
        window.fetch_snapshot_html = original_fetch_snapshot_html
    if len(verified_samples) != 3:
        raise AssertionError("AI 修复后多样本重采数量错误")
    sample_issues = window.analyze_repaired_sample_quality(verified_samples, ["标题", "价格", "正文"])
    if not sample_issues or any(issue.get("sample_count") != 3 for issue in sample_issues):
        raise AssertionError("AI 修复后多样本质量评分未记录样本数")
    sample_report_rows = window.update_repair_quality_report(repair_payload.get("quality_issues", []), sample_issues)
    if not sample_report_rows or "样本 3 条" not in window.repair_quality_report_label.text():
        raise AssertionError("AI 修复验证报告未展示多样本结果")
    failing_report_rows = window.build_repair_quality_report(
        [
            {"field": "价格", "status": "需处理", "score": 25, "problem": "空值率 100%", "advice": "修复价格"},
            {"field": "标题", "status": "需确认", "score": 65, "problem": "重复率 100%", "advice": "修复标题"},
        ],
        [
            {"field": "价格", "status": "需处理", "score": 10, "problem": "空值率 100%", "advice": "建议换选择器", "sample_count": 3},
            {"field": "标题", "status": "需确认", "score": 65, "problem": "重复率 100%", "advice": "建议缩小选择器", "sample_count": 3},
        ],
    )
    secondary_issues = window.prepare_secondary_repair_prompt(failing_report_rows)
    if len(secondary_issues) != 2 or "继续修复上一轮没有变好的字段" not in window.ai_prompt_input.toPlainText():
        raise AssertionError("AI 修复失败后未自动生成二次提示")
    captured_secondary_repair = []
    original_run_ai_worker_for_secondary = window.run_ai_worker
    window.latest_preview_rules = [FieldRule("价格", ".bad-price"), FieldRule("标题", ".bad-title")]
    window.latest_preview_html = html
    window.latest_preview_url = "https://example.com/item"
    window.run_ai_worker = lambda action, payload=None: captured_secondary_repair.append((action, payload or {}))
    try:
        window.ai_repair_problem_fields()
    finally:
        window.run_ai_worker = original_run_ai_worker_for_secondary
    if not captured_secondary_repair or captured_secondary_repair[0][0] != "repair_fields":
        raise AssertionError("二次提示未触发 AI 修复任务")
    secondary_payload_fields = {item.get("field") for item in captured_secondary_repair[0][1].get("quality_issues", [])}
    if secondary_payload_fields != {"价格", "标题"}:
        raise AssertionError("二次 AI 修复任务未使用失败字段")
    window.auto_apply_repair_after_ai = False
    window.fill_result_quality_table([])
    if window.result_quality_table.rowCount() != 0 or "等待结果" not in window.result_quality_score_label.text():
        raise AssertionError("采集结果质量总览清空后未恢复等待状态")
    window.result_table.selectRow(3)
    window.update_current_detail()
    if "状态：错误" not in window.detail_meta_label.text():
        raise AssertionError("结果详情未展示状态")
    report_rows = db.change_report(20)
    if not report_rows or not any(row.get("字段") == "价格" and "128.50" in row.get("旧值", "") and "99.00" in row.get("新值", "") for row in report_rows):
        raise AssertionError("网页监控变更报告未识别价格变化")
    window.change_report_rows = report_rows
    window.fill_change_report_table(report_rows)
    if window.change_report_table.rowCount() < 1:
        raise AssertionError("变更报告表格未显示数据")
    alert_count = window.refresh_change_alerts(silent=True)
    if alert_count < 1 or window.change_alert_table.rowCount() < 1:
        raise AssertionError("变更提醒表格未显示数据")
    if "发现" not in window.change_alert_status_label.text():
        raise AssertionError("变更提醒状态未展示数量")
    first_alert_status = window.change_alert_table.item(0, 0)
    first_alert_type = window.change_alert_table.item(0, 1)
    if not first_alert_status or first_alert_status.text() != "未读":
        raise AssertionError("变更提醒默认处理状态错误")
    if not first_alert_type or first_alert_type.text() != "变化":
        raise AssertionError("变更提醒状态列错误")
    alert_text = "\n".join(
        window.change_alert_table.item(row, column).text()
        for row in range(window.change_alert_table.rowCount())
        for column in range(window.change_alert_table.columnCount())
        if window.change_alert_table.item(row, column)
    )
    if "价格" not in alert_text or "99.00" not in alert_text:
        raise AssertionError("变更提醒未突出价格变化")
    notice_before = len(window.notification_events)
    if not window.notify_unread_change_alerts(alert_count, window.change_alert_rows[0]):
        raise AssertionError("未读变更通知未触发")
    if len(window.notification_events) != notice_before + 1:
        raise AssertionError("未读变更通知未记录")
    if window.notify_unread_change_alerts(alert_count, window.change_alert_rows[0]):
        raise AssertionError("重复未读变更通知未去重")
    window.show_change_alerts_tab(unread_only=True)
    # show_change_alerts_tab 切换的是历史子页，主 tab 不变
    if not window.show_history_section("变更提醒"):
        raise AssertionError("变更通知应能定位到变更提醒子页")
    if window.change_alert_filter_combo.currentText() != "未读":
        raise AssertionError("变更通知未切换未读筛选")
    window.change_alert_filter_combo.setCurrentText("全部提醒")
    window.change_alert_table.selectRow(0)
    selected_alert_id = window.selected_change_alert_id()
    if not selected_alert_id:
        raise AssertionError("变更提醒未生成稳定 ID")
    window.set_selected_change_alert_status("已处理")
    if load_change_alert_states().get(selected_alert_id, {}).get("status") != "已处理":
        raise AssertionError("变更提醒处理状态未保存")
    window.refresh_change_alerts(silent=True)
    handled_row = next((item for item in window.change_alert_rows if item.get("ID") == selected_alert_id), {})
    if handled_row.get("处理状态") != "已处理":
        raise AssertionError("变更提醒处理状态刷新后丢失")
    window.change_alert_filter_combo.setCurrentText("已处理")
    if window.change_alert_table.rowCount() != 1:
        raise AssertionError("变更提醒已处理筛选失败")
    window.change_alert_table.selectRow(0)
    window.set_selected_change_alert_status("忽略")
    ignored_row = next((item for item in window.change_alert_rows if item.get("ID") == selected_alert_id), {})
    if ignored_row.get("处理状态") != "忽略":
        raise AssertionError("变更提醒忽略状态未保存")
    window.change_alert_filter_combo.setCurrentText("忽略")
    if window.change_alert_table.rowCount() != 1:
        raise AssertionError("变更提醒忽略筛选失败")
    window.change_alert_table.selectRow(0)
    window.set_selected_change_alert_status("未读")
    unread_row = next((item for item in window.change_alert_rows if item.get("ID") == selected_alert_id), {})
    if unread_row.get("处理状态") != "未读":
        raise AssertionError("变更提醒标回未读失败")
    window.change_alert_filter_combo.setCurrentText("全部提醒")
    change_alert_json = os.path.join(data_dir, "change_alerts.json")
    window.export_change_alerts_to_file(change_alert_json)
    with open(change_alert_json, "r", encoding="utf-8") as f:
        alert_payload = json.load(f)
    alert_payload_text = json.dumps(alert_payload, ensure_ascii=False)
    if "处理状态" not in alert_payload_text or "未读" not in alert_payload_text or "变化" not in alert_payload_text or "价格" not in alert_payload_text or "99.00" not in alert_payload_text:
        raise AssertionError("变更提醒导出内容错误")
    change_report_json = os.path.join(data_dir, "change_report.json")
    export_table_data(
        change_report_json,
        ["监控时间", "网址", "域名", "字段", "旧值", "新值", "标题"],
        [[item.get(column, "") for column in ["监控时间", "网址", "域名", "字段", "旧值", "新值", "标题"]] for item in report_rows],
        sheet_name="变更报告",
    )
    with open(change_report_json, "r", encoding="utf-8") as f:
        report_payload = json.load(f)
    if "价格" not in json.dumps(report_payload, ensure_ascii=False):
        raise AssertionError("变更报告导出内容错误")
    formula_record = dict(record)
    formula_record["title"] = "=cmd|' /C calc'!A0"
    formula_record["body"] = "+SUM(1,2)"
    formula_csv = os.path.join(data_dir, "formula_export.csv")
    formula_xlsx = os.path.join(data_dir, "formula_export.xlsx")
    export_records(formula_csv, [formula_record])
    export_records(formula_xlsx, [formula_record])
    with open(formula_csv, "r", encoding="utf-8-sig") as f:
        formula_csv_text = f.read()
    if "'=cmd|" not in formula_csv_text or "'+SUM" not in formula_csv_text:
        raise AssertionError("CSV 导出未转义公式前缀")
    from openpyxl import load_workbook as load_formula_workbook

    formula_workbook = load_formula_workbook(formula_xlsx, data_only=False)
    formula_sheet = formula_workbook["采集结果"]
    if not str(formula_sheet["E2"].value).startswith("'=") or not str(formula_sheet["I2"].value).startswith("'+"):
        raise AssertionError("XLSX 导出未转义公式前缀")
    formula_workbook.close()
    window.url_input.setPlainText("https://example.com/item\nhttps://example.com/list")
    window.subpage_checkbox.setChecked(True)
    window.subpage_limit_input.setValue(3)
    window.selected_subpage_urls = ["https://example.com/detail/1"]
    window.ai_settings = save_ai_settings(
        {
            "provider": "custom",
            "provider_name": "任务档案自检 API",
            "api_format": "openai_compatible",
            "base_url": "https://api.example.test/v1",
            "models_url": "https://api.example.test/v1/models",
            "model": "archive-self-test-model",
            "api_key": "archive-test-key",
        },
        ai_settings_file,
    )
    with open(ai_settings_file, "r", encoding="utf-8") as f:
        raw_ai_settings_text = f.read()
    if os.name == "nt" and "archive-test-key" in raw_ai_settings_text:
        raise AssertionError("AI Key 不应以明文保存到配置文件")
    if load_ai_settings(ai_settings_file).get("api_key") != "archive-test-key":
        raise AssertionError("AI Key 加密保存后未能正确解密读取")
    cleanup_browser_dir = os.path.join(data_dir, "browser-profile")
    os.makedirs(cleanup_browser_dir, exist_ok=True)
    with open(os.path.join(cleanup_browser_dir, "Login Data"), "w", encoding="utf-8") as f:
        f.write("login")
    window.database = None
    gc.collect()
    QApplication.processEvents()
    cleanup_result = cleanup_user_data(
        {
            "api_settings": True,
            "ai_logs": True,
            "history": True,
            "browser_profile": True,
            "templates": False,
        }
    )
    if cleanup_result.get("failed"):
        raise AssertionError(f"清理本机数据失败：{cleanup_result.get('failed')}")
    for removed_path in (ai_settings_file, db_file, ai_call_log_file, os.path.join(data_dir, "ai_repair_history.jsonl")):
        if os.path.exists(removed_path):
            raise AssertionError(f"清理本机数据后仍存在：{removed_path}")
    if os.path.exists(cleanup_browser_dir):
        raise AssertionError("清理本机数据后浏览器登录态目录仍存在")
    if not os.path.exists(template_file):
        raise AssertionError("清理本机数据不应删除模板库")
    window.database = CollectorDatabase(db_file)
    window.ai_settings = save_ai_settings(
        {
            "provider": "custom",
            "provider_name": "任务档案自检 API",
            "api_format": "openai_compatible",
            "base_url": "https://api.example.test/v1",
            "models_url": "https://api.example.test/v1/models",
            "model": "archive-self-test-model",
            "api_key": "archive-test-key",
        },
        ai_settings_file,
    )
    db = window.database
    run_config = window.current_run_config(window.urls_from_input())
    run_risks = [{"级别": "需确认", "检查项": "自检", "说明": "任务档案风险快照"}]
    run_id = db.start_run(run_config, run_risks)
    run_record = dict(record)
    run_record["url"] = "https://example.com/item"
    run_record["title"] = "任务关联结果 A"
    run_record["fingerprint"] = ""
    run_record["run_id"] = run_id
    db.save_record(run_record, skip_unchanged=False)
    db.finish_run(run_id, status="finished", result_count=2, notes="任务档案自检完成")
    linked_records = db.records_for_run(run_id)
    if len(linked_records) != 1 or linked_records[0].get("title") != "任务关联结果 A":
        raise AssertionError("任务档案未能按 run_id 查询关联结果")
    runs = db.recent_runs(10)
    archive_run = next((item for item in runs if item.get("id") == run_id), None)
    if not archive_run:
        raise AssertionError("任务运行档案未入库")
    if archive_run.get("urls") != ["https://example.com/item", "https://example.com/list"]:
        raise AssertionError("任务运行档案未保存网址")
    if archive_run.get("template_name") != window.selected_template_name():
        raise AssertionError("任务运行档案未保存模板")
    if archive_run.get("ai_provider") != "custom" or archive_run.get("model") != "archive-self-test-model":
        raise AssertionError("任务运行档案未保存 AI 厂商/模型")
    if archive_run.get("config", {}).get("selected_subpage_urls") != ["https://example.com/detail/1"]:
        raise AssertionError("任务运行档案未保存子页面选择")
    archive_firecrawl = archive_run.get("config", {}).get("firecrawl", {})
    if not archive_firecrawl.get("enabled") or archive_firecrawl.get("base_url") != "https://api.firecrawl.dev":
        raise AssertionError(f"任务运行档案未保存 Firecrawl 配置摘要：{archive_firecrawl}")
    if (
        not archive_firecrawl.get("use_search")
        or not archive_firecrawl.get("use_extract")
        or not archive_firecrawl.get("use_batch")
        or not archive_firecrawl.get("use_crawl")
        or not archive_firecrawl.get("use_parse")
        or not archive_firecrawl.get("use_interact")
    ):
        raise AssertionError(f"任务运行档案未保存 Firecrawl 高级配置：{archive_firecrawl}")
    if archive_firecrawl.get("api_key") or "fc-ui-self-test" in json.dumps(archive_run.get("config", {}), ensure_ascii=False):
        raise AssertionError("任务运行档案不应保存 Firecrawl 明文 Key")
    if "任务档案风险快照" not in json.dumps(archive_run.get("risks", []), ensure_ascii=False):
        raise AssertionError("任务运行档案未保存风险快照")
    schedule_item = window.add_schedule_from_current_config(minutes=15)
    if not schedule_item or window.schedule_table.rowCount() < 1:
        raise AssertionError("计划采集未加入任务列表")
    saved_schedule_rows = load_schedules()
    saved_schedule_item = next((item for item in saved_schedule_rows if item.get("id") == schedule_item.get("id")), None)
    if not saved_schedule_item or saved_schedule_item.get("interval_minutes") != 15:
        raise AssertionError("计划采集未持久化保存")
    schedule_row = -1
    for row in range(window.schedule_table.rowCount()):
        item = window.schedule_table.item(row, 8)
        if item and item.text() == schedule_item.get("id"):
            schedule_row = row
            break
    if schedule_row < 0:
        raise AssertionError("计划采集未显示在后台任务表")
    window.schedule_table.selectRow(schedule_row)
    window.toggle_selected_schedule()
    saved_schedule_item = next((item for item in load_schedules() if item.get("id") == schedule_item.get("id")), None)
    if saved_schedule_item.get("enabled"):
        raise AssertionError("计划采集停用失败")
    window.schedule_table.selectRow(schedule_row)
    window.toggle_selected_schedule()
    saved_schedule_item = next((item for item in load_schedules() if item.get("id") == schedule_item.get("id")), None)
    if not saved_schedule_item.get("enabled"):
        raise AssertionError("计划采集启用失败")
    schedule_runs = []
    original_start_collecting_for_schedule = window.start_collecting
    window.start_collecting = lambda: schedule_runs.append(window.urls_from_input())
    window.firecrawl_enabled_checkbox.setChecked(False)
    window.firecrawl_base_url_input.setText("https://changed.invalid")
    window.firecrawl_map_checkbox.setChecked(False)
    window.firecrawl_search_checkbox.setChecked(False)
    window.firecrawl_search_query_input.setText("")
    window.firecrawl_search_limit_input.setValue(1)
    window.firecrawl_extract_checkbox.setChecked(False)
    window.firecrawl_extract_prompt_input.setText("")
    window.firecrawl_batch_checkbox.setChecked(False)
    window.firecrawl_batch_concurrency_input.setValue(1)
    window.firecrawl_crawl_checkbox.setChecked(False)
    window.firecrawl_crawl_limit_input.setValue(1)
    window.firecrawl_crawl_depth_input.setValue(1)
    window.firecrawl_parse_checkbox.setChecked(False)
    window.firecrawl_interact_checkbox.setChecked(False)
    window.firecrawl_interact_wait_input.setValue(0)
    window.firecrawl_interact_prompt_input.setText("")
    window.run_selected_schedule_now()
    window.start_collecting = original_start_collecting_for_schedule
    if schedule_runs != [["https://example.com/item", "https://example.com/list"]]:
        raise AssertionError("立即运行计划采集未回填当前任务配置")
    if (
        not window.firecrawl_enabled_checkbox.isChecked()
        or window.firecrawl_base_url_input.text() != "https://api.firecrawl.dev"
        or not window.firecrawl_map_checkbox.isChecked()
        or not window.firecrawl_search_checkbox.isChecked()
        or window.firecrawl_search_query_input.text() != "智能采集 自检"
        or window.firecrawl_search_limit_input.value() != 7
        or not window.firecrawl_extract_checkbox.isChecked()
        or "标题" not in window.firecrawl_extract_prompt_input.text()
        or not window.firecrawl_batch_checkbox.isChecked()
        or window.firecrawl_batch_concurrency_input.value() != 6
        or not window.firecrawl_crawl_checkbox.isChecked()
        or window.firecrawl_crawl_limit_input.value() != 12
        or window.firecrawl_crawl_depth_input.value() != 3
        or not window.firecrawl_parse_checkbox.isChecked()
        or not window.firecrawl_interact_checkbox.isChecked()
        or window.firecrawl_interact_wait_input.value() != 1500
        or "展开更多" not in window.firecrawl_interact_prompt_input.text()
    ):
        raise AssertionError("立即运行计划采集未回填 Firecrawl 配置")
    window.mark_schedule_run(schedule_item.get("id"), "完成", "自检计划完成", count_run=True)
    saved_schedule_item = next((item for item in load_schedules() if item.get("id") == schedule_item.get("id")), None)
    if saved_schedule_item.get("run_count") != 1:
        raise AssertionError("计划采集运行次数未更新")
    schedule_row = -1
    for row in range(window.schedule_table.rowCount()):
        item = window.schedule_table.item(row, 8)
        if item and item.text() == schedule_item.get("id"):
            schedule_row = row
            break
    if schedule_row < 0:
        raise AssertionError("计划采集删除前未找到自检计划")
    window.schedule_table.selectRow(schedule_row)
    window.delete_selected_schedule()
    if any(item.get("id") == schedule_item.get("id") for item in load_schedules()):
        raise AssertionError("计划采集删除失败")
    window.run_records = runs
    window.fill_run_table(runs)
    if window.run_table.rowCount() < 1 or "archive-self-test-model" not in window.run_table.item(0, 7).text():
        raise AssertionError("任务运行档案表格未显示数据")
    archive_row = -1
    for row in range(window.run_table.rowCount()):
        id_item = window.run_table.item(row, 0)
        if id_item and int(id_item.text()) == run_id:
            archive_row = row
            break
    if archive_row < 0:
        raise AssertionError("任务运行档案表格未找到自检记录")
    window.run_table.selectRow(archive_row)
    window.update_run_detail()
    if f"任务 #{run_id}" not in window.run_detail_title_label.text():
        raise AssertionError("任务档案详情标题未显示选中任务")
    if window.run_detail_url_table.rowCount() != 2:
        raise AssertionError("任务档案详情未显示网址列表")
    if window.run_detail_risk_table.rowCount() != 1:
        raise AssertionError("任务档案详情未显示风险检查")
    detail_text = window.run_detail_json_output.toPlainText()
    if "archive-self-test-model" not in detail_text or "https://example.com/detail/1" not in detail_text:
        raise AssertionError("任务档案详情未显示完整配置")
    if "真实浏览器" not in window.run_detail_summary_output.toPlainText():
        raise AssertionError("任务档案详情摘要未显示浏览器配置")
    if (
        "Firecrawl：开启" not in window.run_detail_summary_output.toPlainText()
        or "Search 开启" not in window.run_detail_summary_output.toPlainText()
        or "Crawl 开启" not in window.run_detail_summary_output.toPlainText()
        or "Interact 开启" not in window.run_detail_summary_output.toPlainText()
    ):
        raise AssertionError("任务档案详情摘要未显示 Firecrawl 配置")
    if window.run_detail_result_table.rowCount() != 1:
        raise AssertionError("任务档案详情未显示本次采集结果")
    if "已关联结果：1 条" not in window.run_detail_summary_output.toPlainText():
        raise AssertionError("任务档案详情摘要未显示关联结果数量")
    run_results_xlsx = os.path.join(data_dir, "run_results_export.xlsx")
    run_results_csv = os.path.join(data_dir, "run_results_export.csv")
    export_records(run_results_xlsx, db.records_for_run(run_id))
    export_records(run_results_csv, db.records_for_run(run_id))
    for exported_path in (run_results_xlsx, run_results_csv):
        if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
            raise AssertionError(f"当前任务结果导出失败：{exported_path}")
    from openpyxl import load_workbook

    workbook = load_workbook(run_results_xlsx)
    if workbook.sheetnames[:3] != ["采集结果", "字段说明", "导出摘要"]:
        raise AssertionError("Excel 工作簿未按采集结果/字段说明/导出摘要组织")
    result_sheet = workbook["采集结果"]
    if result_sheet.freeze_panes != "A2" or not result_sheet.auto_filter.ref:
        raise AssertionError("Excel 工作簿未冻结表头或启用筛选")
    if workbook["字段说明"].cell(1, 1).value != "字段" or workbook["导出摘要"].cell(3, 2).value != 1:
        raise AssertionError("Excel 工作簿字段说明或摘要内容错误")
    with open(run_results_csv, "r", encoding="utf-8-sig") as f:
        if "任务关联结果 A" not in f.read():
            raise AssertionError("当前任务结果 CSV 导出内容错误")
    sheets_text = records_to_tsv(db.records_for_run(run_id))
    if not sheets_text.startswith("采集时间\t网址\t域名") or "任务关联结果 A" not in sheets_text:
        raise AssertionError("Sheets 复制文本格式错误")
    window.records = linked_records
    if not window.copy_records_to_sheets(window.records, "自检本次结果"):
        raise AssertionError("复制本次结果到 Sheets 失败")
    if "采集时间\t网址\t域名" not in window.last_clipboard_text or "任务关联结果 A" not in window.last_clipboard_text:
        raise AssertionError("复制到 Sheets 未写入剪贴板文本")
    if not window.copy_current_results_to_sheets():
        raise AssertionError("复制当前结果到 Sheets 失败")
    if "已复制" not in window.result_export_hint_label.text():
        raise AssertionError("复制当前结果后导出引导未反馈成功")
    window.update_collect_progress(
        {
            "processed": 1,
            "success": 1,
            "failed": 0,
            "total": 2,
            "current_url": "https://example.com/item",
            "status": "running",
        }
    )
    if window.collect_progress_bar.maximum() != 2 or window.collect_progress_bar.value() != 1:
        raise AssertionError("采集进度条未按任务进度更新")
    if "已处理 1/2" not in window.collect_progress_label.text() or "https://example.com/item" not in window.collect_progress_label.text():
        raise AssertionError("采集进度文字未显示处理数量和当前网址")
    window.update_collect_progress(
        {
            "processed": 1,
            "success": 1,
            "failed": 0,
            "total": 2,
            "current_url": "https://example.com/item",
            "stage": "采集页面",
            "status": "running",
        }
    )
    if "阶段：采集页面" not in window.collect_progress_label.text():
        raise AssertionError("采集进度文字未显示内核阶段")
    window.url_input.setPlainText("https://example.com/item\nhttps://example.com/list")
    window.page_limit_input.setValue(3)
    window.subpage_checkbox.setChecked(True)
    window.selected_subpage_urls = ["https://example.com/detail/1"]
    estimated_queue = window.estimated_task_queue(window.urls_from_input())
    if len(estimated_queue) != 8:
        raise AssertionError("任务预估队列数量错误")
    window.fill_task_queue_table(estimated_queue)
    if window.task_queue_table.rowCount() != 8:
        raise AssertionError("任务预估队列表未显示")
    window.update_task_queue_progress({"stage": "采集页面", "current_url": "https://example.com/item"})
    if window.task_queue_table.item(0, 0).text() != "运行中" or window.task_queue_table.item(0, 2).text() != "采集页面":
        raise AssertionError("运行队列未显示当前采集阶段")
    window.update_task_queue_progress({"stage": "页面完成", "current_url": "https://example.com/item"})
    if window.task_queue_table.item(0, 0).text() != "已完成":
        raise AssertionError("运行队列未显示完成状态")
    self_test_stage("queue workflow OK")
    queue_result_record = dict(record)
    queue_result_record["url"] = "https://example.com/item"
    queue_result_record["title"] = "队列关联结果"
    queue_result_record["error"] = ""
    window.add_record(queue_result_record)
    if not any(source.get("url") == "https://example.com/item" and int(source.get("result_count") or 0) >= 1 for source in window.task_queue_rows):
        raise AssertionError("当前队列未关联结果数量")
    window.task_queue_status_filter.setCurrentText("全部状态")
    window.task_queue_table.selectRow(0)
    window.view_selected_queue_result()
    selected_current = window.selected_record_from_table(window.result_table)
    if not selected_current or selected_current.get("url") != "https://example.com/item":
        raise AssertionError("当前队列查看结果未定位结果表")
    queue_error_record = dict(record)
    queue_error_record["url"] = "https://example.com/list"
    queue_error_record["title"] = ""
    queue_error_record["error"] = "Timeout 30000ms exceeded"
    window.add_record(queue_error_record)
    failed_source = next((source for source in window.task_queue_rows if source.get("url") == "https://example.com/list"), {})
    if failed_source.get("error_category") != "网络超时" or "访问间隔" not in failed_source.get("error_advice", ""):
        raise AssertionError("当前队列未显示错误分类和修复建议")
    window.update_task_queue_progress({"stage": "页面完成", "current_url": "https://example.com/list", "failed_item": True})
    window.task_queue_status_filter.setCurrentText("失败")
    if window.task_queue_table.rowCount() != 1 or window.task_queue_table.item(0, 3).text() != "https://example.com/list":
        raise AssertionError("运行队列失败筛选未生效")
    window.task_queue_table.selectRow(0)
    window.update_queue_detail_panel()
    if "网络超时" not in window.queue_detail_output.toPlainText() or "访问间隔" not in window.queue_detail_output.toPlainText():
        raise AssertionError("运行队列详情未显示失败原因和建议")
    if "失败" not in window.queue_summary_label.text():
        raise AssertionError("运行队列摘要未显示失败数量")
    recovery_text = window.failure_recovery_label.text()
    for expected in ("失败自恢复", "网络超时", "重试失败项", "启用真实浏览器", "调低速度"):
        if expected not in recovery_text:
            raise AssertionError(f"失败自恢复面板缺失：{expected}")
    window.use_browser_checkbox.setChecked(False)
    if not window.enable_browser_recovery() or not window.use_browser_checkbox.isChecked():
        raise AssertionError("失败自恢复未启用真实浏览器")
    window.delay_input.setValue(0)
    if not window.slow_down_recovery() or window.delay_input.value() < 3:
        raise AssertionError("失败自恢复未调低访问速度")
    window.copy_selected_queue_error()
    clipboard_text = getattr(window, "last_clipboard_text", "") or QApplication.clipboard().text()
    if "https://example.com/list" not in clipboard_text or "Timeout 30000ms exceeded" not in clipboard_text:
        raise AssertionError("运行队列错误复制内容不完整")
    selected_retry_calls = []
    original_start_collecting_for_selected_queue = window.start_collecting
    window.start_collecting = lambda: selected_retry_calls.append(window.urls_from_input())
    window.retry_selected_queue_item()
    window.start_collecting = original_start_collecting_for_selected_queue
    if selected_retry_calls != [["https://example.com/list"]]:
        raise AssertionError("选中队列项重试未回填并触发采集")
    window.task_queue_status_filter.setCurrentText("未完成")
    if window.task_queue_table.rowCount() < 1:
        raise AssertionError("运行队列未完成筛选未生效")
    retry_calls = []
    original_start_collecting_for_queue = window.start_collecting
    window.start_collecting = lambda: retry_calls.append(window.urls_from_input())
    window.retry_incomplete_queue_items()
    window.start_collecting = original_start_collecting_for_queue
    if retry_calls != [["https://example.com/list"]]:
        raise AssertionError("失败/未完成队列重试未回填并触发采集")
    window.task_queue_status_filter.setCurrentText("全部状态")
    queue_run_id = db.start_run(run_config, run_risks)
    window.current_run_id = queue_run_id
    window.persist_current_run_queue_snapshot()
    queue_archive_record = dict(record)
    queue_archive_record["url"] = "https://example.com/item"
    queue_archive_record["title"] = "队列档案关联结果"
    queue_archive_record["fingerprint"] = ""
    queue_archive_record["run_id"] = queue_run_id
    db.save_record(queue_archive_record, skip_unchanged=False)
    queue_run = next((item for item in db.recent_runs(20) if item.get("id") == queue_run_id), None)
    queue_snapshot = (queue_run.get("config") or {}).get("task_queue_snapshot") if queue_run else []
    if not queue_snapshot or not any(item.get("status") == "失败" and item.get("url") == "https://example.com/list" for item in queue_snapshot):
        raise AssertionError("任务队列快照未写入任务档案")
    window.run_records = db.recent_runs(20)
    window.fill_run_table(window.run_records)
    queue_row = -1
    for row in range(window.run_table.rowCount()):
        id_item = window.run_table.item(row, 0)
        if id_item and int(id_item.text()) == queue_run_id:
            queue_row = row
            break
    if queue_row < 0:
        raise AssertionError("任务队列快照档案未显示在任务表")
    window.run_table.selectRow(queue_row)
    window.update_run_detail()
    if window.run_detail_queue_table.rowCount() != len(queue_snapshot):
        raise AssertionError("任务档案详情未显示队列快照")
    if window.run_detail_queue_table.columnCount() != 8:
        raise AssertionError("任务档案队列快照未显示结果数/错误列")
    if "task_queue_snapshot" not in window.run_detail_json_output.toPlainText():
        raise AssertionError("任务档案详情 JSON 未包含队列快照")
    if "队列快照" not in window.run_detail_summary_output.toPlainText():
        raise AssertionError("任务档案详情摘要未显示队列快照数量")
    resumable_urls = window.resumable_queue_urls_from_run(queue_run)
    if resumable_urls != ["https://example.com/list"]:
        raise AssertionError("任务档案未正确识别可继续队列网址")
    resume_calls = []
    original_start_collecting_for_resume = window.start_collecting
    window.start_collecting = lambda: resume_calls.append(window.urls_from_input())
    window.resume_selected_run_queue()
    window.start_collecting = original_start_collecting_for_resume
    if resume_calls != [["https://example.com/list"]]:
        raise AssertionError("任务档案继续运行未按队列失败项启动")
    if window.page_limit_input.value() != run_config.get("page_limit"):
        raise AssertionError("任务档案继续运行未复用原分页配置")
    window.run_table.selectRow(queue_row)
    window.update_run_detail()
    matched_queue_row = -1
    for row in range(window.run_detail_queue_table.rowCount()):
        url_item = window.run_detail_queue_table.item(row, 3)
        if url_item and url_item.text() == "https://example.com/item":
            matched_queue_row = row
            break
    if matched_queue_row < 0:
        raise AssertionError("任务档案队列快照未包含有结果的网址")
    result_count_item = window.run_detail_queue_table.item(matched_queue_row, 4)
    if not result_count_item or int(result_count_item.text() or "0") < 1:
        raise AssertionError("任务档案队列快照未关联结果数量")
    error_type_item = window.run_detail_queue_table.item(matched_queue_row, 5)
    advice_item = window.run_detail_queue_table.item(matched_queue_row, 6)
    if error_type_item is None or advice_item is None:
        raise AssertionError("任务档案队列快照未显示错误类型/建议列")
    failed_queue_row = -1
    for row in range(window.run_detail_queue_table.rowCount()):
        url_item = window.run_detail_queue_table.item(row, 3)
        if url_item and url_item.text() == "https://example.com/list":
            failed_queue_row = row
            break
    if failed_queue_row < 0:
        raise AssertionError("任务档案队列快照未包含失败网址")
    failed_category_item = window.run_detail_queue_table.item(failed_queue_row, 5)
    failed_advice_item = window.run_detail_queue_table.item(failed_queue_row, 6)
    if not failed_category_item or failed_category_item.text() != "网络超时" or "访问间隔" not in (failed_advice_item.text() if failed_advice_item else ""):
        raise AssertionError("任务档案队列快照未保留错误分类和修复建议")
    window.run_detail_queue_table.selectRow(matched_queue_row)
    window.view_selected_run_queue_result()
    selected_run_record = window.selected_record_from_table(window.run_detail_result_table)
    if not selected_run_record or selected_run_record.get("url") != "https://example.com/item":
        raise AssertionError("任务档案队列查看结果未定位结果表")
    window.current_run_id = None
    window.url_input.setPlainText("https://changed.example.test/")
    window.page_limit_input.setValue(1)
    window.scroll_times_input.setValue(0)
    window.delay_input.setValue(5)
    window.subpage_checkbox.setChecked(False)
    window.selected_subpage_urls = []
    window.firecrawl_enabled_checkbox.setChecked(False)
    window.firecrawl_base_url_input.setText("https://changed.invalid")
    window.firecrawl_map_checkbox.setChecked(False)
    window.firecrawl_search_checkbox.setChecked(False)
    window.firecrawl_search_query_input.setText("")
    window.firecrawl_search_limit_input.setValue(1)
    window.firecrawl_extract_checkbox.setChecked(False)
    window.firecrawl_extract_prompt_input.setText("")
    window.firecrawl_batch_checkbox.setChecked(False)
    window.firecrawl_batch_concurrency_input.setValue(1)
    window.firecrawl_crawl_checkbox.setChecked(False)
    window.firecrawl_crawl_limit_input.setValue(1)
    window.firecrawl_crawl_depth_input.setValue(1)
    window.firecrawl_parse_checkbox.setChecked(False)
    window.firecrawl_interact_checkbox.setChecked(False)
    window.firecrawl_interact_wait_input.setValue(0)
    window.firecrawl_interact_prompt_input.setText("")
    window.reuse_selected_run_config()
    if window.urls_from_input() != ["https://example.com/item", "https://example.com/list"]:
        raise AssertionError("任务档案复用未恢复网址")
    if window.page_limit_input.value() != run_config.get("page_limit"):
        raise AssertionError("任务档案复用未恢复翻页上限")
    if window.scroll_times_input.value() != run_config.get("scroll_times"):
        raise AssertionError("任务档案复用未恢复滚动次数")
    if not window.subpage_checkbox.isChecked() or window.selected_subpage_urls != ["https://example.com/detail/1"]:
        raise AssertionError("任务档案复用未恢复子页面设置")
    if (
        not window.firecrawl_enabled_checkbox.isChecked()
        or window.firecrawl_base_url_input.text() != "https://api.firecrawl.dev"
        or not window.firecrawl_map_checkbox.isChecked()
        or not window.firecrawl_search_checkbox.isChecked()
        or window.firecrawl_search_query_input.text() != "智能采集 自检"
        or window.firecrawl_search_limit_input.value() != 7
        or not window.firecrawl_extract_checkbox.isChecked()
        or "标题" not in window.firecrawl_extract_prompt_input.text()
        or not window.firecrawl_batch_checkbox.isChecked()
        or window.firecrawl_batch_concurrency_input.value() != 6
        or not window.firecrawl_crawl_checkbox.isChecked()
        or window.firecrawl_crawl_limit_input.value() != 12
        or window.firecrawl_crawl_depth_input.value() != 3
        or not window.firecrawl_parse_checkbox.isChecked()
        or not window.firecrawl_interact_checkbox.isChecked()
        or window.firecrawl_interact_wait_input.value() != 1500
        or "展开更多" not in window.firecrawl_interact_prompt_input.text()
    ):
        raise AssertionError("任务档案复用未恢复 Firecrawl 配置")
    if window.current_ai_model_text() != "archive-self-test-model":
        raise AssertionError("任务档案复用未恢复 AI 模型")
    rerun_calls = []
    original_start_collecting = window.start_collecting
    window.start_collecting = lambda: rerun_calls.append(window.urls_from_input())
    window.rerun_selected_task()
    window.start_collecting = original_start_collecting
    if rerun_calls != [["https://example.com/item", "https://example.com/list"]]:
        raise AssertionError("任务档案重跑未进入采集启动链路")
    stopped_run_id = db.start_run({**run_config, "urls": ["https://example.com/stopped"]}, run_risks)
    window.current_run_id = stopped_run_id
    window.current_run_start_count = len(window.records)
    window.current_run_progress = {"processed": 1, "success": 0, "failed": 1, "total": 1, "current_url": "https://example.com/stopped"}
    window.on_collect_finished({"status": "stopped", "emitted_count": 0, "notes": "用户停止采集，已返回结果 0 条。"})
    stopped_run = next((item for item in db.recent_runs(20) if item.get("id") == stopped_run_id), None)
    if not stopped_run or stopped_run.get("status") != "stopped" or "用户停止" not in stopped_run.get("notes", "") or "进度摘要" not in stopped_run.get("notes", ""):
        raise AssertionError("任务停止状态未写入档案")
    failed_run_id = db.start_run({**run_config, "urls": ["https://example.com/failed"]}, run_risks)
    window.current_run_id = failed_run_id
    window.current_run_start_count = len(window.records)
    window.on_collect_finished({"status": "failed", "emitted_count": 0, "notes": "采集异常，已返回结果 0 条。"})
    failed_run = next((item for item in db.recent_runs(20) if item.get("id") == failed_run_id), None)
    if not failed_run or failed_run.get("status") != "failed" or "采集异常" not in failed_run.get("notes", ""):
        raise AssertionError("任务失败状态未写入档案")
    partial_run_id = db.start_run({**run_config, "urls": ["https://example.com/partial"]}, run_risks)
    partial_record = dict(record)
    partial_record["url"] = "https://example.com/partial"
    partial_record["title"] = "部分成功结果"
    partial_record["fingerprint"] = ""
    partial_record["run_id"] = partial_run_id
    db.save_record(partial_record, skip_unchanged=False)
    window.add_record(partial_record)
    window.current_run_id = partial_run_id
    window.current_run_start_count = len(window.records) - 1
    window.on_collect_finished({"status": "partial", "emitted_count": 1, "notes": "采集异常，已返回结果 1 条。"})
    partial_run = next((item for item in db.recent_runs(20) if item.get("id") == partial_run_id), None)
    if not partial_run or partial_run.get("status") != "partial" or partial_run.get("result_count") != 1:
        raise AssertionError("任务部分成功状态未写入档案")
    window.run_records = db.recent_runs(20)
    window.fill_run_table(window.run_records)
    visible_statuses = [
        window.run_table.item(row, 3).text()
        for row in range(window.run_table.rowCount())
        if window.run_table.item(row, 3)
    ]
    for expected_status in ("已停止", "失败", "部分成功"):
        if expected_status not in visible_statuses:
            raise AssertionError(f"任务状态未显示中文：{expected_status}")
    run_archive_json = os.path.join(data_dir, "run_archive.json")
    export_table_data(
        run_archive_json,
        ["ID", "开始时间", "结束时间", "状态", "网址", "模板", "AI 厂商", "模型", "结果数", "配置快照", "风险检查"],
        [
            [
                item.get("id", ""),
                item.get("started_at", ""),
                item.get("finished_at", ""),
                item.get("status", ""),
                "\n".join(item.get("urls", []) or []),
                item.get("template_name", ""),
                item.get("ai_provider", ""),
                item.get("model", ""),
                item.get("result_count", 0),
                json.dumps(item.get("config", {}), ensure_ascii=False),
                json.dumps(item.get("risks", []), ensure_ascii=False),
            ]
            for item in runs
        ],
        sheet_name="任务运行档案",
    )
    with open(run_archive_json, "r", encoding="utf-8") as f:
        run_archive_payload = json.load(f)
    if "archive-self-test-model" not in json.dumps(run_archive_payload, ensure_ascii=False):
        raise AssertionError("任务运行档案导出内容错误")
    window.subpage_checkbox.setChecked(False)
    window.subpage_limit_input.setValue(0)
    window.selected_subpage_urls = []
    class FakeAIHandler(BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            return

        def do_GET(self):
            if self.path.endswith("/models"):
                payload = {"data": [{"id": "fake-model"}, {"id": "fake-vision"}]}
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps(payload).encode("utf-8"))
                return
            self.send_response(404)
            self.end_headers()

        def do_POST(self):
            length = int(self.headers.get("Content-Length", "0") or 0)
            raw = self.rfile.read(length).decode("utf-8", errors="replace")
            auth = self.headers.get("Authorization", "")
            if "bad-key" in auth:
                self.send_response(401)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"error": {"message": "invalid key"}}).encode("utf-8"))
                return
            if "建议" in raw or "fields" in raw:
                if "修复" in raw or "quality_issues" in raw:
                    content = {
                        "fields": [
                            {"name": "标题", "selector": "h1#main-title", "attr": "text", "multiple": False, "reason": "修复为更精确的主标题"},
                            {"name": "价格", "selector": "div.price", "attr": "text", "multiple": False, "reason": "修复为空值或重复问题"},
                        ]
                    }
                else:
                    content = {
                        "fields": [
                            {"name": "标题", "selector": "h1", "attr": "text", "multiple": False, "reason": "主标题"},
                            {"name": "价格", "selector": ".price", "attr": "text", "multiple": False, "reason": "价格区域"},
                        ]
                    }
            elif "连接测试" in raw:
                content = {"ok": True, "message": "连接成功"}
            elif "自然语言" in raw or "actions" in raw:
                content = {
                    "template": {
                        "name": "AI 自检模板",
                        "domain": "example.com",
                        "template_type": "ecommerce",
                        "next_page_selector": "a.next",
                        "field_rules": [{"name": "标题", "selector": "h1", "attr": "text", "multiple": False}],
                    },
                    "options": {"use_browser": True, "scroll_times": 1, "page_limit": 2, "subpage_limit": 1},
                    "actions": [{"type": "extract", "field_rules": [{"name": "标题", "selector": "h1", "attr": "text", "multiple": False}]}],
                }
            else:
                content = {"columns": ["字段", "值"], "rows": [["状态", "成功"]]}
            payload = {"choices": [{"message": {"content": json.dumps(content, ensure_ascii=False)}}]}
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(payload).encode("utf-8"))

    server = HTTPServer(("127.0.0.1", 0), FakeAIHandler)
    server_thread = threading.Thread(target=server.serve_forever, daemon=True)
    server_thread.start()
    try:
        fake_base = f"http://127.0.0.1:{server.server_address[1]}/v1"
        settings = save_ai_settings(
            {
                "provider": "custom",
                "provider_name": "自检假 API",
                "api_format": "openai_compatible",
                "base_url": fake_base,
                "models_url": f"{fake_base}/models",
                "model": "fake-model",
                "models": ["fake-model"],
                "api_key": "test-key",
            },
            ai_settings_file,
        )
        models = AIClient(settings).fetch_models()
        if "fake-model" not in models:
            raise AssertionError("AI 模型拉取失败")
        window.ai_settings = settings
        window.load_ai_settings_to_ui()
        window.on_ai_result("fetch_models", models)
        saved_custom_settings = load_ai_settings().get("providers", {}).get("custom", {})
        saved_models = saved_custom_settings.get("model_cache", [])
        if "fake-vision" not in saved_models:
            raise AssertionError("在线拉取模型未缓存到用户设置")
        if not saved_custom_settings.get("models_updated_at"):
            raise AssertionError("单厂商拉取模型未记录刷新时间")
        batch_seed_settings = load_ai_settings()
        batch_providers = batch_seed_settings.get("providers", {})
        batch_providers["openai"] = {
            **batch_providers.get("openai", {}),
            "provider": "openai",
            "provider_name": "OpenAI",
            "api_format": "openai_compatible",
            "base_url": fake_base,
            "models_url": f"{fake_base}/models",
            "model": "fake-model",
            "api_key": "test-key",
        }
        batch_providers["deepseek"] = {
            **batch_providers.get("deepseek", {}),
            "provider": "deepseek",
            "provider_name": "DeepSeek",
            "api_format": "openai_compatible",
            "base_url": fake_base,
            "models_url": f"{fake_base}/missing-models",
            "model": "deepseek-chat",
            "api_key": "test-key",
        }
        batch_seed_settings["providers"] = batch_providers
        batch_result = refresh_ai_provider_models(batch_seed_settings, ["openai", "deepseek", "thunderbit"], now_text="2026-06-09 09:00:00")
        batch_rows = batch_result.get("results", [])
        if not any(item.get("provider") == "openai" and item.get("status") == "成功" and item.get("model_count") == 2 for item in batch_rows):
            raise AssertionError("批量刷新模型未记录成功厂商")
        if not any(item.get("provider") == "deepseek" and item.get("status") == "失败" for item in batch_rows):
            raise AssertionError("批量刷新模型未记录失败厂商")
        if not any(item.get("provider") == "thunderbit" and item.get("status") == "跳过" for item in batch_rows):
            raise AssertionError("批量刷新模型未跳过第三方抽取接口")
        window.on_ai_result("refresh_provider_models", batch_result)
        refreshed_settings = load_ai_settings()
        refreshed_openai = refreshed_settings.get("providers", {}).get("openai", {})
        refreshed_deepseek = refreshed_settings.get("providers", {}).get("deepseek", {})
        if "fake-vision" not in refreshed_openai.get("model_cache", []) or refreshed_openai.get("models_updated_at") != "2026-06-09 09:00:00":
            raise AssertionError("批量刷新成功结果未保存模型和时间")
        if not refreshed_deepseek.get("models_refresh_error"):
            raise AssertionError("批量刷新失败原因未保存")
        window.refresh_ai_provider_overview()
        overview_text = "\n".join(
            window.ai_provider_overview_table.item(row, 7).text()
            for row in range(window.ai_provider_overview_table.rowCount())
            if window.ai_provider_overview_table.item(row, 7)
        )
        if "2026-06-09 09:00:00" not in overview_text or "失败：" not in overview_text:
            raise AssertionError("厂商总览未展示模型刷新时间或失败原因")
        connectivity_seed = load_ai_settings()
        connectivity_providers = connectivity_seed.get("providers", {})
        connectivity_providers["openai"] = {
            **connectivity_providers.get("openai", {}),
            "provider": "openai",
            "provider_name": "OpenAI",
            "api_format": "openai_compatible",
            "base_url": fake_base,
            "models_url": f"{fake_base}/models",
            "model": "fake-model",
            "api_key": "test-key",
            "active_api_key_name": "好 Key",
            "api_keys": [{"name": "好 Key", "key": "test-key"}],
        }
        connectivity_providers["deepseek"] = {
            **connectivity_providers.get("deepseek", {}),
            "provider": "deepseek",
            "provider_name": "DeepSeek",
            "api_format": "openai_compatible",
            "base_url": fake_base,
            "models_url": f"{fake_base}/models",
            "model": "fake-model",
            "api_key": "bad-key",
            "active_api_key_name": "坏 Key",
            "api_keys": [{"name": "坏 Key", "key": "bad-key"}],
        }
        connectivity_seed["providers"] = connectivity_providers
        connectivity_result = test_ai_provider_connectivity(
            connectivity_seed,
            ["openai", "deepseek", "thunderbit"],
            now_text="2026-06-09 09:30:00",
        )
        connectivity_rows = connectivity_result.get("results", [])
        if not any(item.get("provider") == "openai" and item.get("status") == "成功" for item in connectivity_rows):
            raise AssertionError("批量模型测试未记录成功厂商")
        if not any(item.get("provider") == "deepseek" and item.get("status") == "失败" and "invalid key" in item.get("message", "").lower() for item in connectivity_rows):
            raise AssertionError("批量模型测试未记录失败原因")
        if not any(item.get("provider") == "thunderbit" and item.get("status") == "跳过" for item in connectivity_rows):
            raise AssertionError("批量模型测试未跳过第三方抽取接口")
        window.on_ai_result("test_provider_connectivity", connectivity_result)
        connectivity_settings = load_ai_settings()
        tested_openai = connectivity_settings.get("providers", {}).get("openai", {})
        tested_deepseek = connectivity_settings.get("providers", {}).get("deepseek", {})
        if tested_openai.get("connection_status") != "成功" or tested_openai.get("connection_tested_at") != "2026-06-09 09:30:00":
            raise AssertionError("批量模型测试成功状态未保存")
        if tested_deepseek.get("connection_status") != "失败" or "invalid key" not in tested_deepseek.get("connection_error", "").lower():
            raise AssertionError("批量模型测试失败状态未保存")
        failed_key = next((item for item in tested_deepseek.get("api_keys", []) if item.get("name") == "坏 Key"), {})
        if failed_key.get("status") != "失败" or "invalid key" not in failed_key.get("last_error", "").lower():
            raise AssertionError("批量模型测试未同步 Key 失败状态")
        window.refresh_ai_provider_overview()
        connectivity_text = "\n".join(
            window.ai_provider_overview_table.item(row, 8).text()
            for row in range(window.ai_provider_overview_table.rowCount())
            if window.ai_provider_overview_table.item(row, 8)
        )
        if "2026-06-09 09:30:00" not in connectivity_text or "invalid key" not in connectivity_text.lower():
            raise AssertionError("厂商总览未展示连接测试时间或失败原因")
        test_result = AIClient(settings).test_connection()
        if not test_result.get("ok"):
            raise AssertionError("AI API 测试失败")
        self_test_stage("ai api connection OK")
        good_diagnosis = diagnose_ai_settings(settings)
        if not good_diagnosis.get("ok"):
            raise AssertionError("AI 配置诊断误判正常配置")
        window.on_ai_result("diagnose_api", good_diagnosis)
        if window.ai_diagnosis_table.rowCount() < 6:
            raise AssertionError("AI 配置诊断结果未进入表格")
        suggested = ai_suggest_fields("https://example.com/item", html, "建议列", settings)
        if not suggested.get("fields"):
            raise AssertionError("AI 建议列失败")
        retry_settings = dict(settings)
        retry_settings.update(
            {
                "api_key": "bad-key",
                "active_api_key_name": "坏 Key",
                "api_keys": [
                    {"name": "坏 Key", "key": "bad-key", "status": "失败", "last_error": "invalid key"},
                    {"name": "好 Key", "key": "test-key", "status": "可用", "last_tested_at": "2026-06-09 00:00:00"},
                ],
            }
        )
        retry_worker = AIWorker(
            "suggest_fields",
            retry_settings,
            {"url": "https://example.com/item", "html": html, "goal": "建议列"},
        )
        retry_result = retry_worker.run_with_key_retry()
        if retry_result.get("_auto_switched_key") != "好 Key" or not retry_result.get("fields"):
            raise AssertionError("AI 任务未自动切换可用 Key 重试")
        retry_worker.write_call_log("成功", result=retry_result, duration_ms=123)
        retry_worker_without_key = AIWorker(
            "suggest_fields",
            {**retry_settings, "api_keys": [{"name": "坏 Key", "key": "bad-key", "status": "失败"}]},
            {"url": "https://example.com/item", "html": html, "goal": "建议列"},
        )
        try:
            retry_worker_without_key.run_with_key_retry()
            raise AssertionError("无可用 Key 时不应重试成功")
        except RuntimeError as exc:
            retry_worker_without_key.write_call_log("失败", error_text=str(exc), duration_ms=45)
            if "401" not in str(exc) and "invalid key" not in str(exc).lower():
                raise
        ai_logs = load_ai_call_logs(20)
        if not any(item.get("status") == "成功" and item.get("auto_switched_key") == "好 Key" for item in ai_logs):
            raise AssertionError("AI 调用日志未记录自动换 Key 成功")
        if not any(item.get("status") == "失败" and "invalid key" in item.get("error", "").lower() for item in ai_logs):
            raise AssertionError("AI 调用日志未记录失败原因")
        window.fill_ai_call_log_table(ai_logs)
        if window.ai_call_log_table.rowCount() < 2:
            raise AssertionError("AI 调用日志未进入 UI 表格")
        summary_rows = summarize_ai_call_logs(ai_logs)
        summary_row = next(
            (
                item
                for item in summary_rows
                if item.get("model") == "fake-model"
                and item.get("key_name") == "好 Key"
            ),
            {},
        )
        if summary_row.get("total_calls") != 1 or summary_row.get("success_count") != 1:
            raise AssertionError("AI 用量汇总未统计成功调用")
        failed_summary_row = next(
            (
                item
                for item in summary_rows
                if item.get("key_name") == "坏 Key" and item.get("failure_count") == 1
            ),
            {},
        )
        if "invalid key" not in failed_summary_row.get("latest_error", "").lower():
            raise AssertionError("AI 用量汇总未保留最近失败原因")
        window.fill_ai_call_summary_table(summary_rows)
        if window.ai_call_summary_table.rowCount() < 2:
            raise AssertionError("AI 用量汇总未进入 UI 表格")
        ai_log_export_xlsx = os.path.join(data_dir, "ai_call_logs_export.xlsx")
        ai_log_export_csv = os.path.join(data_dir, "ai_call_logs_export.csv")
        ai_log_export_json = os.path.join(data_dir, "ai_call_logs_export.json")
        ai_summary_export_xlsx = os.path.join(data_dir, "ai_call_summary_export.xlsx")
        ai_summary_export_csv = os.path.join(data_dir, "ai_call_summary_export.csv")
        ai_summary_export_json = os.path.join(data_dir, "ai_call_summary_export.json")
        window.export_ai_call_logs_to_file(ai_log_export_xlsx)
        window.export_ai_call_logs_to_file(ai_log_export_csv)
        window.export_ai_call_logs_to_file(ai_log_export_json)
        window.export_ai_call_summary_to_file(ai_summary_export_xlsx)
        window.export_ai_call_summary_to_file(ai_summary_export_csv)
        window.export_ai_call_summary_to_file(ai_summary_export_json)
        for exported_path in (
            ai_log_export_xlsx,
            ai_log_export_csv,
            ai_log_export_json,
            ai_summary_export_xlsx,
            ai_summary_export_csv,
            ai_summary_export_json,
        ):
            if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
                raise AssertionError(f"AI 调用日志导出失败：{exported_path}")
        with open(ai_log_export_json, "r", encoding="utf-8") as f:
            exported_log_payload = json.load(f)
        if not exported_log_payload.get("rows"):
            raise AssertionError("AI 调用日志 JSON 导出内容为空")
        with open(ai_summary_export_json, "r", encoding="utf-8") as f:
            exported_summary_payload = json.load(f)
        if not any(row.get("Key", "").startswith("好 Key") for row in exported_summary_payload.get("rows", [])):
            raise AssertionError("AI 用量汇总 JSON 导出内容错误")
        window.clear_ai_call_logs_and_refresh()
        if load_ai_call_logs(20) or window.ai_call_log_table.rowCount() != 0 or window.ai_call_summary_table.rowCount() != 0:
            raise AssertionError("AI 调用日志清空失败")
        task = ai_parse_task("按自然语言生成采集任务 actions", page_snapshot_from_html("https://example.com/item", html), settings)
        if not task.get("actions"):
            raise AssertionError("AI Agent 动作规划失败")
        window.latest_ai_result = task
        window.apply_ai_task(task)
        if window.ai_task_plan_table.rowCount() < 3 or "AI 自检模板" not in window.ai_task_plan_label.text():
            raise AssertionError("自然语言任务计划未进入预览表")
        plan_text = "\n".join(
            window.ai_task_plan_table.item(row, column).text()
            for row in range(window.ai_task_plan_table.rowCount())
            for column in range(window.ai_task_plan_table.columnCount())
            if window.ai_task_plan_table.item(row, column)
        )
        if "extract" not in plan_text or "标题" not in plan_text:
            raise AssertionError("自然语言任务计划预览缺少动作或字段")
        window.copy_current_ai_task_plan()
        copied_plan_text = QApplication.clipboard().text() or getattr(window, "last_clipboard_text", "")
        if "AI 自检模板" not in copied_plan_text or "actions" not in copied_plan_text:
            raise AssertionError("自然语言任务计划复制失败")
        if not window.apply_current_ai_task_plan():
            raise AssertionError("自然语言任务计划未应用成功")
        if window.template_combo.currentText() != "AI 自检模板" or window.page_limit_input.value() != 2:
            raise AssertionError("自然语言任务计划未写入模板或采集参数")
        if not window.expert_mode_enabled:
            raise AssertionError("自然语言任务计划应用后应切换到专家模式")
        if window.tabs.tabText(window.tabs.currentIndex()) != "批量采集":
            raise AssertionError("自然语言任务计划应用后应定位到批量采集页")
        window.apply_ai_fields(suggested)
        if window.ai_suggest_table.rowCount() < 2:
            raise AssertionError("AI 建议列未进入确认表格")
        second_enable = window.ai_suggest_table.item(1, 0)
        second_enable.setCheckState(Qt.CheckState.Unchecked)
        window.apply_checked_ai_fields_to_template()
        if window.field_table.rowCount() != 1:
            raise AssertionError("AI 建议列确认应用失败")
        if window.field_table.item(0, 0).text() != "标题":
            raise AssertionError("AI 建议列编辑结果未写入模板 UI")
        window.show_ai_suggested_fields(suggested.get("fields", []))
        preview_record = UniversalExtractor(
            SiteTemplate(
                "AI 预采自检模板",
                field_rules=window.suggested_field_rules_from_table(),
            )
        ).extract(html, "https://example.com/item")
        window.show_preview_record(preview_record, window.suggested_field_rules_from_table())
        if window.ai_table.columnCount() < 3 or window.ai_table.rowCount() != 1:
            raise AssertionError("AI 预采表格未生成")
        if "测试商品 A" not in (window.ai_table.item(0, 1).text() if window.ai_table.item(0, 1) else ""):
            raise AssertionError("AI 预采字段值错误")
        if window.ai_quality_table.rowCount() < 2:
            raise AssertionError("AI 预采质量检测未生成")
        if window.ai_quality_table.columnCount() != 6 or "字段质量评分" not in window.ai_quality_score_label.text():
            raise AssertionError("AI 字段质量评分未显示")
        table_columns, table_rows = window.ai_table_data()
        if "标题" not in table_columns or not table_rows:
            raise AssertionError("AI 表格数据读取失败")
        ai_export_xlsx = os.path.join(data_dir, "ai_table_export.xlsx")
        ai_export_csv = os.path.join(data_dir, "ai_table_export.csv")
        ai_export_json = os.path.join(data_dir, "ai_table_export.json")
        export_table_data(ai_export_xlsx, table_columns, table_rows, sheet_name="AI表格结果")
        export_table_data(ai_export_csv, table_columns, table_rows, sheet_name="AI表格结果")
        export_table_data(ai_export_json, table_columns, table_rows, sheet_name="AI表格结果")
        for exported_path in (ai_export_xlsx, ai_export_csv, ai_export_json):
            if not os.path.exists(exported_path) or os.path.getsize(exported_path) <= 0:
                raise AssertionError(f"AI 表格导出失败：{exported_path}")
        from openpyxl import load_workbook

        ai_workbook = load_workbook(ai_export_xlsx)
        if ai_workbook.sheetnames[:3] != ["AI表格结果", "字段说明", "导出摘要"]:
            raise AssertionError("AI 表格 Excel 未生成工作簿结构")
        if ai_workbook["AI表格结果"].freeze_panes != "A2" or not ai_workbook["AI表格结果"].auto_filter.ref:
            raise AssertionError("AI 表格 Excel 未冻结表头或启用筛选")
        if ai_workbook["导出摘要"].cell(5, 2).value != len(table_rows):
            raise AssertionError("AI 表格 Excel 摘要行数错误")
        with open(ai_export_json, "r", encoding="utf-8") as f:
            exported_json = json.load(f)
        if not exported_json.get("rows") or "测试商品 A" not in json.dumps(exported_json, ensure_ascii=False):
            raise AssertionError("AI 表格 JSON 导出内容错误")
        window.fill_ai_table(["字段A", "字段B"], [["含\t制表符", "含\n换行"]])
        window.copy_ai_table_to_clipboard()
        clipboard_text = QApplication.clipboard().text() or getattr(window, "last_clipboard_text", "")
        if "含 制表符" not in clipboard_text or "含 换行" not in clipboard_text or "\t制表符" in clipboard_text:
            raise AssertionError("AI 表格复制到剪贴板失败")
        window.fill_ai_table(table_columns, table_rows)
        quality_issues = window.analyze_preview_quality(
            [
                FieldRule("空字段", ".not-exists"),
                FieldRule("重复一", "h1"),
                FieldRule("重复二", "h1"),
            ],
            {"空字段": "", "重复一": "同一个值", "重复二": "同一个值"},
        )
        problems = [issue.get("problem", "") for issue in quality_issues]
        if not any("空值" in problem for problem in problems):
            raise AssertionError("AI 质量检测未识别空值列")
        if not any("重复" in problem for problem in problems):
            raise AssertionError("AI 质量检测未识别重复列")
        summary = window.quality_summary(quality_issues)
        if summary.get("score", 100) >= 80 or summary.get("need_fix") != 1 or summary.get("need_confirm") < 1:
            raise AssertionError("AI 字段质量评分未识别低质量字段")
        window.fill_quality_table(quality_issues)
        if "需要修复" not in window.ai_quality_score_label.text():
            raise AssertionError("AI 字段质量评分未提示修复")
        if int(window.ai_quality_table.item(0, 1).text()) >= 50:
            raise AssertionError("AI 字段质量评分未给空字段低分")
        repaired = ai_repair_fields(
            "https://example.com/item",
            html,
            [FieldRule("标题", "h1"), FieldRule("价格", ".missing")],
            quality_issues,
            "修复字段",
            settings,
        )
        if not repaired.get("fields") or repaired["fields"][0].get("selector") != "h1#main-title":
            raise AssertionError("AI 问题列修复失败")
        window.apply_repaired_fields(repaired)
        if window.ai_suggest_table.item(0, 2).text() != "h1#main-title":
            raise AssertionError("AI 修复字段未回填建议表")
        if not window.apply_repaired_fields_to_template() or window.field_table.item(0, 1).text() != "h1#main-title":
            raise AssertionError("AI 修复字段未一键应用到模板")
        self_test_stage("ai repair fields OK")
        window.latest_preview_url = "https://example.com/item"
        window.latest_preview_html = html
        window.latest_preview_rules = [FieldRule("标题", ".missing-title")]
        window.fill_quality_table(
            window.analyze_preview_quality(window.latest_preview_rules, {"标题": ""})
        )
        window.auto_apply_repair_after_ai = True
        window.apply_repaired_fields(
            {
                "fields": [
                    {
                        "name": "标题",
                        "selector": "h1#main-title",
                        "attr": "text",
                        "multiple": False,
                        "reason": "自动修复标题选择器",
                    }
                ]
            }
        )
        if window.auto_apply_repair_after_ai:
            raise AssertionError("AI 修复后未清理自动应用状态")
        if window.field_table.item(0, 1).text() != "h1#main-title":
            raise AssertionError("AI 修复未自动应用到模板")
        if "需要修复" in window.ai_quality_score_label.text():
            raise AssertionError("AI 修复自动重评分仍提示需要修复")
        if not window.ai_table.item(0, 1) or "测试商品 A" not in window.ai_table.item(0, 1).text():
            raise AssertionError("AI 修复后未自动重新预采")
    finally:
        server.shutdown()
        server.server_close()
    class FakeSiteHandler(BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            return

        def do_GET(self):
            request_path = self.path.split("?", 1)[0]
            if request_path == "/image.png":
                image_data = (
                    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
                    b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
                    b"\x00\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe\x02\xfeA\xe2=\x9b"
                    b"\x00\x00\x00\x00IEND\xaeB`\x82"
                )
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.send_header("Content-Length", str(len(image_data)))
                self.end_headers()
                self.wfile.write(image_data)
                return
            pages = {
                "/": """
                <html><body>
                  <h1>列表页</h1>
                  <a href="/product/10001">测试商品详情</a>
                  <a href="/article/20002">测试文章详情</a>
                  <a href="/help">帮助中心</a>
                  <a class="next" href="/page2">下一页</a>
                  <a class="more" href="/feed">加载更多</a>
                  <a href="https://outside.example.com/product/999">站外商品</a>
                </body></html>
                """,
                "/page2": """
                <html><body>
                  <h1>第二页</h1>
                  <a href="/product/10002">第二页商品详情</a>
                </body></html>
                """,
                "/product/10002": """
                <html><body>
                  <h1>第二页商品</h1>
                  <p>第二页详情资料，包含库存、规格和发货说明。</p>
                  <img src="/product-10002.png" alt="第二页商品图">
                </body></html>
                """,
                "/feed": """
                <html><body>
                  <h1>更多列表</h1>
                  <p>模拟滚动后加载出的更多商品列表。</p>
                  <a href="/product/10003">滚动商品详情</a>
                </body></html>
                """,
                "/product/10003": """
                <html><body>
                  <h1>滚动商品</h1>
                  <p>滚动加载后进入的商品详情，包含颜色、尺码、库存和图片资料。</p>
                  <img src="/product-10003.png" alt="滚动商品图">
                  <table><tr><th>尺码</th><td>M/L/XL</td></tr><tr><th>库存</th><td>18</td></tr></table>
                </body></html>
                """,
                "/product/10001": """
                <html>
                  <head>
                    <meta name="description" content="结构化详情页描述，包含售后、库存和规格资料">
                    <meta property="og:image" content="/image.png">
                    <script type="application/ld+json">
                    {
                      "@context": "https://schema.org",
                      "@type": "Product",
                      "name": "结构化测试商品",
                      "description": "JSON-LD 中的商品描述和卖点",
                      "sku": "SKU-10001",
                      "brand": {"@type": "Brand", "name": "测试品牌"},
                      "image": ["/image.png", "/image-large.png"],
                      "offers": {
                        "@type": "Offer",
                        "price": "199.00",
                        "priceCurrency": "CNY",
                        "availability": "https://schema.org/InStock"
                      }
                    }
                    </script>
                  </head>
                  <body>
                    <h1>选中详情页</h1>
                    <article>
                      <p>这是一段选中详情页正文。</p>
                      <p>包含材质、尺寸、库存、售后和发货周期等完整说明。</p>
                    </article>
                    <img data-src="/image.png" alt="懒加载商品图">
                    <img srcset="/image-small.png 1x, /image-large.png 2x" alt="高清商品图">
                    <dl class="specs">
                      <dt>颜色</dt><dd>深海蓝</dd>
                      <dt>库存</dt><dd>现货 42 件</dd>
                    </dl>
                    <ul class="params">
                      <li>材质：航空铝</li>
                      <li>包装：礼盒装</li>
                    </ul>
                    <a href="/product/10001/reviews" aria-label="用户评价">评价</a>
                  </body>
                </html>
                """,
                "/article/20002": "<html><body><h1>未选文章页</h1><p>不应该被深抓。</p></body></html>",
                "/help": "<html><body><h1>帮助中心</h1></body></html>",
            }
            html_page = pages.get(request_path, "<html><body><h1>404</h1></body></html>")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(html_page.encode("utf-8"))

    site_server = HTTPServer(("127.0.0.1", 0), FakeSiteHandler)
    site_thread = threading.Thread(target=site_server.serve_forever, daemon=True)
    site_thread.start()
    try:
        site_base = f"http://127.0.0.1:{site_server.server_address[1]}/"
        discovery_logs = []
        collector = UniversalCollector(logger=lambda message: discovery_logs.append(message))
        candidates = collector.scan_subpage_links(site_base, use_browser=False, limit=20)
        if not candidates or not any("product/10001" in item.get("url", "") for item in candidates):
            raise AssertionError("子页面链接扫描未发现详情候选")
        if not any(item.get("selected") and "product/10001" in item.get("url", "") for item in candidates):
            raise AssertionError("详情页候选未默认选中")
        tricky_links = [
            {"url": "/#/product/999", "text": "商品详情"},
            {"url": "/en", "text": "English"},
            {"url": "/category/product", "text": "商品分类"},
            {"url": "/search?q=phone", "text": "搜索手机"},
            {"url": "/product/10001?sort=price", "text": "排序"},
            {"url": "/product/10002", "text": "商品详情"},
        ]
        tricky_candidates = collector.rank_subpage_links(tricky_links, site_base, limit=20)
        selected_tricky_urls = [item.get("url", "") for item in tricky_candidates if item.get("selected")]
        if not any("/product/10002" in item for item in selected_tricky_urls):
            raise AssertionError(f"强详情链接未被自动选中：{tricky_candidates}")
        for rejected_fragment in ("/en", "/category/product", "/search", "sort=price"):
            if any(rejected_fragment in item for item in selected_tricky_urls):
                raise AssertionError(f"导航/语言/筛选链接被错误自动选中：{selected_tricky_urls}")
        if any(item.get("type") == "SPA 路由" and item.get("selected") for item in tricky_candidates):
            raise AssertionError(f"SPA 锚点路由不应默认深抓：{tricky_candidates}")
        window.show_subpage_link_candidates(candidates)
        if window.subpage_link_table.rowCount() < 3:
            raise AssertionError("子页面候选表格未显示扫描结果")
        for row in range(window.subpage_link_table.rowCount()):
            url_item = window.subpage_link_table.item(row, 3)
            checked_item = window.subpage_link_table.item(row, 0)
            if url_item and "product/10001" in url_item.text():
                checked_item.setCheckState(Qt.CheckState.Checked)
            elif checked_item:
                checked_item.setCheckState(Qt.CheckState.Unchecked)
        window.apply_selected_subpage_links()
        if len(window.selected_subpage_urls) != 1 or "product/10001" not in window.selected_subpage_urls[0]:
            raise AssertionError("选中子页面未写入采集设置")
        selected_results = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=1,
            delay_seconds=0,
            scrape_subpages=True,
            subpage_limit=10,
            selected_subpage_urls=window.selected_subpage_urls,
            skip_unchanged=False,
        )
        selected_urls = [item.get("url", "") for item in selected_results]
        if len(selected_results) != 2 or not any("product/10001" in item for item in selected_urls):
            raise AssertionError("手动选择的子页面未被深抓")
        if any("article/20002" in item for item in selected_urls):
            raise AssertionError("未勾选子页面被错误深抓")
        selected_detail = next((item for item in selected_results if "product/10001" in item.get("url", "")), {})
        detail_body = selected_detail.get("body", "")
        detail_images = selected_detail.get("images", [])
        detail_tables = selected_detail.get("tables", [])
        if selected_detail.get("title") != "结构化测试商品":
            raise AssertionError("详情页未优先使用结构化标题")
        if "199.00" not in selected_detail.get("price", "") or "CNY" not in selected_detail.get("price", ""):
            raise AssertionError("详情页未抓取结构化价格")
        if int(selected_detail.get("completeness_score") or 0) < 85 or "完整" not in selected_detail.get("completeness_label", ""):
            raise AssertionError(f"详情页资料完整度评分过低：{selected_detail.get('completeness_label')}")
        for expected_text in ("SKU-10001", "测试品牌", "深海蓝", "航空铝", "结构化详情页描述"):
            if expected_text not in detail_body:
                raise AssertionError(f"详情页富资料未进入正文：{expected_text}")
        if not any("image-large.png" in image.get("url", "") for image in detail_images):
            raise AssertionError("详情页未抓取 srcset 高清图片")
        if not any("image.png" in image.get("url", "") for image in detail_images):
            raise AssertionError("详情页未抓取 meta/data-src 图片")
        flat_table_text = json.dumps(detail_tables, ensure_ascii=False)
        if "库存" not in flat_table_text or "礼盒装" not in flat_table_text:
            raise AssertionError("详情页规格参数未进入表格资料")
        progress_events = []
        progress_results = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=1,
            delay_seconds=0,
            scrape_subpages=True,
            subpage_limit=1,
            selected_subpage_urls=[],
            skip_unchanged=False,
            progress_callback=lambda event: progress_events.append(event),
        )
        progress_stages = [event.get("stage") for event in progress_events]
        if not progress_results or "采集页面" not in progress_stages or "页面完成" not in progress_stages:
            raise AssertionError("采集内核未实时回调页面进度")
        if not any(event.get("processed", 0) >= 1 for event in progress_events):
            raise AssertionError("采集内核进度回调未更新处理数量")
        firecrawl_collect_logs = []
        firecrawl_collect_events = []
        firecrawl_collect_progress = []
        firecrawl_collector = UniversalCollector(logger=lambda message: firecrawl_collect_logs.append(message))
        firecrawl_collector.expand_urls_with_firecrawl_search = lambda source_urls, firecrawl_client, firecrawl_config: (
            firecrawl_collect_events.append({"search_query": firecrawl_config.search_query, "search_limit": firecrawl_config.search_limit})
            or list(source_urls)
            + [site_base.rstrip("/") + "/search-extra"]
        )
        firecrawl_collector.expand_pages_with_firecrawl_map = lambda url, firecrawl_client, page_limit: [
            normalize_url(url),
            site_base.rstrip("/") + "/page2",
        ][:page_limit]

        def fake_collect_one_firecrawl(target_url, template, firecrawl_client):
            firecrawl_collect_events.append(
                {
                    "url": normalize_url(target_url),
                    "template": template.name,
                    "base_url": firecrawl_client.config.base_url,
                    "use_extract": firecrawl_client.config.use_extract,
                }
            )
            normalized_target = normalize_url(target_url)
            return {
                "collected_at": "2026-06-10 10:00:00",
                "url": normalized_target,
                "domain": "127.0.0.1",
                "template_name": f"{template.name} + Firecrawl",
                "title": "Firecrawl 采集结果",
                "price": "",
                "published_time": "",
                "author": "",
                "body": "Firecrawl 分支采集到的正文资料，用于验证 collect_urls 已接入远程增强采集。",
                "images": [],
                "links": [{"url": site_base.rstrip("/") + "/product/10001", "text": "详情"}],
                "tables": [],
                "fingerprint": f"firecrawl-{len(firecrawl_collect_events)}",
            }

        firecrawl_collector.collect_one_firecrawl = fake_collect_one_firecrawl
        firecrawl_collect_results = firecrawl_collector.collect_urls(
            [site_base],
            use_browser=True,
            page_limit=2,
            delay_seconds=0,
            scrape_subpages=False,
            skip_unchanged=False,
            firecrawl_config={
                "enabled": True,
                "api_key": "fc-collect-self-test",
                "base_url": "https://api.firecrawl.dev",
                "use_map": True,
                "map_limit": 2,
                "use_search": True,
                "search_query": "自检扩源",
                "search_limit": 1,
                "use_extract": True,
                "extract_prompt": "提取字段",
            },
            progress_callback=lambda event: firecrawl_collect_progress.append(event),
        )
        firecrawl_collect_stages = [event.get("stage") for event in firecrawl_collect_progress]
        if len(firecrawl_collect_results) < 3 or not any("/page2" in item.get("url", "") for item in firecrawl_collect_results) or not any("/search-extra" in item.get("url", "") for item in firecrawl_collect_results):
            raise AssertionError(f"Firecrawl collect_urls 分支未采集 Search/Map 扩展页：{firecrawl_collect_results}")
        if "Firecrawl 搜索" not in firecrawl_collect_stages or "Firecrawl 映射" not in firecrawl_collect_stages or "Firecrawl 采集" not in firecrawl_collect_stages:
            raise AssertionError(f"Firecrawl collect_urls 分支未回调专用阶段：{firecrawl_collect_stages}")
        if not any(item.get("search_query") == "自检扩源" for item in firecrawl_collect_events):
            raise AssertionError(f"Firecrawl collect_urls 未调用 Search 扩源：{firecrawl_collect_events}")
        if not any(item.get("base_url") == "https://api.firecrawl.dev" and item.get("use_extract") for item in firecrawl_collect_events):
            raise AssertionError(f"Firecrawl collect_urls 未传入客户端配置：{firecrawl_collect_events}")
        if not any("Firecrawl 增强已启用" in message for message in firecrawl_collect_logs):
            raise AssertionError(f"Firecrawl collect_urls 未写入启用日志：{firecrawl_collect_logs}")
        original_firecrawl_scrape = FirecrawlClient.scrape
        original_firecrawl_interact = FirecrawlClient.interact
        try:
            FirecrawlClient.scrape = lambda self, url: {
                "firecrawl_job_id": "interact-collect-job",
                "url": normalize_url(url),
                "markdown": "# Interact 页\n\n原始内容",
                "metadata": {"title": "Interact 页"},
            }
            FirecrawlClient.interact = lambda self, job_id, prompt=None, code=None: {
                "success": True,
                "output": f"{job_id}: 展开后的隐藏字段",
                "exit_code": 0,
            }
            interact_branch_results = UniversalCollector(logger=lambda message: None).collect_urls(
                [site_base],
                use_browser=False,
                page_limit=1,
                delay_seconds=0,
                skip_unchanged=False,
                firecrawl_config={
                    "enabled": True,
                    "api_key": "fc-interact-self-test",
                    "base_url": "https://api.firecrawl.dev",
                    "use_interact": True,
                    "interact_prompt": "点击展开更多",
                },
            )
            if not interact_branch_results or "隐藏字段" not in interact_branch_results[0].get("body", ""):
                raise AssertionError(f"Firecrawl Interact 分支未合并交互结果：{interact_branch_results}")
        finally:
            FirecrawlClient.scrape = original_firecrawl_scrape
            FirecrawlClient.interact = original_firecrawl_interact
        original_firecrawl_crawl = FirecrawlClient.crawl
        original_firecrawl_batch = FirecrawlClient.batch_scrape
        try:
            FirecrawlClient.crawl = lambda self, url: {
                "status": "completed",
                "completed": 2,
                "total": 2,
                "data": [
                    {
                        "url": site_base.rstrip("/") + "/crawl-a",
                        "markdown": "# Crawl A\n\nFirecrawl Crawl 深抓正文 A",
                        "metadata": {"title": "Crawl A"},
                    },
                    {
                        "url": site_base.rstrip("/") + "/crawl-b",
                        "markdown": "# Crawl B\n\nFirecrawl Crawl 深抓正文 B",
                        "metadata": {"title": "Crawl B"},
                    },
                ],
            }
            crawl_branch_logs = []
            crawl_branch_results = UniversalCollector(logger=lambda message: crawl_branch_logs.append(message)).collect_urls(
                [site_base],
                use_browser=True,
                page_limit=1,
                delay_seconds=0,
                skip_unchanged=False,
                firecrawl_config={
                    "enabled": True,
                    "api_key": "fc-crawl-self-test",
                    "base_url": "https://api.firecrawl.dev",
                    "use_crawl": True,
                    "crawl_limit": 2,
                    "crawl_max_depth": 2,
                },
            )
            if len(crawl_branch_results) != 2 or not any("/crawl-b" in item.get("url", "") for item in crawl_branch_results):
                raise AssertionError(f"Firecrawl Crawl 分支未保存深抓文档：{crawl_branch_results}")
            if not any("Firecrawl Crawl 完成" in message for message in crawl_branch_logs):
                raise AssertionError(f"Firecrawl Crawl 分支未写日志：{crawl_branch_logs}")
            FirecrawlClient.batch_scrape = lambda self, urls: {
                "status": "completed",
                "completed": len(urls),
                "total": len(urls),
                "data": [
                    {
                        "url": normalize_url(url),
                        "markdown": f"# Batch {index}\n\nFirecrawl Batch 批量正文 {index}",
                        "metadata": {"title": f"Batch {index}"},
                    }
                    for index, url in enumerate(urls, 1)
                ],
            }
            batch_branch_logs = []
            batch_branch_results = UniversalCollector(logger=lambda message: batch_branch_logs.append(message)).collect_urls(
                [site_base, site_base.rstrip("/") + "/page2"],
                use_browser=True,
                page_limit=1,
                delay_seconds=0,
                skip_unchanged=False,
                firecrawl_config={
                    "enabled": True,
                    "api_key": "fc-batch-self-test",
                    "base_url": "https://api.firecrawl.dev",
                    "use_batch": True,
                    "batch_max_concurrency": 3,
                },
            )
            if len(batch_branch_results) != 2 or not any("/page2" in item.get("url", "") for item in batch_branch_results):
                raise AssertionError(f"Firecrawl Batch 分支未保存批量文档：{batch_branch_results}")
            if not any("Firecrawl Batch 完成" in message for message in batch_branch_logs):
                raise AssertionError(f"Firecrawl Batch 分支未写日志：{batch_branch_logs}")
        finally:
            FirecrawlClient.crawl = original_firecrawl_crawl
            FirecrawlClient.batch_scrape = original_firecrawl_batch
        auto_page_results = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=2,
            delay_seconds=0,
            scrape_subpages=False,
            skip_unchanged=False,
        )
        auto_page_urls = [item.get("url", "") for item in auto_page_results]
        if not any("/page2" in item for item in auto_page_urls):
            raise AssertionError(f"自动翻页未在没有 CSS 选择器时发现第二页：{auto_page_urls}")
        ranked_auto_subpages = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=2,
            delay_seconds=0,
            scrape_subpages=True,
            subpage_limit=3,
            selected_subpage_urls=[],
            skip_unchanged=False,
        )
        ranked_urls = [item.get("url", "") for item in ranked_auto_subpages]
        if not any("product/10001" in item for item in ranked_urls) or not any("product/10002" in item for item in ranked_urls):
            raise AssertionError(f"自动深抓未覆盖分页里的商品详情：{ranked_urls}")
        if any("/help" in item for item in ranked_urls):
            raise AssertionError(f"自动深抓错误抓取了帮助/导航页：{ranked_urls}")
        ordinary_site_results = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=1,
            delay_seconds=0,
            scrape_subpages=False,
            skip_unchanged=False,
        )
        complete_site_results = collector.collect_urls(
            [site_base],
            use_browser=False,
            page_limit=3,
            delay_seconds=0,
            scrape_subpages=True,
            subpage_limit=6,
            selected_subpage_urls=[],
            skip_unchanged=False,
        )
        ordinary_urls = [item.get("url", "") for item in ordinary_site_results]
        complete_urls = [item.get("url", "") for item in complete_site_results]
        for expected_url in ("/page2", "/feed", "/product/10001", "/product/10002", "/product/10003"):
            if not any(expected_url in item for item in complete_urls):
                raise AssertionError(f"完整采集样例集未覆盖 {expected_url}：{complete_urls}")
        if len(complete_site_results) <= len(ordinary_site_results):
            raise AssertionError(f"完整采集没有比普通采集覆盖更多页面：普通={ordinary_urls} 完整={complete_urls}")
        ordinary_body_len = sum(len(item.get("body", "") or "") for item in ordinary_site_results)
        complete_body_len = sum(len(item.get("body", "") or "") for item in complete_site_results)
        ordinary_images = sum(len(item.get("images", []) or []) for item in ordinary_site_results)
        complete_images = sum(len(item.get("images", []) or []) for item in complete_site_results)
        complete_tables = sum(len(item.get("tables", []) or []) for item in complete_site_results)
        if complete_body_len <= ordinary_body_len or complete_images <= ordinary_images or complete_tables < 1:
            raise AssertionError(
                f"完整采集样例集未证明资料增加：body {ordinary_body_len}->{complete_body_len}, "
                f"images {ordinary_images}->{complete_images}, tables {complete_tables}"
            )
        joined_discovery_logs = "\n".join(discovery_logs)
        for expected_log in ("自动翻页候选", "自动发现", "子页面发现", "排除示例"):
            if expected_log not in joined_discovery_logs:
                raise AssertionError(f"采集发现日志缺少 {expected_log}：{joined_discovery_logs}")
        window.record_crawl_discovery_message("自动翻页候选：http://example.com 发现 2 个，示例=http://example.com/page2")
        window.record_crawl_discovery_message("子页面发现：候选 4 个，选中 2 个，选中=商品页:http://example.com/p/1，排除示例=导航页:http://example.com/help")
        if "自动翻页候选" not in window.simple_discovery_label.text() or "子页面发现" not in window.simple_discovery_label.text():
            raise AssertionError(f"普通首页未展示采集发现记录：{window.simple_discovery_label.text()}")
        browser_reuse_events = []
        reuse_collector = UniversalCollector(logger=lambda message: browser_reuse_events.append(("log", message)))

        def fake_open_browser_session(keep_login_state=False, headless=True):
            browser_reuse_events.append(("open", keep_login_state, headless))
            return {"context": "fake-context"}

        def fake_close_browser_session(session):
            browser_reuse_events.append(("close", session.get("context")))

        def fake_fetch_with_browser_session(session, url, scroll_times=0):
            browser_reuse_events.append(("fetch", session.get("context"), url))
            return f"<html><body><h1>复用页面</h1><p>{url}</p></body></html>"

        reuse_collector.open_browser_session = fake_open_browser_session
        reuse_collector.close_browser_session = fake_close_browser_session
        reuse_collector.fetch_with_browser_session = fake_fetch_with_browser_session
        reuse_results = reuse_collector.collect_urls(
            [site_base, site_base.rstrip("/") + "/page2"],
            use_browser=True,
            page_limit=1,
            delay_seconds=0,
            scrape_subpages=False,
            skip_unchanged=False,
        )
        if len(reuse_results) != 2:
            raise AssertionError("浏览器会话复用样例未采集两个 URL")
        if len([event for event in browser_reuse_events if event[0] == "open"]) != 1:
            raise AssertionError(f"批量采集不应每页重开浏览器：{browser_reuse_events}")
        if len([event for event in browser_reuse_events if event[0] == "close"]) != 1:
            raise AssertionError(f"批量采集应在结束时关闭一次浏览器会话：{browser_reuse_events}")
        if len([event for event in browser_reuse_events if event[0] == "fetch"]) != 2:
            raise AssertionError(f"两个 URL 应复用同一浏览器会话抓取：{browser_reuse_events}")
        fallback_logs = []
        fallback_collector = UniversalCollector(logger=lambda message: fallback_logs.append(message))
        fallback_collector.fetch_with_playwright = lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("browser unavailable"))
        fallback_results = fallback_collector.collect_urls(
            [site_base],
            use_browser=True,
            page_limit=1,
            delay_seconds=0,
            skip_unchanged=False,
        )
        if not fallback_results or fallback_results[0].get("title") != "列表页" or fallback_results[0].get("error"):
            raise AssertionError("浏览器采集失败后未自动退回普通网页读取")
        if not any("自动改用普通网页读取" in message for message in fallback_logs):
            raise AssertionError("浏览器采集降级未写入日志")
        pagination = collector.preview_pagination(
            site_base,
            next_page_selector="a.next",
            page_limit=2,
            scroll_times=2,
            keep_login_state=False,
        )
        if len(pagination.get("urls", [])) != 2 or not any("/page2" in item for item in pagination.get("urls", [])):
            raise AssertionError("分页预览未发现第二页")
        window.show_pagination_preview(pagination.get("rows", []))
        if window.pagination_table.rowCount() != 2:
            raise AssertionError("分页预览表格未显示翻页结果")
        window.ai_next_page_selector_input.setText("a.next")
        window.ai_page_limit_input.setValue(2)
        window.ai_scroll_times_input.setValue(2)
        window.apply_pagination_settings()
        if window.next_page_selector_input.text() != "a.next":
            raise AssertionError("下一页选择器未应用到模板设置")
        if window.page_limit_input.value() != 2 or window.scroll_times_input.value() != 2:
            raise AssertionError("分页/滚动参数未应用到高级采集")
        image_dir = os.path.join(data_dir, "downloaded_images")
        image_records = [
            {
                "url": site_base,
                "title": "图片来源页",
                "images": [{"url": site_base.rstrip("/") + "/image.png"}],
            }
        ]
        image_rows = download_images_from_records(image_records, image_dir, logger=lambda message: None)
        if len(image_rows) != 1 or image_rows[0].get("status") != "已保存" or not os.path.exists(image_rows[0].get("file_path", "")):
            raise AssertionError("图片下载结构化结果失败")
        window.records = image_records
        original_get_existing_directory = QFileDialog.getExistingDirectory
        original_message_information = QMessageBox.information
        try:
            QFileDialog.getExistingDirectory = staticmethod(lambda *args, **kwargs: image_dir)
            QMessageBox.information = staticmethod(lambda *args, **kwargs: None)
            window.download_current_images()
        finally:
            QFileDialog.getExistingDirectory = original_get_existing_directory
            QMessageBox.information = original_message_information
        image_columns, image_table_rows = window.ai_table_data()
        if image_columns[:3] != ["状态", "保存路径", "图片网址"] or not image_table_rows or image_table_rows[0][0] != "已保存":
            raise AssertionError("图片下载结果未进入结构化表格")
    finally:
        site_server.shutdown()
        site_server.server_close()
    self_test_stage("OK")
    if getattr(window, "schedule_tick_timer", None):
        window.schedule_tick_timer.stop()
    if getattr(window, "schedule_timer", None):
        window.schedule_timer.stop()
    window.close()
    app.processEvents()
    app.quit()
