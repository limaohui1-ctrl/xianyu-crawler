"""Export helpers for collector records and table data."""

import csv
import json
import re
import time


MAX_TEXT_LENGTH = 20000
FORMULA_PREFIXES = ("=", "+", "-", "@")

FIELD_HEADERS = [
    "采集时间",
    "网址",
    "域名",
    "模板",
    "标题",
    "价格",
    "时间",
    "作者",
    "正文",
    "图片",
    "链接",
    "表格",
    "内容指纹",
    "是否变化",
    "错误",
]

FIELD_DESCRIPTIONS = {
    "采集时间": "本条记录生成的本地时间。",
    "网址": "采集页面或子页面 URL。",
    "域名": "网址所属域名，便于筛选来源网站。",
    "模板": "采集时使用的网站模板。",
    "标题": "页面标题或模板抽取到的主标题。",
    "价格": "商品、房源或列表项价格字段。",
    "时间": "页面公开展示的发布时间、更新时间或相关时间。",
    "作者": "页面作者、商家、公司或发布人。",
    "正文": "页面主要正文或详情内容。",
    "图片": "抽取到的图片 URL，多个值用换行分隔。",
    "链接": "抽取到的链接，多个值用换行分隔。",
    "表格": "页面表格内容的文本化结果。",
    "内容指纹": "用于变更监控的内容摘要指纹。",
    "是否变化": "与上次同网址采集相比是否发生变化。",
    "错误": "本条记录采集失败或部分失败时的错误信息。",
}


def now_text():
    return time.strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value, limit=MAX_TEXT_LENGTH):
    if value is None:
        return ""
    text = re.sub(r"[\r\t]+", " ", str(value))
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ ]{2,}", " ", text)
    text = text.strip()
    if len(text) > limit:
        return text[:limit].rstrip() + "..."
    return text


def list_to_text(value, limit=32000):
    if isinstance(value, str):
        return clean_text(value, limit)
    if value is None:
        return ""
    return clean_text(json.dumps(value, ensure_ascii=False), limit)


def safe_export_cell(value, limit=32000):
    text = list_to_text(value, limit) if isinstance(value, (list, tuple, dict)) else clean_text(value, limit)
    if text and text[0] in FORMULA_PREFIXES:
        return "'" + text
    return text


def records_to_rows(records):
    rows = []
    for record in records:
        rows.append(
            [
                record.get("collected_at", ""),
                record.get("url", ""),
                record.get("domain", ""),
                record.get("template_name", ""),
                record.get("title", ""),
                record.get("price", ""),
                record.get("published_time", ""),
                record.get("author", ""),
                record.get("body", ""),
                list_to_text(record.get("images", [])),
                list_to_text(record.get("links", [])),
                list_to_text(record.get("tables", [])),
                record.get("fingerprint", ""),
                "是" if record.get("changed") else "否",
                record.get("error", ""),
            ]
        )
    return rows


def export_csv(file_path, records):
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(FIELD_HEADERS)
        writer.writerows([[safe_export_cell(value) for value in row] for row in records_to_rows(records)])


def records_to_tsv(records):
    lines = ["\t".join(FIELD_HEADERS)]
    for row in records_to_rows(records):
        values = []
        for value in row:
            text = safe_export_cell(value, 32000)
            text = str(text).replace("\r", " ").replace("\n", " ").replace("\t", " ")
            values.append(text)
        lines.append("\t".join(values))
    return "\n".join(lines)


def freeze_and_filter_sheet(sheet):
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def apply_excel_widths(sheet, max_width=60):
    from openpyxl.utils import get_column_letter

    for column_cells in sheet.columns:
        width = 10
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        sheet.column_dimensions[column_letter].width = width


def append_summary_sheet(workbook, records):
    sheet = workbook.create_sheet("导出摘要")
    records = list(records or [])
    domains = sorted({clean_text(record.get("domain", ""), 120) for record in records if record.get("domain")})
    templates = sorted({clean_text(record.get("template_name", ""), 120) for record in records if record.get("template_name")})
    rows = [
        ["项目", "值"],
        ["导出时间", now_text()],
        ["记录数", len(records)],
        ["域名数", len(domains)],
        ["模板数", len(templates)],
        ["域名", "\n".join(domains)],
        ["模板", "\n".join(templates)],
    ]
    for row in rows:
        sheet.append(row)
    apply_excel_widths(sheet)
    freeze_and_filter_sheet(sheet)


def append_field_description_sheet(workbook):
    sheet = workbook.create_sheet("字段说明")
    sheet.append(["字段", "说明"])
    for header in FIELD_HEADERS:
        sheet.append([header, FIELD_DESCRIPTIONS.get(header, "")])
    apply_excel_widths(sheet)
    freeze_and_filter_sheet(sheet)


def export_xlsx(file_path, records):
    from openpyxl import Workbook

    records = list(records or [])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采集结果"
    sheet.append(FIELD_HEADERS)
    for row in records_to_rows(records):
        sheet.append([safe_export_cell(value) for value in row])
    apply_excel_widths(sheet)
    freeze_and_filter_sheet(sheet)
    append_field_description_sheet(workbook)
    append_summary_sheet(workbook, records)
    workbook.save(file_path)


def export_records(file_path, records):
    if file_path.lower().endswith(".xlsx"):
        return export_xlsx(file_path, records)
    return export_csv(file_path, records)


def normalize_table_rows(rows):
    normalized = []
    for source_row in rows or []:
        if isinstance(source_row, list):
            normalized.append(source_row)
        elif isinstance(source_row, tuple):
            normalized.append(list(source_row))
        elif isinstance(source_row, dict):
            normalized.append(list(source_row.values()))
        else:
            normalized.append([source_row])
    return normalized


def table_data_to_tsv(columns, rows):
    columns = [str(item) for item in (columns or [])]
    rows = normalize_table_rows(rows)
    lines = []
    if columns:
        lines.append("\t".join(safe_export_cell(value, 32000).replace("\t", " ").replace("\n", " ") for value in columns))
    for row in rows:
        lines.append(
            "\t".join(
                safe_export_cell(value, 32000).replace("\t", " ").replace("\n", " ")
                for value in row
            )
        )
    return "\n".join(lines)


def append_table_summary_sheet(workbook, columns, rows, sheet_name="表格"):
    sheet = workbook.create_sheet("导出摘要")
    rows = list(rows or [])
    summary_rows = [
        ["项目", "值"],
        ["导出时间", now_text()],
        ["表格名称", sheet_name],
        ["列数", len(columns or [])],
        ["行数", len(rows)],
    ]
    for row in summary_rows:
        sheet.append(row)
    apply_excel_widths(sheet)
    freeze_and_filter_sheet(sheet)


def append_table_field_sheet(workbook, columns):
    sheet = workbook.create_sheet("字段说明")
    sheet.append(["字段", "说明"])
    for column in columns or []:
        sheet.append([column, "AI 或文件识别得到的结构化字段。"])
    apply_excel_widths(sheet)
    freeze_and_filter_sheet(sheet)


def export_table_data(file_path, columns, rows, sheet_name="表格"):
    columns = [str(item) for item in (columns or [])]
    rows = normalize_table_rows(rows)
    if file_path.lower().endswith(".json"):
        payload_rows = []
        for row in rows:
            payload_rows.append(
                {
                    columns[index] if index < len(columns) else f"字段{index + 1}": row[index]
                    for index in range(max(len(columns), len(row)))
                    if index < len(row)
                }
            )
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump({"columns": columns, "rows": payload_rows}, f, ensure_ascii=False, indent=2)
        return file_path
    if file_path.lower().endswith(".xlsx"):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = clean_text(sheet_name or "表格", 30) or "表格"
        if columns:
            sheet.append([safe_export_cell(value) for value in columns])
        for row in rows:
            sheet.append([safe_export_cell(value) for value in row])
        apply_excel_widths(sheet)
        freeze_and_filter_sheet(sheet)
        append_table_field_sheet(workbook, columns)
        append_table_summary_sheet(workbook, columns, rows, sheet.title)
        workbook.save(file_path)
        return file_path
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        if columns:
            writer.writerow([safe_export_cell(value) for value in columns])
        writer.writerows([[safe_export_cell(value) for value in row] for row in rows])
    return file_path
